# GyG RESUMENES DE PAGO
#
# Genera los resúmenes de pagos recibidos de las familias seleccionadas en el
# período indicado
#

"""
    PENDIENTE POR HACER
    -   Fijar como celda superior izquierda de la hoja de cálculo 'A solicitud' unas 5 filas
        antes del último grupo generado
    -   

    HISTORICO
    -   Se cambiaron las ubicaciones de los archivos resultantes a la carpeta GyG Recibos dentro de
        la carpeta actual para compatibilidad entre Windows y macOS (21/10/2019)
    -   CORREGIR: Ajustar función "envia_por_correo()" para manejar múltiples destinatarios
        de un mismo correo (13/10/2019)
    -   Cambiar el manejo de cuotas para usar las rutinas en la clase Cuota (GyG_cuotas)
        (11/10/2019)
    -   Ajustar el resumen_de_cuotas() para indicar que, a partir de Septiembre 2019 se maneja una
        cuota mensual de 1 dólar, actualizada semanalmente. (04/09/2019)
    -   CORREGIR: Las cuotas mostradas para los colegios y otros inmuebles especiales son las
        mismas que la del resto de los vecinos, por lo que los montos desplegados no son correctos
        (10/06/2019)
    -   Cambiar el texto del botón btAcepta a '', 'Genera resumen' o 'Genera resúmenes' si
        len(lbVecinos.curselection()) = {0, 1, >1}, respectivamente -- 12/03/2019
    -   Opción para desplegar o no el texto 'a solicitud de la parte interesada '
        (12/03/2019)
    -   Cambiar el cursor a HourGlass al iniciar las rutina ckHistorico_Seleccion() y
        genera_resumenes() y devolverlo al estándar al finalizar
    -   Generar resumen de saldos, tanto del archivo normal como del histórico
    -   Seleccionar la impresión de resúmenes de dos posibles fuentes: Histórico (en
        bolívares fuertes) o el archivo estándar (en bolívares soberanos)
    -   ¿Cómo manejar la fecha real de inicio de una familia, para utilizar ésta como
        fecha de inicio, si es posterior a la indicada? Igual para la fecha de fina-
        lización
        ^--> En la hoja "R.VIGILANCIA (reordenado)" están las columnas F.Desde y F.Hasta
             que pueden ser usadas para ello
        ^--> No es necesario, no se mostrarán pagos antes o después de las fechas en
             las cuales la familia canceló cuotas. La relación de cuotas mensuales
             aprobadas, no tiene que ver con la familia en sí
    -   Revisar formas más eficientes para seleccionar las fechas de inicio y final:
        ¿Combobox? ¿anticipar fechas posibles?
        ^--> Colocar como mes y año hasta el mes y año del día de hoy, para abarcar
             el período completo de pagos
    -   En el resumen de cuotas, ignorar los meses en los cuales no hay una cuota
        definida (NaN)
    -   Indicar la cantidad de vecinos seleccionados
    -   
    
"""

# Selecciona las librerías a utilizar
print('Cargando librerías...')
import tkinter as tk
from tkinter import ttk, font

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

from pandas import read_excel, isnull, notnull, DataFrame
from datetime import datetime
from dateutil.relativedelta import relativedelta

import yagmail
import credentials

import re
import sys
import os
import pickle
from locale import setlocale, LC_ALL

from reportlab.lib import colors, pdfencrypt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

import GyG_constantes
from GyG_cuotas import *
from GyG_utilitarios import *


# Define textos y constantes
pdf_file = GyG_constantes.pdf_resumen      # 'GyG Resumen {resumen:05d}.pdf'
pdf_path = GyG_constantes.ruta_resumenes   # 'C:/Users/MColosso/Google Drive/GyG Recibos/Resúmenes/'

excel_wb_historico = GyG_constantes.pagos_wb_historico   # '1.1. GyG Recibos - SOLO CONSULTA.xlsm'
excel_wb_estandar  = GyG_constantes.pagos_wb_estandar    # '1.1. GyG Recibos.xlsm'

consulta_historico = False
despliega_saldos   = False
cambio_archivo_origen = False

cuotas_en_lista    = True   # Elige entre mostrar las cuotas en una lista o en un único párrafo

excel_workbook     = excel_wb_estandar
excel_vigilancia   = GyG_constantes.pagos_ws_vigilancia   # 'Vigilancia'
excel_resumen      = GyG_constantes.pagos_ws_resumen      # 'RESUMEN VIGILANCIA'
                                                          # (anteriormente: 'R.VIGILANCIA (reordenado)')
excel_saldos       = GyG_constantes.pagos_ws_saldos       # 'Saldos'
resumen_workbook   = GyG_constantes.resumen_workbook      # '4.2. GyG Resúmenes - Control resúmenes.xlsx'
resumen_worksheet  = GyG_constantes.resumen_worksheet     # 'A solicitud'
excel_cuotas       = GyG_constantes.pagos_ws_cuotas       # 'CUOTA'

nombreMeses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
               'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

normal_cursor = ''
busy_cursor   = '@C:/Windows/Cursors/aero_busy.ani'
#busy_cursor   = ('@' if sys.platform.startswith('win') else '') + \
#                os.path.join(GyG_constantes.rec_imágenes, '3D_red_busy.cur.ani')    #'aero_busy.ani')
#busy_cursor   = "@recursos/imágenes/3D_red_busy.cur.ani"  #'aero_busy.ani'
                # Normal: @./imagenes/aero_arrow.cur        # aero normal.cur   # Normal.cur
                # Busy:               3D_red_busy.cur.ani   # aero busy.ani     # Busy.ani

ini_file_path = 'resumenes_de_pago.ini'

geometria = dict()

EOL = '<br/>'

dummy = setlocale(LC_ALL, 'es_es')

ahora = datetime.now()
am_pm = 'am' if ahora.hour < 12 else 'm' if ahora.hour < 13 else 'pm'
fecha_actual = '{:%d/%m/%Y %I:%M} {}'.format(ahora, am_pm)

envia_correos = False   # variable relacionada con el checkbutton ckCorreo

# Opciones a manejar: <check>, <warn>, <error>, <info>, <pdf>, <mail>, <whatsapp>
imgCheck = None   # tk.PhotoImage()
imgWarn  = None   # tk.PhotoImage()
imgError = None   # tk.PhotoImage()
imgInfo  = None   # tk.PhotoImage()
imgStop  = None   # tk.PhotoImage()
imgPdf   = None   # tk.PhotoImage()
imgEmail = None   # tk.PhotoImage()
stDelimiters = ('<check>', '<warn>', '<error>', '<info>', '<stop>', '<pdf>', '<email>')
stImages     = (imgCheck,  imgWarn,  imgError,  imgInfo,  imgStop,  imgPdf,  imgEmail )
stPattern    = '|'.join(['(' + r + ')' for r in map(re.escape, stDelimiters)])


def valida_parametros():
    global seleccion_vecinos, dt_inicial, dt_final

    actualiza_status(txStatus, '\n\nValidando opciones: ')

    seleccion_vecinos = [lbVecinos.get(idx) for idx in lbVecinos.curselection()]
    valida_vecinos = len(seleccion_vecinos) >= 1
    if not valida_vecinos:
        actualiza_status(txStatus, '\n- <error> No se seleccionó vecino alguno para gererar un resumen')

    dt_inicial = datetime(int(spDesdeAño.get()), nombreMeses.index(spDesdeMes.get()) + 1, 1)
    dt_final   = datetime(int(spHastaAño.get()), nombreMeses.index(spHastaMes.get()) + 1, 1)
    valida_fechas = dt_inicial <= dt_final
    if not valida_fechas:
        actualiza_status(txStatus, f'\n- <error> La fecha inicial ({dt_inicial:%B/%Y}) es posterior ' + \
                                   f'a la fecha final ({dt_final:%B/%Y})')

    if valida_vecinos and valida_fechas:
        actualiza_status(txStatus, '<check>')

    if valida_vecinos and valida_fechas:
        set_cursor_wait()
        genera_resumenes()
        set_cursor_standard()
    else:
        actualiza_status(txStatus, '\n\n<stop> Proceso terminado: 0 resumenes generados')
        if envia_correos:
            actualiza_status(txStatus, ', 0 enviados')


def genera_resumenes():
    global hoja_resumen, ultimo_resumen
    global generados, enviados

    def último_correo_o_celular(beneficiario):
        df_test = df[df['Beneficiario'] == beneficiario]
        return df_test.iloc[-1]['E-mail o celular']

    def genera_y_envia_resumen():
        global generados, enviados   # enviado, 

        if beneficiario_anterior != '':
            actualiza_status(txStatus, f"\n - {beneficiario_anterior}:")
            enviado = False
            if genera_pdf(r_anterior, resumen, idx + ultimo_resumen):
                generados += 1
                enviado = envia_por_correo(r_anterior,
                                           decode_email_o_celular(último_correo_o_celular(r_anterior['Beneficiario'])),
                                           idx + ultimo_resumen)
                if enviado:
                    enviados += 1
                    actualiza_status(txStatus, ', <email>: <check>')
            linea_de_resumen(idx + ultimo_resumen,
                             idx + row_offset,
                             beneficiario_anterior,
                             último_correo_o_celular(beneficiario_anterior),
                             r_anterior['Dirección'],
                             enviado)

    print('Generando resúmenes...')

    if cambio_archivo_origen:
        carga_pagos_de_vigilancia()
        carga_saldos_de_cuentas()

    actualiza_status(txStatus, '\nGenerando resúmenes:')

    df = df_saldos.copy() if despliega_saldos else df_pagos.copy()

    # Elimina aquellos registros de pago que no fueron seleccionados
    df.drop(df.index[(df['Fecha'] < dt_inicial) | (df['Fecha'] >= dt_final + relativedelta(months=1))], inplace=True)
    df.drop(df.index[~df['Beneficiario'].isin(seleccion_vecinos)], inplace=True)

    # Ordena registros por beneficiario y fecha
    df.sort_values(by=['Beneficiario', 'Fecha'], inplace=True)

    try:
        wb_resumen = load_workbook(filename=resumen_workbook)
    except:
        # Crea la hoja resumen (si no existe)
        actualiza_status(txStatus, f'\n- Generando la hoja de cálculo "{resumen_workbook}"...')

        wb_resumen = Workbook()
        hoja_resumen = wb_resumen.active
        define_atributos(hoja_resumen)

    if resumen_worksheet in wb_resumen.sheetnames:
        # open worksheet
        hoja_resumen = wb_resumen[resumen_worksheet]
        ultimo_resumen = hoja_resumen[f'A{hoja_resumen.max_row}'].value
        row_offset = hoja_resumen.max_row
    else:
        # create worksheet
        hoja_resumen = wb_resumen.create_sheet()  # insert at the end (default)
        define_atributos(hoja_resumen)
        ultimo_resumen = 0
        row_offset = 1

    if isnull(ultimo_resumen):
        row = last_row = hoja_resumen.max_row
        while isnull(hoja_resumen[f'A{row}'].value):
            row -=1
        ultimo_resumen = hoja_resumen[f'A{row}'].value
        row_offset = row
        actualiza_status(txStatus, f' <warn> {last_row - row} líneas en blanco al final de la hoja de cálculo')

    # Fija la primera línea como encabezado
    hoja_resumen.freeze_panes = hoja_resumen['A2']

    # proceso
    beneficiario_anterior = ''
    r_anterior = None
    idx = 0
    generados = 0
    enviados = 0
    if despliega_saldos:
        for i in range(df.shape[0]):
            r = df.iloc[i]
            if r['Beneficiario'] != beneficiario_anterior:
                genera_y_envia_resumen()
                resumen = [['Fecha', 'Concepto', 'Monto', 'Saldo']]
                idx += 1
                beneficiario_anterior = r['Beneficiario']
                r_anterior = r
            resumen.append([r['Fecha'].strftime('%d/%m/%Y'),
                            r['Concepto'],
                            '' if isnull(r['Monto']) else '{:,.02f}'.format(r['Monto']),
                            '{:,.02f}'.format(r['Saldo'])])
    else:
        for i in range(df.shape[0]):
            r = df.iloc[i]
            if r['Beneficiario'] != beneficiario_anterior:
                genera_y_envia_resumen()
                resumen = [['Recibo', 'Fecha', 'Monto', 'Concepto']]
                idx += 1
                beneficiario_anterior = r['Beneficiario']
                r_anterior = r
            resumen.append(['{:05d}'.format(r['Nro. Recibo']),
                           r['Fecha'].strftime('%d/%m/%Y'),
                           r['Monto (ed)'],
                           r['Concepto']])

    genera_y_envia_resumen()
    linea_de_separacion(idx + row_offset)

    # Selecciona la última celda como la celda activa
    hoja_resumen.sheet_view.selection[0].activeCell = f'A{idx + row_offset}'
    hoja_resumen.sheet_view.selection[0].sqref      = f'A{idx + row_offset}'

# -- Prueba para tratar de fijar la celda superior izquierda de la hoja de control de resúmenes
#
#    celda_superior = idx + row_offset - 5
#    if celda_superior < 1:
#        celda_superior = 1
#    print(f'DEBUG: topLeftCell = A{celda_superior}\n       activeCell = A{idx + row_offset}')
#    hoja_resumen.SheetView(topLeftCell = f'A{celda_superior}')
##    hoja_resumen.views.topLeftCell = f'A{celda_superior}'
#
# ----------------------------------------------------------------------------------------------

    hoja_resumen.cell(column=1, row=idx+row_offset)
    
    try:
        wb_resumen.save(resumen_workbook)
        s  =  's' if generados != 1 else ''
        resumen_es = 'resúmenes' if generados != 1 else 'resumen'
        s2 =  's' if enviados != 1 else ''
        actualiza_status(txStatus, f'\n\n<info> Proceso terminado: {generados} {resumen_es} generado{s}')
        vecinos_seleccionados = len(lbVecinos.curselection())
        if generados != vecinos_seleccionados:
            diferencia = vecinos_seleccionados - generados
            s =  's' if diferencia != 1 else ''
            actualiza_status(txStatus, f', {diferencia} vecino{s} seleccionado{s} sin pagos en el período')
        if envia_correos:
            actualiza_status(txStatus, f', {enviados} enviado{s2}')
    except:
        error_msg  = str(sys.exc_info()[1])
        if sys.platform.startswith('win'):
            error_msg  = error_msg.replace('\\', '/')
        actualiza_status(txStatus, f'\n\n<stop> ERROR: {error_msg}')


def genera_pdf(r, tabla, idx):
    global rango_fechas

    ENCABEZADO_P = '<para alignment="center"><font size=15><b>A QUIEN PUEDA INTERESAR</b></font></para>'
    ENCABEZADO_S = '<font size=18><b>{Beneficiario}</b></font>'
    CUERPO  = '<font size=12>Resumen de {Tipo} de <b>{Beneficiario}, <i>{Direccion}</i></b>, ' + \
              '{Rango_fechas}:</font>'
#   FOOTER1  Se genera dinámicamente en el cuerpo, con las cuotas vigentes durante el período solicitado
    FOOTER2 = '<font size=12>Resumen emitido {Origen}el día ' + \
              '{Fecha:%d de %B de %Y}.</font>'.format(Fecha=ahora)
    EQUIPO_COBRANZA = '<font size=12><i>Equipo de Cobranza' + EOL + 'Junta Directiva GyG</i></font>'

    FONT_NAME = 'Times-Roman'

    # Security of .pdf files
    allow_print = pdfencrypt.StandardEncryption("",           # sin contraseña
                                                canPrint=1,
                                                canModify=0,
                                                canCopy=0,
                                                canAnnotate=0,
                                                strength=128)

    filename = os.path.join(pdf_path, pdf_file.format(resumen=idx))
    doc = SimpleDocTemplate(filename, pagesize=letter, encrypt=allow_print)

    # container for the 'Flowable' objects
    elements = []

    # Genera las líneas de encabezado
    styleSheet = getSampleStyleSheet()
    normalStyle = styleSheet['Normal']
    normalStyle.fontName = FONT_NAME
    normalStyle.alignment = 4                             # TA_JUSTIFY
    if despliega_saldos:
        elements.append(Paragraph(ENCABEZADO_S.format(Beneficiario=r['Beneficiario']), normalStyle))
    else:
        elements.append(Paragraph(ENCABEZADO_P, normalStyle))
    elements.append(Spacer(1, 24))
    if dt_inicial == dt_final:
        rango_fechas = 'em el mes de {:%B %Y}'.format(dt_final)
    else:
        rango_fechas = 'durante el período comprendido entre ' + \
                       ('{:%B} y {:%B %Y}' if dt_inicial.year == dt_final.year
                        else '{:%B %Y} y {:%B %Y}').format(dt_inicial, dt_final)
    ptext = CUERPO.format(Tipo=('movimientos' if despliega_saldos else 'pagos recibidos'),
                          Beneficiario=r['Beneficiario'],
                          Direccion=r['Dirección'],
                          Rango_fechas=rango_fechas)
    elements.append(Paragraph(ptext, normalStyle))
    elements.append(Spacer(1, 16))

    # Habilita el desbordamiento en la columna de Concepto, convirtiéndola en párrafo
    bodytextStyle = styleSheet['BodyText']
    bodytextStyle.fontName = FONT_NAME
    bodytextStyle.alignment = 4          # TA_JUSTIFY
    col_concepto = 1 if despliega_saldos else 3
    for i in range(1, len(tabla)):
        tabla[i][col_concepto] = Paragraph('<para wordwrap=CJK>' + tabla[i][col_concepto] + '</para>', bodytextStyle)
    
    # Anexa la tabla de valores
    colwidths = [60, 255, 70, 70] if despliega_saldos else [50, 60, 70, 275]
    t = Table(tabla, colWidths=colwidths)
    # Table coordinates are given as (column, row) which follows the spreadsheet 'A1' model,
    # but not the more natural (for mathematicians) 'RC' ordering. The top left cell is (0, 0)
    # the bottom right is (-1, -1)
    if despliega_saldos:
        t.setStyle(TableStyle([('VALIGN',     (0, 0), (-1,  0), 'TOP'),
                               ('TEXTCOLOR',  (0, 0), (-1,  0), colors.white),
                               ('BACKGROUND', (0, 0), (-1,  0), colors.blue),
                               ('ALIGN',      (0, 0), (-1,  0), 'CENTER'),
                               ('ALIGN',      (0, 1), ( 0, -1), 'CENTER'),
                               ('ALIGN',      (2, 1), ( 3, -1), 'RIGHT'),
                               ('VALIGN',     (0, 1), (-1, -1), 'TOP'),
                               ('FONTNAME',   (0, 0), (-1, -1), FONT_NAME)]))
    else:
        t.setStyle(TableStyle([('VALIGN',     (0, 0), (-1,  0), 'TOP'),
                               ('TEXTCOLOR',  (0, 0), (-1,  0), colors.white),
                               ('BACKGROUND', (0, 0), (-1,  0), colors.blue),
                               ('ALIGN',      (0, 0), (-1,  0), 'CENTER'),
                               ('ALIGN',      (0, 1), ( 1, -1), 'CENTER'),
                               ('ALIGN',      (2, 1), ( 2, -1), 'RIGHT'),
                               ('VALIGN',     (0, 1), (-1, -1), 'TOP'),
                               ('FONTNAME',   (0, 0), (-1, -1), FONT_NAME)]))
    elements.append(t)

    FOOTER1 = '<font size=12>' + \
              cuotas_obj.resumen_de_cuotas(r['Beneficiario'], 
                                           fecha_inicial=dt_inicial, fecha_final=dt_final, 
                                           formato=fmtPdf, lista=True) + \
              EOL + '</font>'
    # Genera la firma
    if ckCuotas.get() > 0:
        elements.append(Spacer(1, 16))
        elements.append(Paragraph(FOOTER1, normalStyle))

    elements.append(Spacer(1, 16))
    fue_solicitado = ckSolicitado.get() > 0
#    origen_de_la_solicitud = '' if despliega_saldos else 'a solicitud de la parte interesada '
    origen_de_la_solicitud = 'a solicitud de la parte interesada ' if not despliega_saldos and fue_solicitado else ''
    elements.append(Paragraph(FOOTER2.format(Origen=origen_de_la_solicitud), normalStyle))
    elements.append(Spacer(1, 24))
    elements.append(Paragraph(EQUIPO_COBRANZA, normalStyle))

    # write the document to disk
    try:
        doc.build(elements)
        pdf_generado = True
        actualiza_status(txStatus, ' <pdf>: <check>')
    except:
        pdf_generado = False
        error_msg  = str(sys.exc_info()[1])
        if sys.platform.startswith('win'):
            error_msg  = error_msg.replace('\\', '/')
        actualiza_status(txStatus, f' <pdf>: <error>Error generando "{pdf_file.format(resumen=idx)}": {error_msg}')

    return pdf_generado


def envia_por_correo(r, email_o_celular, idx):

    if not envia_correos:
        return False   # El correo no fue enviado
    es_correo  = isinstance(email_o_celular, list) or is_email(email_o_celular)
    if not es_correo:
        actualiza_status(txStatus, f', <email>: <error>"{email_o_celular}" no es una dirección de correo válida')
        return False

    # Compose some of the basic message headers
    to = dict()
    for send_to in email_o_celular[0]:
        to.update({send_to: r['Beneficiario']})
    para   = to
    asunto = 'GyG Cuadra Segura - Resumen de pagos'
    cuerpo = [f"Resumen de pagos recibidos de <b>{r['Beneficiario']}, <i>{r['Dirección']}</i></b>, " + \
              f"{rango_fechas}",
              EOL,
              "<i>Equipo de Cobranza. Junta Directiva GyG</i>"]
    anexo = os.path.join(pdf_path, pdf_file.format(resumen=idx))

    try:
        email.send(to=para, subject=asunto, contents=cuerpo, attachments=anexo)
        email_sent = True
    except:
        email_sent = False
        error_msg  = str(sys.exc_info()[1])
        if sys.platform.startswith('win'):
            error_msg  = error_msg.replace('\\', '/')
        actualiza_status(txStatus, f', <email>: <error>Error enviando "{pdf_file.format(resumen=idx)}": {error_msg}')

    return email_sent


#
# Rutinas generales
#

def is_email(send_to):
    return re.search(r'[\w.-]+@[\w.-]+', send_to)

def decode_email_o_celular(email_o_celular):
    # return [str.strip(x) for x in re.split('\|', email_o_celular)]
    destinos = [str.strip(x) for x in re.split('\|', email_o_celular)]

    correos = list()
    otros_destinos = list()
    for destino in destinos:
        if is_email(destino):
            correos.append(destino)
        else:
            otros_destinos.append(destino)

    destinos = list()
    if len(correos) > 0:
        destinos.append(correos)
    for dest in otros_destinos:
        destinos.append(dest)

    return destinos


def ckSelTodo_Seleccion():
    if ckSelTodo.get() == 1:
        lbVecinos.selection_set(first=0, last=lbVecinos.size())
    else:
        lbVecinos.selection_clear(first=0, last=lbVecinos.size())
    muestra_cantidad_de_vecinos()

def ckCorreo_Seleccion():
    global email, envia_correos

    envia_correos = (ckCorreo.get() == 1)
    if envia_correos:
        email = yagmail.SMTP({credentials.email_account: credentials.email_fromaddr},
                             credentials.email_password)
    else:
        email.close()

def ckOrigen_Seleccion():
    global despliega_saldos

    despliega_saldos = (ckOrigen.get() == 'Saldos')
    if despliega_saldos:
        carga_lista_de_vecinos_con_saldo()
    else:
        carga_lista_de_vecinos()

def ckHistorico_Seleccion():
    global consulta_historico, despliega_saldos, cambio_archivo_origen, excel_workbook

    set_cursor_wait()
    consulta_historico = (ckHistorico.get() == 1)
    excel_workbook = excel_wb_historico if consulta_historico else excel_wb_estandar
    abre_hoja_de_recibos_de_pago()
    carga_lista_de_vecinos()
    carga_saldos_de_cuentas()
    cambio_archivo_origen = not cambio_archivo_origen
    ckOrigen.set('Pagos')
    despliega_saldos = False
    set_cursor_standard()


def onselect(evt):
    # Marca el Checkbutton cbSelTodo en gris indicando "selección parcial"
    w = evt.widget
    ckSelTodo.set('')
    muestra_cantidad_de_vecinos()


def set_cursor_wait():
    if sys.platform.startswith('win'):
        ventana_ppal.config(cursor=busy_cursor)
        ventana_ppal.update()


def set_cursor_standard():
    if sys.platform.startswith('win'):
        ventana_ppal.config(cursor=normal_cursor)
#        ventana_ppal.update()


def muestra_cantidad_de_vecinos():
    l = len(lbVecinos.curselection())
    s = 's' if l > 1 else ''
    tkNumVecinos.set(f"{l} vecino{s} seleccionado{s}" if l > 0 else '')
    if l > 2:
        l = 2   # limita el índice a utilizar para el texto del botón Acepta
    btAcepta['state'] = tk.NORMAL if l > 0 else tk.DISABLED
    btAcepta['text'] = ['', 'Genera resumen', 'Genera resúmenes'][l]


def actualiza_status(var, mensaje, ir_al_final=True):
    output = re.split(stPattern, mensaje)
    output = [x for x in output if (x != None) and (x != '')]

    for str in output:
        if str in stDelimiters:
            var.image_create(tk.END, image=stImages[stDelimiters.index(str)])
        else:
            var.insert(tk.END, str)

    if ir_al_final:
        txStatus.yview_moveto(1.0)


#
# Rutinas relacionadas con el control de resúmenes
#

def define_atributos(hoja_resumen):
    hoja_resumen.title = resumen_worksheet
    hoja_resumen['A1'] = 'Nro. Recibo'
    hoja_resumen['B1'] = 'Enviado'
    hoja_resumen['C1'] = 'Beneficiario'
    hoja_resumen['D1'] = 'E-mail o celular'
    hoja_resumen['E1'] = 'Dirección'
    hoja_resumen['F1'] = 'F. Desde'
    hoja_resumen['G1'] = 'F. Hasta'
    hoja_resumen['H1'] = 'F. Emisión'
    hoja_resumen['I1'] = 'Observaciones'
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        hoja_resumen[col+'1'].style = 'Accent1'
        hoja_resumen[col+'1'].alignment = Alignment(horizontal='center', wrap_text=True)
    hoja_resumen.column_dimensions['A'].width =  8.95
    hoja_resumen.column_dimensions['B'].width =  5.75
    hoja_resumen.column_dimensions['C'].width = 25.15
    hoja_resumen.column_dimensions['D'].width = 32.29
    hoja_resumen.column_dimensions['E'].width = 23.15
    hoja_resumen.column_dimensions['F'].width = 16.50
    hoja_resumen.column_dimensions['G'].width = 16.50
    hoja_resumen.column_dimensions['H'].width = 20.50
    hoja_resumen.column_dimensions['I'].width = 23.15
    hoja_resumen.auto_filter.ref = "A1:I1"


def linea_de_resumen(num_resumen, num_linea, beneficiario, email_o_celular, direccion, enviado):
    col_a = 'A{}'.format(num_linea)
    col_b = 'B{}'.format(num_linea)
    col_c = 'C{}'.format(num_linea)
    col_d = 'D{}'.format(num_linea)
    col_e = 'E{}'.format(num_linea)
    col_f = 'F{}'.format(num_linea)
    col_g = 'G{}'.format(num_linea)
    col_h = 'H{}'.format(num_linea)
    col_i = 'I{}'.format(num_linea)
    hoja_resumen[col_a] = num_resumen
    if enviado:
        hoja_resumen[col_b] = 'ü'
    hoja_resumen[col_c] = beneficiario
    hoja_resumen[col_d] = email_o_celular
    hoja_resumen[col_e] = direccion
    hoja_resumen[col_f] = '{:%B/%Y}'.format(dt_inicial)
    hoja_resumen[col_g] = '{:%B/%Y}'.format(dt_final)
    hoja_resumen[col_h] = fecha_actual
    observaciones = 'Histórico' if consulta_historico else ''
    if consulta_historico and despliega_saldos:
        observaciones += ', '
    if despliega_saldos:
        observaciones += 'Saldos'
    hoja_resumen[col_i] = observaciones
    hoja_resumen[col_a].number_format = '00000'
    hoja_resumen[col_b].font = Font(name='Wingdings')
    for columna in [col_a, col_b, col_f, col_g, col_h]:
        hoja_resumen[columna].alignment = Alignment(horizontal='center')
    hoja_resumen[col_i].alignment = Alignment(indent=1)


def linea_de_separacion(num_linea):
    linea_inferior = Border(bottom=Side(style='thin'))
    # Options must be one of {left, right, top, bottom,
    #                        diagonal, diagonal_direction, vertical, horizontal}
    # Style must be one of {‘hair’, ‘mediumDashDotDot’, ‘medium’, ‘mediumDashDot’,
    #                       ‘dashed’, ‘double’, ‘dashDotDot’, ‘dotted’, ‘slantDashDot’,
    #                       ‘thick’, ‘thin’, ‘dashDot’, ‘mediumDashed’}
    for col in "ABCDEFGHI":
        hoja_resumen[f"{col}{num_linea}"].border = linea_inferior
    hoja_resumen.cell(row=num_linea+1, column=1)


#
# Interfase de usuario
#

def genera_interfase():
    global ventana_ppal, ckSelTodo, lbVecinos, cbSelTodo, txStatus
    global spDesdeMes, spDesdeAño, spHastaMes, spHastaAño, ckOrigen, ckHistorico, ckCuotas, ckCorreo, ckSolicitado, btAcepta
    global imgCheck, imgWarn, imgError, imgInfo, imgStop, imgPdf, imgEmail
    global stDelimiters, stImages, stPattern
    global geometria
    global tkNumVecinos

    # Colores tomados de https://www.webucator.com/blog/2015/03/python-color-constants-module/
    background =       '#D3D3D3'   # LightGray
    activebackground = '#BFBFBF'   # Gray75

    ventana_ppal = tk.Tk()
    geometria = dict()

    # Try to load last saved window data
    try:
        # if the file is there, get geometry from file 
        with open(ini_file_path, 'rb') as ini_file:
#            print(f"DEBUG: genera_interfase(): Antes: {geometria=}")
            geometria = pickle.load(ini_file)
#            print(f"                         Después: {geometria=}")
            ventana_ppal.geometry(geometria['win' if sys.platform.startswith('win') else 'osx'])
    except:
        # if the file is not there, use default geometry
        # width x height + horizontal position + vertical position on the screen
        error_msg  = str(sys.exc_info()[1])
        if sys.platform.startswith('win'):
            error_msg  = error_msg.replace('\\', '/')
        print(f"*** Error en genera_interfase(): {error_msg}")
        ventana_ppal.geometry('362x390+15+15')

    ventana_ppal.configure(bg=background, cursor=normal_cursor)
    ventana_ppal.title('Resúmenes a Solicitud')
#    ventana_ppal.bind("<Configure>", save_size)
    ventana_ppal.protocol('WM_DELETE_WINDOW', finaliza_app)

    banda0 = ttk.Frame(ventana_ppal, width=10, borderwidth=0)
    banda1 = ttk.Frame(ventana_ppal, height=100, width=120)
    frVecinos = ttk.LabelFrame(banda1, text='Vecinos', height=100, width=140, padding=(5, 0),
                               borderwidth=2)
    banda2 = ttk.Frame(ventana_ppal, width=10, borderwidth=0)
    banda3 = ttk.Frame(ventana_ppal, height=100, width=120)
    frPeriodo = ttk.LabelFrame(banda3, text='Período', height=100, width=120, padding=(5, 0),
                               borderwidth=2)
    frOrigen  = ttk.LabelFrame(banda3, text='Origen', height=100, width=120, padding=(5, 0),
                               borderwidth=2)
    frStatus  = ttk.LabelFrame(ventana_ppal, text='Status',  height=100,  width=250, padding=(5, 0),
                               borderwidth=2)
    frButtons = tk.Frame(ventana_ppal, height=20, width=100)


    yScroll = tk.Scrollbar(frVecinos, orient=tk.VERTICAL)
    lbVecinos = tk.Listbox(frVecinos, height=14, width=25,   # height x width = lines x characters
                                                             # each line is about 5 pixels height
                                      activestyle='none', # exportselection=0,
                                      selectmode=tk.EXTENDED, # textvariable=stVecinos,
                                      yscrollcommand=yScroll.set)
    yScroll['command'] = lbVecinos.yview
    lbVecinos.bind('<<ListboxSelect>>', onselect)

    ckSelTodo = tk.IntVar()
    cbSelTodo = tk.Checkbutton(frVecinos, text='Selecciona todo', bg=background, pady=0,
                                          variable=ckSelTodo, command=ckSelTodo_Seleccion)
    cbSelTodo.deselect()

    tkNumVecinos = tk.StringVar()
    font_8     = font.Font(size=8)
    txBlank_1  = tk.Label(frVecinos, textvariable=tkNumVecinos, bg=background, font=font_8)


    spTxtHMes  = tk.StringVar()
    spIntHAño  = tk.IntVar()
    txtDesde   = tk.Label(frPeriodo, text='Desde:', bg=background)
    spDesdeMes = tk.Spinbox(frPeriodo, values=nombreMeses, wrap=True, width=12)
    spDesdeAño = tk.Spinbox(frPeriodo, from_=2017, to=2020, wrap=True, width=5)
    txtHasta   = tk.Label(frPeriodo, text='Hasta:', bg=background)
    spHastaMes = tk.Spinbox(frPeriodo, values=nombreMeses, wrap=True, width=12,
                                       textvariable=spTxtHMes)
    spHastaAño = tk.Spinbox(frPeriodo, from_=2017, to=2020, wrap=True, width=5,
                                       textvariable=spIntHAño)
    spTxtHMes.set(nombreMeses[ahora.month - 1])
    spIntHAño.set(ahora.year)
    
    ckOrigen   = tk.StringVar()
    cbOrigen_P = tk.Radiobutton(frOrigen, text='Pagos recibidos', value='Pagos', bg=background, pady=0,
                                          variable=ckOrigen, command=ckOrigen_Seleccion)
    cbOrigen_S = tk.Radiobutton(frOrigen, text='Saldo de cuentas', value='Saldos', bg=background, pady=0,
                                          variable=ckOrigen, command=ckOrigen_Seleccion)
    ckOrigen.set('Pagos')

    txBlank_3   = tk.Label(frOrigen, text='------------------------------', bg=background, height=0)
    txBlank_3.config(font=("Courier", 2))

    ckHistorico = tk.IntVar()
    cbHistorico = tk.Checkbutton(frOrigen, text='Histórico en Bs.F', bg=background, pady=0,
                                           variable=ckHistorico, command=ckHistorico_Seleccion)

    ckCuotas   = tk.IntVar()
    cbCuotas   = tk.Checkbutton(banda3, text='Incluye cuotas mens.', bg=background, pady=0,
                                        variable=ckCuotas)
    cbCuotas.select()

    ckSolicitado = tk.IntVar()
    cbSolicitado = tk.Checkbutton(banda3, text='Fue solicitado', bg=background, pady=0,
                                          variable=ckSolicitado)
    cbSolicitado.select()

    ckCorreo   = tk.IntVar()
    cbCorreo   = tk.Checkbutton(banda3, text='Envía como e-mail', bg=background, pady=0,
                                        variable=ckCorreo, command=ckCorreo_Seleccion)

    txBlank_2  = tk.Label(frPeriodo, text='', bg=background, height=0)
    txBlank_2.config(font=("Courier", 4))


    sScroll    = tk.Scrollbar(frStatus, orient=tk.VERTICAL)
    txStatus   = tk.Text(frStatus, padx=0, pady=0, bg=background,
                                   height=4, font=font_8, relief='flat',
                                   width=52, wrap=tk.WORD,
                                   yscrollcommand=sScroll.set)
    sScroll.configure(command=txStatus.yview)
    txStatus.insert(tk.END, 'Seleccione los vecinos y fechas deseados para esta emisión')

    s = ttk.Style()
    s.theme_use('classic')   # ('clam', 'alt', 'default', 'classic')

    btAcepta   = tk.Button(frButtons, text='', command=valida_parametros,
                                      padx=10, borderwidth=0, bg=background, activebackground=activebackground,
                                      state=tk.DISABLED)
    btCancela  = tk.Button(frButtons, text='Cancela', command=finaliza_app,
                                      padx=10, borderwidth=0, bg=background, activebackground=activebackground)

    # VENTANA_PPAL:
    banda0.grid(row=0, column=0)
    banda1.grid(row=0, column=1, sticky=tk.N)
    banda2.grid(row=0, column=2)
    banda3.grid(row=0, column=3, sticky=tk.N)
    frStatus.grid(row=2, column=1, columnspan=3, sticky=tk.E+tk.W)
    frButtons.grid(row=4, column=1, columnspan=3)

    # FRAME banda1:
    frVecinos.grid(row=1, column=0)

    # FRAME frVecinos:
    cbSelTodo.grid(row=0, column=0, sticky=tk.W)
    lbVecinos.grid(row=1, column=0)
    yScroll.grid(row=1, column=1, sticky=tk.N+tk.S)
    txBlank_1.grid(row=2, column=0, sticky=tk.W)

    # FRAME banda3:
    frPeriodo.grid(row=0, column=0, sticky=tk.N)
    frOrigen.grid(row=1, column=0, sticky=tk.E+tk.W)
    cbCuotas.grid(row=2, column=0, sticky=tk.W)
    cbSolicitado.grid(row=3, column=0, sticky=tk.W)
    cbCorreo.grid(row=4, column=0, sticky=tk.W)

    # FRAME frPeriodo:
    txtDesde.grid(row=0, column=0, sticky=tk.W)
    spDesdeMes.grid(row=1, column=0)
    spDesdeAño.grid(row=1, column=1)
    txtHasta.grid(row=2, column=0, sticky=tk.W)
    spHastaMes.grid(row=3, column=0)
    spHastaAño.grid(row=3, column=1)
    txBlank_2.grid(row=4, column=0)

    # FRAME frOrigen:
    cbOrigen_P.grid(row=2, column=0, sticky=tk.W)
    cbOrigen_S.grid(row=3, column=0, sticky=tk.W)
    txBlank_3.grid(row=1, column=0, sticky=tk.W)
    cbHistorico.grid(row=0, column=0, sticky=tk.W)

    # FRAME frStatus:
    txStatus.grid(row=0, column=0, sticky=tk.N+tk.S+tk.E+tk.W)
    sScroll.grid(row=0, column=1, sticky=tk.N+tk.S)

    # FRAME frButtons:
    btAcepta.grid(row=0, column=0, sticky=tk.E)
    btCancela.grid(row=0, column=2, sticky=tk.E)

    imgCheck = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'check.gif'))
    imgWarn  = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'warn.gif'))
    imgError = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'error.gif'))
    imgInfo  = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'info.gif'))
    imgStop  = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'stop.gif'))
    imgPdf   = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'pdf.gif'))
    imgEmail = tk.PhotoImage(file=os.path.join(GyG_constantes.rec_imágenes, 'mail.gif'))

    stDelimiters = ('<check>', '<warn>', '<error>', '<info>', '<stop>', '<pdf>', '<email>')
    stImages     = (imgCheck,  imgWarn,  imgError,  imgInfo,  imgStop,  imgPdf,  imgEmail )
    stPattern    = '|'.join(['(' + r + ')' for r in map(re.escape, stDelimiters)])


# Guarda las dimensiones y ubicación de la ventana principal
# def save_size(event):
def finaliza_app():
    print('Guardando configuración...\n')
    with open(ini_file_path, 'wb') as conf:
#        print(f"DEBUG: finaliza_app(): Antes: {geometria=}")
        geometria['win' if sys.platform.startswith('win') else 'osx'] = ventana_ppal.geometry()
#        print(f"                     Después: {geometria=}")
        pickle.dump(geometria, conf)
    ventana_ppal.destroy()


# Abre la hoja de cálculo de Recibos de Pago
def abre_hoja_de_recibos_de_pago():
    global df_resumen, df_cuotas, otros_inmuebles, cuotas_especiales
    global última_cuota, última_cuota_otros_inmuebles, última_cuota_cuotas_especiales
    global cuotas_obj

    print(f'Cargando hoja de cálculo "{excel_workbook}"...')
    df_resumen = read_excel(excel_workbook, sheet_name=excel_resumen)

    # Inicializa el handler para el manejo de las cuotas
    cuotas_obj = Cuota(excel_workbook)


# Carga la lista de vecinos en el ListBox
def carga_lista_de_vecinos():
    lista_vecinos = df_resumen['Beneficiario'][1: df_resumen[df_resumen['Dirección'] == 'TOTAL'].index[0] - 1].tolist()
    lista_vecinos.sort()
    lbVecinos.delete(0, tk.END)
    for vecino in lista_vecinos:
        lbVecinos.insert(tk.END, vecino)
    cbSelTodo.deselect()
    muestra_cantidad_de_vecinos()


# Carga la lista de vecinos con saldos en el ListBox
def carga_lista_de_vecinos_con_saldo():
    lista_vecinos = df_vecinos_con_saldo['Beneficiario'].tolist()
    lista_vecinos.sort()
    lbVecinos.delete(0, tk.END)
    for vecino in lista_vecinos:
        lbVecinos.insert(tk.END, vecino)
    cbSelTodo.deselect()
    muestra_cantidad_de_vecinos()


# Carga la hoja con los pagos para la vigilancia y elimina aquellos registros que no fueron
# seleccionados para imprimir
def carga_pagos_de_vigilancia():
    global df_pagos

    df_pagos = read_excel(excel_workbook, sheet_name=excel_vigilancia)
    df_pagos = df_pagos[['Beneficiario', 'Dirección', 'E-mail o celular',
                         'Nro. Recibo', 'Fecha', 'Monto (ed)', 'Concepto',
                         'Enviado']]
    df_pagos.dropna(subset=['Fecha'], inplace=True)
    df_pagos.drop(df_pagos.index[df_pagos['Enviado'] == 'û'], inplace=True)


def carga_saldos_de_cuentas():
    global df_vecinos_con_saldo, df_saldos
    # df_vecinos_con_saldo.columns = ['Beneficiario', 'Dirección', 'E-mail o celular']
    # df_saldos.columns = ['Beneficiario', 'Concepto', 'Fecha', 'Monto', 'Saldo']

    # Lee la hoja de cálculo de saldos
    df = read_excel(excel_workbook, sheet_name=excel_saldos)

    # Guarda los nombres de las columnas originales
    cuentas = (df.columns).tolist()
    cuentas = [col.partition(',')[0] for col in cuentas[4:] if not col.startswith('Unnamed: ')]

    # Cambia los nombres de las columnas al estilo 'C'+<consecutivo>
    columnas = [f'C{col}' for col in range(df.shape[1])]
    df.columns = columnas

    # Selecciona los nombres de los vecinos con cuentas
    df_vecinos_con_saldo = df[['C0', 'C1', 'C2']]
    df_vecinos_con_saldo.columns = ['Beneficiario', 'Dirección', 'E-mail o celular']
    df_vecinos_con_saldo = df_vecinos_con_saldo.iloc[2:]
    df_vecinos_con_saldo.dropna(subset=['Beneficiario'], inplace=True)
 #   df_vecinos_con_saldo.hide_index = True

    # Genera el dataframe "df_saldos" con Beneficiario, Fecha, Concepto, Monto, Saldo
    offset = 4   # 5 columnas antes del primer grupo
    ancho = 5    # Fecha (+1), Concepto (+2), Monto (+3), Saldo (+4), <columna en blanco> (+5)

    idx = offset
    beneficiarios = list()
    direcciones = list()
    correos = list()
    vecinos = list()
    fechas = list()
    conceptos = list()
    montos = list()
    saldos = list()
    for vecino in cuentas:
        df_saldo = df[[f'C{idx + 1}', f'C{idx + 2}', f'C{idx + 3}', f'C{idx + 4}']]
        df_saldo = df_saldo.iloc[2:]
        df_saldo.dropna(subset=[f'C{idx + 2}'], inplace=True)

        r = df_vecinos_con_saldo[df_vecinos_con_saldo['Beneficiario'] == vecino]
        beneficiarios += [vecino] * df_saldo.shape[0]
        direcciones   += [r['Dirección'].to_string(index=False)] * df_saldo.shape[0]
        correos       += [r['E-mail o celular'].to_string(index=False)] * df_saldo.shape[0]
        fechas        += df_saldo[f'C{idx + 1}'].tolist()
        conceptos     += df_saldo[f'C{idx + 2}'].tolist()
        montos        += df_saldo[f'C{idx + 3}'].tolist()
        saldos        += df_saldo[f'C{idx + 4}'].tolist()

        idx += ancho

    df_saldos = DataFrame.from_dict({'Beneficiario': beneficiarios,
                                     'Dirección':    direcciones,
                                     'E-mail o celular': correos,
                                     'Fecha':        fechas,
                                     'Concepto':     conceptos,
                                     'Monto':        montos,
                                     'Saldo':        saldos})


#
# PROCESO
#

abre_hoja_de_recibos_de_pago()

genera_interfase()

carga_lista_de_vecinos()
carga_saldos_de_cuentas()
carga_pagos_de_vigilancia()

# Lazo principal
ventana_ppal.mainloop()
