# GyG ANALISIS DE PAGOS
#
# Determina los pagos no realizados en el mes actual 

"""
    POR HACER
    -   Agrupar en "mangos bajitos", "alto impacto" y el resto, para facilitar la gestión de cobranza
        (¿cómo poder establecer esta clasificación...?)
    -   En agunos casos se ubicó a un vecino en una categoría particular, a fin de que se mostrase perma-
        nentemente en ese segmento de la Cartelera Virtual. Para estos casos, no mostrar una propuesta
        de cambio de categoría
    -   ¿Hay vecinos cuyo comportamiento de pagos es similar al de otros? Ello pudiera ayudar para influen-
        ciar sobre algunos de ellos para lograr un cambio en el grupo
    -   Agregar al resumen:
          . Para aquellos vecinos que pagan la cuota completa, indicar los meses que están por debajo de
            la cuota (incluye meses no cancelados que quedaron "encerrados" entre otros)

    HISTORICO
    -   Se corrigió un error de datos que generaba que, en la rutina 'meses_último_pago()', el último pago
        estuviese vacío ('-'): El texto del 'Concepto' en la hoja 'Vigilancia' contenía un nombre de mes
        errado ('Febero' en lugar de 'Febrero'). Al ajustar todos los meses errados, ya no se presentó
        el error de ejecución (04/02/2020)
    -   CORREGIR: La reconversión monetaria crea fórmulas más complejas que pudieran generar un error al
        realizar la distribución del pago en meses anteriores y posteriores al mes actual. (26/12/2019)
          . Se incorporó la librería PyParsing para la evaluación de las fórmulas.
          . Se unificaron los procedimientos 'separa_meses()' y 'separa_pagos()', corrigiendo adiciomal-
            mente algunos errores.
    -   Al convertir los nombres de los beneficiarios en fórmulas en la hoja 'RESUMEN VIGILANCIA', se generó
        un error en la rutina "distribución_de_pagos()": la columna 'Beneficiarios' contiene fórmulas.
          -> Convertida columna 'Beneficiarios' en valores, leyéndola con pandas.read_excel() y tomando el
             resto de las columnas con openpyxl.load_workbook() (13/12/2019)
    -   Cambiar la forma para la lectura de opciones desde el standard input, estandarizando su uso e
        implementando la opción '--toma_opciones_por_defecto' en la linea de comandos (08/12/2019)
    -   Se agrega opción para ordenar o no los resultados, y hacerlo comparable con saldos_pendientes.py
        (02/11/2019)
    -   Ajustar la codificación de caracteres para que el archivo de saldos sea legible en Apple y
        Windows
          -> Se generan archivos con codificaciones 'UTF-8' (Apple) y 'cp1252' (Windows) en carpetas
             separadas (29/10/2019)
    -   Se cambiaron las ubicaciones de los archivos resultantes a la carpeta GyG Recibos dentro de
        la carpeta actual para compatibilidad entre Windows y macOS (21/10/2019)
    -   Determinar la cantidad de pagos al 100% realizados comparando directamente el monto cancelado con
        el que estaba vigente para la fecha de pago, y no tomarlo directamente del archivo Excel.
        (07/10/2019)
    -   Cambiar el manejo de cuotas para usar las rutinas en la clase Cuota (GyG_cuotas) (25/07/2019)
    -   "Buenos días vecinos, en reunión de Junta Directiva se estableció la cuota para el pago de la
        vigilancia en un dolar en base a la tasa de cierre semanal del Banco Central, actualmente
        en Bs. 23.400
        Las cuotas pendientes por cancelar, hasta Agosto 2019, quedarán en los montos fijos ya
        establecidos. A partir del mes de Septiembre, toda cuota atrasada se cancelará con la tasa
        de la semana en curso"
          -> Ajustado (11/09/2019)
    -   Ajustar a 3 la cantidad de meses a revisar para la proposición de cambio de categoría
        (corregido 12/08/2019)  -- Se mantienen en 5 meses los reflejados en el cuadro de resumen
    -   Los pagos recibidos en un mes en particular afectan a meses anteriores y siguientes, además del mes
        en curso. ¿Cómo se distribuyen estos pagos? (24/07/2019)
    -   Mostar saldo deudor en TODAS las categorías, no sólo en 'Cuota completa' (07/06/2019)
          . La cuota del 'Colegio El Trigal' (y, anteriormente, la de 'La Casita Encantada'), no es la misma
            que la del resto de los vecinos, por lo que el monto indicado no es correcto (corregido 09/06/2019)
    -   Ajustes de redacción en los mensajes (05/06/2019)
    -   En genera_pagos_mensuales() hay que restar 4 del conteo de pagos por mes (probablemente corresponden
        a las lineas TOTAL, NÚMERO DE PAGOS, PAGOS 100% CUOTA y PAGOS 100% EQUIVALENTES) ¿cómo eliminar
        esta constante?
          . 'linea_totales' contiene el índice de la fila que contiene los totales, no el número físico de la
            linea. La diferencia es producto de las eliminaciones de renglones en el proceso de limpieza
          . Se reinició el índice para ubicar correctamente la linea de totales (corregido 15/05/2019)
    -   Mostrar el saldo pendiente en bolívares en los vecinos que pagan cuota completa (29/04/2019)
    -   En aquellos vecinos que parcicipan en el programa de comidas para los vigilantes, incluir un comen-
        tario al respecto (14/04/2019)
    -   Mostrar la lista de vecinos en orden alfabético (09/04/2019)
    -   Se muestran en el análisis los vecinos en fecha posterior a "F.Hasta" y "Categoría" no nula
        -- Corregido 02/02/2019
    -   Agregar al resumen:
          . Cantidad de pagos recibidos iguales o superiores a la cuota
          . Cantidad de pagos recibidos equivalentes a pagos al 100% de la cuota (total recaudado / cuota)
    -   Mostrar el saldo a la fecha de aquellos vecinos quienes han hecho depósitos por adelantado
    -   Manejar adecuadamente aquellos registros con Promedio o Variación == None (=> no se encontraron
        suficientes pagos en un período de <nMeses> meses)
    -   No eliminar los registros de aquellos vecinos que pagan cuota completa y el monto cancelado en el
        mes de análisis sea inferior a la cuota del mes
        ( df_resumen = df_resumen.loc[isnull(df_resumen[datetime(año, mes, 1)])] )
    -   Mostrar la cantidad de cuotas recaudadas, el monto total y el monto de la cuota de los últimos
        <nMeses>
    -   Colocar la propuesta de cambio de categoría al final del análisis del vecino, si corresponde
    -   Destacar las propuestas de cambio de categoría cuando la resultante sea diferente a "No participa"
    -   Evaluar cambios de categoría: Cuota completa, Colaboración o No participa
        ("más de xx meses cancelando {cuota completa | colaboración}. Cambiar su clasificación a 'xxx'")

"""

print('Cargando librerías...')
import GyG_constantes
from GyG_cuotas import *
from GyG_utilitarios import *
from pandas import DataFrame, read_excel, isnull, notnull, to_numeric
from numpy import mean, std, NaN
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from pyparsing import Word, Regex, Literal, OneOrMore, ParseException
import re
import sys
import os
import numbers
# from re import match, findall
import locale
dummy = locale.setlocale(locale.LC_ALL, 'es_es')

# Define textos
nombre_análisis = GyG_constantes.txt_analisis_de_pago     # "GyG Analisis de Pagos {:%Y-%m (%B)}.txt"
attach_path     = GyG_constantes.ruta_analisis_de_pagos   # "./GyG Recibos/Análisis de Pago"

excel_workbook          = GyG_constantes.pagos_wb_estandar             # '1.1. GyG Recibos.xlsm'
excel_resumen           = GyG_constantes.pagos_ws_resumen              # 'RESUMEN VIGILANCIA'
excel_worksheet_resumen = GyG_constantes.pagos_ws_resumen_reordenado   # 'R.VIGILANCIA (reordenado)'
excel_worksheet_detalle = GyG_constantes.pagos_ws_vigilancia           # 'Vigilancia'
excel_worksheet_Vecinos = GyG_constantes.pagos_ws_vecinos              # 'Vecinos'
excel_worksheet_cuotas  = GyG_constantes.pagos_ws_cuotas               # 'CUOTA'
excel_worksheet_saldos  = GyG_constantes.pagos_ws_saldos               # 'Saldos'

nMeses                  = 3  # Meses a analizar para propuestas de cambio de categoría
nMeses_resumen          = 5  # Meses a mostrar en el cuadro resumen al inicio del análisis


# Gramática de expresiones aritméticas
operator  = Regex(r'(?<!--[\+\-\^\*/%])[\+\-]|[\^\*/%!]')
function  = Regex(r'[a-zA-Z_][a-zA-Z0-9_]*(?=([ \t]+)?\()')
variable  = Regex(r'[+-]?[a-zA-Z_][a-zA-Z0-9_]*(?!([ \t]+)?\()')
number    = Regex(r'[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?')       # Ej. -125,54e-5
lbrace    = Word('(')
rbrace    = Word(')')
linebreak = Word('\n')
skip      = Word(' \t')                                             # espacio o tabulador

lexOnly = operator | function | variable | number | lbrace \
    | rbrace | linebreak | skip
lexAllOnly = OneOrMore(lexOnly)


toma_opciones_por_defecto = False
if len(sys.argv) > 1:
    toma_opciones_por_defecto = sys.argv[1] == '--toma_opciones_por_defecto'


#
# DEFINE ALGUNAS RUTINAS UTILITARIAS
#

def esVigilancia(x):
    return x == 'Vigilancia'

def get_street(address):
    return address.index(' ', address.index(' ') + 1)

def seleccionaRegistro(beneficiarios, categorías, montos):

    def list_and(l1, l2): return [a and b for a, b in zip(l1, l2)]
    def list_or(l1, l2):  return [a or  b for a, b in zip(l1, l2)]
    def list_not(l1):     return [not a for a in l1]

    def list_lt_cuota(l1):
#        l2 = [a if is_numeric(a) else cuotas_mensuales[last_col] for a in l1]
#        return [a < cuotas_mensuales[last_col] for a in l2]
        l2 = list()
        for beneficiario, monto in zip(beneficiarios, montos):
            f_ultimo_pago = fecha_ultimo_pago(beneficiario, columns[last_col].strftime('%m-%Y'))
            if f_ultimo_pago == None:
                l2.append(False)
            else:
                cuota = cuotas_obj.cuota_vigente(beneficiario, f_ultimo_pago)
                l2.append(monto < cuota)
        return l2

    # Selecciona aquellos que pagan cuota completa y el monto del mes analizado es inferior
    # al establecido o no lo han pagado
    list_1 = list_or(
                        list_and(categorías == 'Cuota completa', list_lt_cuota(montos)),
                        list_and(categorías == 'Cuota completa', isnull(montos))
                    )
    # Selecciona aquellos que colaboran, pero no han pagado el mes analizado
    list_2 = list_and(categorías == 'Colaboración', isnull(montos))
    # Selecciona aquellos que tienen una cuenta con saldo a favor
    df_saldos_gt_0 = df_saldos[df_saldos['Saldo'] > 0]
    list_3 = [len(df_saldos_gt_0[df_saldos_gt_0['Beneficiario'] == b]) > 0 for b in beneficiarios]

    return list_or(list_1, list_or(list_2, list_3))

def dia_promedio(r):
    # Determina el día promedio de pago en los últimos 'nn' meses
    dias = [x.day for x in list(df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']]['Fecha'])]
    dias = dias[-nMeses:]
    if len(dias) == 0:
        promedio = None   # No se encontraron suficientes pagos en un período de <nMeses> meses
    else:
        promedio = int(round(mean(dias), 0))
    return promedio

def desviación(r):
    # Determina la desviación estándar del día promedio de pago en los últimos 'nn' meses
    """
    dias = [x.day for x in list(df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']]['Fecha'])]
    dias = dias[-nMeses:]
    return std(dias)
    """
    fechas = list(df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']]['Fecha'])
    fechas.sort()
    fechas = fechas[-nMeses:]
    dif_fechas = [(fechas[i+1] - fechas[i]).days for i in range(len(fechas) - 1)]
    if len(dif_fechas) == 0:
        desv_est = None   # No se encontraron suficientes pagos en un período de <nMeses> meses
    else:
        desv_est = std(dif_fechas)
    return desv_est
    

meses            = ['enero',      'febrero', 'marzo',     'abril',
                    'mayo',       'junio',   'julio',     'agosto',
                    'septiembre', 'octubre', 'noviembre', 'diciembre']
meses_abrev      = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
                    'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
conectores       = ['a', '-']
textos_anticipos = ['adelanto', 'anticipo'   ]
textos_saldos    = ['ajuste',   'complemento', 'diferencia', 'saldo']
modificadores    = ['anticipo', 'saldo']

tokens_validos = meses + meses_abrev + conectores

def separa_meses(mensaje, as_string=False, muestra_modificador=False):
    tokens_validos = meses + meses_abrev + conectores + modificadores

    mensaje = re.sub("\([^()]*\)", "", mensaje)
    mensaje = mensaje.lower().replace('-', ' a ').replace('/', ' ')
    for token in textos_anticipos:
        mensaje.replace(token, modificadores[0])
    for token in textos_saldos:
        mensaje.replace(token, modificadores[1])
    mensaje = re.sub(r"\W ", " ", mensaje).split()
    mensaje_ed = [x for x in mensaje if (x in tokens_validos) or x.isdigit()]
    last_year = None
    last_month = None
    acción = ''
    mensaje_anterior = None
    mensaje_final = list()
    maneja_conector = False
    for x in reversed(mensaje_ed):
        token = meses[meses_abrev.index(x)] if x in meses_abrev else x
        if token.isdigit():
            if mensaje_anterior != None:
                mensaje_final = mensaje_anterior + mensaje_final
            last_year = token
            last_month = None
            mensaje_anterior = None
        elif token in meses:
            if mensaje_anterior != None:
                mensaje_final = mensaje_anterior + mensaje_final
            if maneja_conector:
                try:
                    n_last_month = meses.index(last_month)
                except:
                    continue    # ignora los mensajes que contienen textos del tipo:
                                # "(saldo a favor: Bs. 69.862,95)"
                n_token = meses.index(token)
                for t in reversed(range(n_token + 1, n_last_month)):
#                        mensaje_final = [f"{meses_abrev[t]}.{last_year}"] + mensaje_final
                    mensaje_final = [f"{t+1:02}-{last_year}"] + mensaje_final
                maneja_conector = False
            last_month = token
#                mensaje_anterior = [f"{meses_abrev[meses.index(last_month)]}.{last_year}"]
            mensaje_anterior = [f"{meses.index(last_month)+1:02}-{last_year}"]
        elif x in conectores:
            maneja_conector = True
        elif x in modificadores and muestra_modificador:
#                mensaje_final = [f"{meses_abrev[meses.index(last_month)]}.{last_year} {x}"] + mensaje_final
            mensaje_final = [f"{meses.index(last_month)+1:02}-{last_year} {x}"] + mensaje_final
            mensaje_anterior = None

    if mensaje_anterior != None:
        mensaje_final = mensaje_anterior + mensaje_final

    if as_string:
        mensaje_final = '|'.join(mensaje_final)

    return mensaje_final


def fecha_ultimo_pago(beneficiario, str_mes_año):
    try:
        df_fecha = df_pagos[(df_pagos['Beneficiario'] == beneficiario) & (df_pagos['Meses'].str.contains(str_mes_año))]
    except:
        return None
    if df_fecha.shape[0] > 0:
        fecha_pago = df_fecha.iloc[-1]['Fecha'].to_pydatetime()
        fecha_objetivo = datetime(int(str_mes_año[3:]), int(str_mes_año[0:2]), 1)
        if fecha_objetivo < GyG_constantes.fecha_de_corte:
            return fecha_objetivo
        else:
            return datetime.today() if fecha_pago < fecha_objetivo else fecha_pago
    else:
        return None

def no_participa_desde(r):
    """
        Busca a partir de qué fecha no se han recibido pagos
        (evalúa desde el mes y año indicado, hasta el 2016)
    """

    x = last_col       # <-------- 'x' es la primera columna vacía
    sum_cuotas = 0.00
    f_desde = columns.index(r['F.Desde'])
    saldo_a_favor   = False
    saldo_pendiente = False

    for idx in reversed(range(f_desde, last_col+1)):
        if notnull(r.iloc[idx]):
            break
        x = idx
        sum_cuotas += cuotas_obj.cuota_actual(r['Beneficiario'], columns[idx]) # cuotas[idx]
    this_col = columns[x] if isnull(r.iloc[last_col]) else columns[x]   # <<<=== anteriormente 'x+1' en lugar de 'x'
    fecha_txt = '2016' if this_col == datetime(2016, 1, 1) else this_col.strftime('%B %Y')

    # Determina el saldo del último mes
    if r.iloc[x] == 'ü':       # 'check'
        deuda = 0.00           # La mensualidad está saldada por completo
    else:
        f_ultimo_pago = fecha_ultimo_pago(r['Beneficiario'], columns[x].strftime('%m-%Y'))
        if (f_ultimo_pago == None) or (r['Categoría'] == 'Colaboración'):
            deuda = 0.00   # Probablemente es un vecino que nunca ha pagado
        else:
            cuota_actual = cuotas_obj.cuota_vigente(r['Beneficiario'], f_ultimo_pago)
            deuda = cuota_actual - r[columns[x]]
            # Si el monto cancelado no cubre la cuota del período, recalcula la deuda del mes en base
            # a la última cuota
            if (deuda > 0.00) and (f_ultimo_pago >= GyG_constantes.fecha_de_corte):
                cuota_actual = cuotas_obj.cuota_vigente(r['Beneficiario'], datetime.now())
                deuda = cuota_actual - r[columns[x]]

    deuda = sum_cuotas + (deuda if deuda > 0.00 else 0.00)

    info_deuda = ''
    if deuda != 0.00:
        if last_col - x <= 0:   # <<<=== anteriormente '1' en lugar de '0'
            if notnull(r[datetime(año, mes, 1)]):
                mensaje = 'tiene un saldo pendiente en ' + fecha_txt
                saldo_pendiente = True
            else:
                mensaje = 'tiene pendiente ' + fecha_txt
        else:
            mensaje = 'tiene cuotas pendientes desde ' + fecha_txt
            saldo_pendiente = True
    else:
        mensaje = 'no tiene saldos pendientes'

    # Valida si tiene un saldo a favor
    if df_saldos['Beneficiario'].str.contains(r['Beneficiario'], regex=False).any():
        saldo = float(df_saldos[df_saldos['Beneficiario'] == r['Beneficiario']]['Saldo'])   # .item()
        saldo = edita_número(saldo, num_decimals=0)
        if saldo != '0':
            mensaje += f'\n  Dispone de un saldo a favor de Bs. {saldo}'
            saldo_a_favor = True

    if (deuda != 0.00):
        if saldo_pendiente:
            sep1, sep2 = (' por ', '')
            if x == last_col:
                sep1 = ' de '
        else:
            sep1, sep2 = (' (', ')')
        if saldo_a_favor:
            sep1, sep2 = (' y un saldo deudor de ', '')   # muestra saldo a favor y saldo deudor
        info_deuda = f"{sep1}Bs. {edita_número(deuda, num_decimals=0)}{sep2}"
    else:
        info_deuda = ""

    return mensaje, info_deuda

def mas_de_un_mes(beneficiario):
    pagos_beneficiario = (df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']])['Concepto'].tolist()
    num_pagos = len(pagos_beneficiario)
    num_pagos_multiples = 0
    for pago in pagos_beneficiario:
        if 'meses' in pago:
            num_pagos_multiples += 1
    pct_pagos_multiples = num_pagos_multiples / num_pagos
    # convierte 'pct_pagos_multiples' en un entero múltiplo de 5
    pct_pagos_multiples = int(round(pct_pagos_multiples * 20) / 20 * 100)
    if num_pagos_multiples > 0:
        return f'\n  En el {pct_pagos_multiples}% de los casos ({num_pagos_multiples} de {num_pagos}), ' + \
                'su pago correspondió a más de un mes'
    else:
        return "\n  Nunca ha pagado más de un mes simultáneamente"

def genera_pagos_mensuales():
    # Reinicia el índice para ubicar correctamente la linea de totales
    df_resumen.reset_index(drop=True, inplace=True)

    # Ubica la línea de totales
    linea_totales = df_resumen.loc[df_resumen['Dirección'] == 'TOTAL'].index[0]
    return [to_numeric(df_resumen[column].iloc[0:linea_totales], errors='coerce').count()
                for column in list(df_resumen.columns)]

def genera_resumen():
    global análisis

    def espacios(width=1, char=' '):
        return char * width

    análisis += f'Resumen de los últimos {nMeses_resumen} meses:\n'
    análisis += espacios(12) + 'Pagos' + espacios(31) + 'Pagos  Equiv.       Pagos         Distribución\n'
    análisis += '  Mes     |  mes | Total Bs. |  Prom. |Cuota(*)| 100% | 100%' + \
                   espacios(6) + 'recibidos  |  Ant.   Mes   Pos.\n'
    análisis += '  ' + espacios(59, '-') + '  ' + espacios(35, '-') + '\n'
    dist_idx = 0
    for idx in range(last_col - nMeses_resumen + 1, last_col + 1):
        mes = format(columns[idx], '%b%Y')   #.capitalize()
        num_pagos = pagos_mensuales[idx]
        tot_pagos = edita_número(totales_mensuales[idx], num_decimals=0)
#       num_pagos_100_pct = int(num_100_pct[idx])   # <<<=== Determinar la cantidad de pagos al 100% realizados
                                                    #        comparando directamente el monto cancelado con el que
                                                    #        estaba vigente para la fecha de pago, y no tomarlo
                                                    #        directamente del archivo Excel.
#       num_pagos_100_pct = cuenta_pagos_completos(columns[idx])
        num_pagos_100_pct = pagos_100_pct[dist_idx] # <<<=== Valores tomados antes de eliminar vecinos no seleccionados
        promedio  = edita_número(round(totales_mensuales[idx] / pagos_mensuales[idx], -1), num_decimals=0)
        if notnull(cuotas_mensuales[idx]):
            cuota = cuotas_mensuales[idx]
            num_pagos_eqv = round(totales_mensuales[idx] / cuota, 1)
            # pct = round((totales_mensuales[idx] / pagos_mensuales[idx]) / cuotas_mensuales[idx] * 100)
        else:
            cuota = 0
            num_pagos_eqv = 0
            # pct = 0
        cuota = edita_número(cuota, num_decimals=0)
        num_pagos_eqv = edita_número(num_pagos_eqv, num_decimals=1)
        total = sum(distribución[dist_idx][1:])
        dist_num_pagos = distribución[dist_idx][0]
        dist_tot_pagos = edita_número(total, num_decimals=0)
        dist_pct_ant   = edita_número(distribución[dist_idx][1] / total * 100, num_decimals=1) + '%'
        dist_pct_mes   = edita_número(distribución[dist_idx][2] / total * 100, num_decimals=1) + '%'
        dist_pct_pos   = edita_número(distribución[dist_idx][3] / total * 100, num_decimals=1) + '%'
        análisis    += f'  {mes:8}{num_pagos:>5}{tot_pagos:>13}{promedio:>9}{cuota:>9}' + \
                       f'{num_pagos_100_pct:>6}{num_pagos_eqv:>8}' + \
                       f'{dist_num_pagos:>6}{dist_tot_pagos:>11}' + \
                       f'{dist_pct_ant:>7}{dist_pct_mes:>7}{dist_pct_pos:>7}\n'
        dist_idx += 1

    análisis += '\n  (*) A partir de septiembre 2019 se muestran los promedios de las cuotas semanales'
    análisis += '\n\n'

def genera_propuesta_categoría():

    def genera_propuesta(r):
        # comparar <r> con <cuotas_mensuales> en los últimos <nMeses> meses [last_col - nMeses + 1: last_col]
        monto_cuotas = [cuotas_obj.cuota_actual(r['Beneficiario'], columns[col]) \
                                for col in range(last_col - nMeses + 1, last_col + 1)]
        pagos = (r.iloc[last_col - nMeses + 1: last_col + 1]).tolist()
        pagos = [p if is_numeric(p) else NaN for p in pagos]
        # Cuota completa: todos los pagos son mayores o iguales a la cuota del mes
        if all(p >= m for p, m in zip(pagos, monto_cuotas)):
            propuesta = 'Cuota completa'
        # Colaboración: todos los pagos son inferiores a la cuota del mes
        elif all(p <  m for p, m in zip(pagos, monto_cuotas)):
            propuesta = 'Colaboración'
        # No colabora: todos los pagos son nulos
        elif all(isnull(p) for p in pagos):
            propuesta = 'No participa'
        # En cualquier otro caso, el resultado es indeterminado
        else:
            propuesta = NaN
        if r['Categoría'] == propuesta:
            propuesta = NaN
        return propuesta

    df = df_resumen[['Beneficiario', 'Categoría', 'Comida']]
    df = df[notnull(df['Beneficiario'])]
    df = df[df['Beneficiario'] != 'CUOTAS MENSUALES']
    df = df[notnull(df['Categoría'])]
    df['Propuesta'] = df_resumen.apply(genera_propuesta, axis=1)
    df = df[notnull(df['Propuesta'])]
    return df

def propuesta_de_cambio(beneficiario):
    r = df_categoría[df_categoría['Beneficiario'] == beneficiario]
    if r.shape[0] == 0:
        return ''
    propuesta = r['Propuesta'].to_string(index=False)
    if propuesta != 'No participa':
        propuesta = propuesta.upper().strip()
    return(f"\n  Se sugiere cambiar su categoría de pago a '{propuesta}'")

def programa_de_comida(comida):
    txt = ''
    if not(isnull(comida) or comida == ' '):
        txt = '\n  Actualmente participa en el programa de comidas para los vigilantes'
    return txt

def get_filename(filename):
    return os.path.basename(filename)

def is_numeric(valor):
    return isinstance(valor, numbers.Number)

def distribución_de_pagos():
    from openpyxl import load_workbook
    import warnings
    warnings.simplefilter("ignore", category=UserWarning)

    def ajusta_fecha(ref):
        """ ajusta_fecha toma una fecha 'ref' de la forma '%m-%Y' (ej. 'may.2019') y la convierte en
            '%Y-%m' ('2019-05') para facilitar su comparación """
        return f"{ref[3:7]}-{ref[0:2]}"

    def análisis_de_pagos(fecha_referencia):

        def evalua_celda(formula, modificador):
            try:
                parsed_expression = lexAllOnly.parseString(formula[1:], parseAll=True)
            except ParseException as err:
                print(f"{err.line}\n{' '*(err.column-1)}^\n{err}\n")
                return None
            result = ''
            count_braces = 0
            start_idx = 0
            curr_idx = 0
            for token in parsed_expression:
                if token == '(':
                    count_braces += 1
                elif token == ')':
                    count_braces -= 1
                elif token in ['+', '-'] and count_braces == 0:
                    result += ''.join(parsed_expression[start_idx: curr_idx])
                    if len(result) > 0 and modificador in [modificadores[0], '']:   # Cancela un Anticipo
                        return eval(result)
                    result = ''
                    start_idx = curr_idx
                curr_idx += 1
            result = ''.join(parsed_expression[start_idx:])
            return eval(result)

        pagos = df_pagos.drop(df_pagos.index[df_pagos['Fecha'] != fecha_referencia])
        total_meses_anteriores = total_mes_actual = total_meses_siguientes = 0
        num_pagos = 0
        for index, pago in pagos.iterrows():
            beneficiario = pago['Beneficiario']
            meses_pagados = separa_meses(pago['Concepto'], muestra_modificador=True)
            num_pagos += 1
            r = df_resumen[df_resumen['Beneficiario'] == beneficiario]
            for mes in separa_meses(pago['Concepto'], muestra_modificador=False):
                modificador = ''
                for mes_pago in meses_pagados:
                    if mes_pago.startswith(mes):
                        modificador = mes_pago[len(mes) + 1:]
                try:
                    pago_mes = r[mes].values[0]
                except:
                    print(f"\n*** ERROR: {str(sys.exc_info()[1])} (Beneficiario: {pago['Beneficiario']}, mes: {mes})")
#                    print(f"DEBUG: df_resumen['Beneficiario']=\n{df_resumen['Beneficiario']}, {beneficiario=}")
#                    print(f"DEBUG: r=\n{r}")
                    sys.exit()
                if pago_mes == None:
                    monto = 0
                elif isinstance(pago_mes, str):
                    try:
                        _ = eval(pago_mes[1:])   # Comprueba si es una fórmula válida
                    except:
                        print(f'*** Error evaluando "{pago_mes}", ({beneficiario}, {mes})')
                        monto = 0
                    else:
                        monto = evalua_celda(pago_mes, modificador)
                else:
                    monto = pago_mes
                if ajusta_fecha(mes) < ajusta_fecha(fecha_referencia):
                    total_meses_anteriores += monto
                elif ajusta_fecha(mes) > ajusta_fecha(fecha_referencia):
                    total_meses_siguientes += monto
                else:   # mes == fecha_referencia
                    total_mes_actual += monto

        return (num_pagos, total_meses_anteriores, total_mes_actual, total_meses_siguientes)


    # Lee la hoja con el resumen de pagos. Usa la libreria OpenPyXL para preservar
    # las fórmulas en las columnas y genera un dataframe de Pandas como resultado

    # Maneja la primera columna ('Beneficiario') como su valor asociado
    df_resumen = read_excel(excel_workbook, sheet_name=excel_resumen, usecols=['Beneficiario'])

    # Y el resto de las columnas como fórmulas
    wb = load_workbook(filename=excel_workbook)
    ws = wb[excel_resumen]
    col_value = None
    for idx, col in enumerate(ws.columns):
        if idx == 0:
            continue
        curr_col = [expr.value for expr in col]
        col_name = curr_col.pop(0)
        try:
            if idx == 2:
                col_name = 'Check'
            elif col_name[0] == '=':
                col_name = col_value + relativedelta(months=1)
        except:
            pass
        col_value = col_name
        df_resumen[col_name] = curr_col

    # Ajusta los nombres de columnas
    new_cols = list()
    for col in df_resumen.columns.values.tolist():
        if isinstance(col, datetime):
            new_cols.append(col.strftime('%m-%Y'))
        else:
            new_cols.append(col)
    df_resumen.columns = new_cols

    # Lee la hoja con el detalle de los pagos recibidos, elimina aquellos cuya Referen-
    # cia no corresponda al pago de Vigilancia y estandariza la fecha de pago
    df_pagos = read_excel(excel_workbook, sheet_name=excel_worksheet_detalle)
    #df_pagos = df_pagos[df_pagos['Categoría'] == 'Vigilancia']
    df_pagos.drop(df_pagos.index[df_pagos['Categoría'] != 'Vigilancia'], inplace=True)
    df_pagos = df_pagos[['Beneficiario', 'Dirección', 'Fecha', 'Monto', 'Concepto', 'Mes', 'Nro. Recibo']]
    df_pagos.sort_values(by=['Beneficiario', 'Fecha'], inplace=True)
    #df_pagos.dropna(subset=['Fecha'], inplace=True)
    df_pagos['Fecha'] = df_pagos['Fecha'].apply(lambda x: f'{x:%m-%Y}')


    def edit(valor, width=9, decimals=0):
        valor_ed = edita_número(valor, num_decimals=decimals)
        return f"{valor_ed:>{width}}"

    def edit_pct(valor, total, width=5, decimals=1):
        return edit(valor/total*100, width-1, decimals) + '%'


    lista_resultados = list()
    for offset in reversed(range(nMeses_resumen)):
        f_ref = f"{date(año, mes, 1) - relativedelta(months=offset) + relativedelta(day=1):%m-%Y}"
        resultado = análisis_de_pagos(f_ref)
        total = sum(resultado[1:])
        lista_resultados.append(resultado)

        print(f"  {f_ref}: {edit(total, width=10)} [" + \
              f"{edit(resultado[1])} ({edit_pct(resultado[1], total)}), " + \
              f"{edit(resultado[2], width=10)} ({edit_pct(resultado[2], total)}), " + \
              f"{edit(resultado[3])} ({edit_pct(resultado[3], total)})"   + "]")

    return lista_resultados

def cuenta_pagos_completos(fecha):
    # pagos_completos = 0
    # Para cada vecino
    #   busca la fecha del último pago del vecino para la fecha indicada
    #     ejemplo: f_ultimo_pago = fecha_ultimo_pago(beneficiario, fecha.strftime('%m-%Y'))
    #   busca la cuota vigente para la fecha del último pago
    #     ejemplo: cuota = cuotas_obj.cuota_vigente(beneficiario, f_ultimo_pago)
    #   incrementa el número de pagos completos si el monto cancelado en el mes es mayor
    #   o igual a la cuota encontrada
    # return pagos_completos
    pagos_completos = 0
    for index, r in df_resumen.iterrows():
        f_ultimo_pago = fecha_ultimo_pago(r['Beneficiario'], fecha.strftime('%m-%Y'))
        if f_ultimo_pago == None:
            continue
        cuota_del_momento = cuotas_obj.cuota_vigente(r['Beneficiario'], f_ultimo_pago)
        monto_del_mes = r[fecha]
        if not is_numeric(monto_del_mes):
            continue
        if monto_del_mes >= cuota_del_momento:
            pagos_completos += 1
    return pagos_completos


def meses_último_pago(beneficiario, fecha):

    def une_lista(lista):
        if not lista:
            return ''
        if len(lista) == 1:
            return lista[0]
        return f"{', '.join(lista[:-1])} y {lista[-1]}"

    # obtiene un string del tipo "05-2019|06-2019|07-2019|07-2019" con posibles duplicados
    último_pago = '|'.join(df_pagos[(df_pagos['Beneficiario'] == beneficiario) & (df_pagos['Fecha'] == fecha)].Meses.tolist())
    # obtiene los grupos  separados por '|' y elimina los duplicados
    último_pago = list(set(último_pago.split('|')))
    # convierte al formato 'yyyy-mm' y lo ordena de menor a mayor
    último_pago = [f[3:] + '-' + f[:2] for f in reversed(último_pago)]
    último_pago.sort()
    # convierte los meses al tipo "may., jun. y jul. 2019"
    año = None
    meses_cancelados = list()
    for f in reversed(último_pago):
        if (año == None) or (f[:4] != año):
            año = f[:4]
#            print(f'DEBUG: {beneficiario:20}, fecha: "{f}", último pago: {último_pago}')
            meses_cancelados.insert(0, meses_abrev[int(f[-2:])-1] + '. ' + año)
        else:
            meses_cancelados.insert(0, meses_abrev[int(f[-2:])-1] + '.')
    último_pago = une_lista(meses_cancelados)

    return último_pago


#
# PROCESO
#

# Determina el mes actual, a fin de utilizarlo como opción por defecto
# mes_actual = datetime.now().strftime('%m-%Y')
hoy = datetime.now()
fecha_análisis = datetime(hoy.year, hoy.month, 1) - timedelta(days=1)
mes_actual = fecha_análisis.strftime('%m-%Y')
print()

# Selecciona el mes y año a procesar
mes_año = input_mes_y_año('Indique el mes y año a analizar', mes_actual, toma_opciones_por_defecto)

# Selecciona si se muestran sólo los saldos deudores o no
solo_deudores = input_si_no('Sólo vecinos con saldos pendientes', 'sí', toma_opciones_por_defecto)

# Selecciona si se ordenan alfabéticamente los vecinos
ordenado = input_si_no('Ordenados alfabéticamente', 'no', toma_opciones_por_defecto)

año = int(mes_año[3:7])
mes = int(mes_año[0:2])
fecha_referencia = datetime(año, mes, 1)
f_ref_último_día = fecha_referencia + relativedelta(months=1) + relativedelta(days=-1)
fecha_análisis = format(datetime(año, mes, 1), '%B/%Y').capitalize()

print("\nDeterminando distribución de pagos...")
distribución = distribución_de_pagos()

# Abre la hoja de cálculo de Recibos de Pago
print()
print('Cargando hoja de cálculo "{filename}"...'.format(filename=excel_workbook))
df_resumen = read_excel(excel_workbook, sheet_name=excel_worksheet_resumen)

# Inicializa el handler para el manejo de las cuotas
cuotas_obj = Cuota(excel_workbook)

# Cambia el nombre de la columna 2016 a datetimme(2016, 1, 1)
df_resumen.rename(columns={2016:datetime(2016, 1, 1)}, inplace=True)

# Define algunas variables necesarias
columns = list(df_resumen.columns.values)
last_col = columns.index(datetime(año, mes, 1))
hoy = datetime.now().day

# Ajusta las columnas F.Desde y F.Hasta en aquellos en los que estén vacíos:
# 01/01/2016 y 01/mes/año+1
df_resumen.loc[df_resumen[isnull(df_resumen['F.Desde'])].index, 'F.Desde'] = date(2016, 1, 1)
df_resumen.loc[df_resumen[isnull(df_resumen['F.Hasta'])].index, 'F.Hasta'] = date(año + 1, mes, 1)

# Elimina aquellos vecinos que compraron (o se iniciaron) en fecha posterior
# a la fecha de análisis
df_resumen = df_resumen[df_resumen['F.Desde'] < f_ref_último_día]

# Elimina aquellos vecinos que vendieron (o cambiaron su razón social) en fecha anterior
# a la fecha de análisis
df_resumen = df_resumen[df_resumen['F.Hasta'] >= f_ref_último_día]

# Lee la pestaña con el detalle de pagos
df_pagos = read_excel(excel_workbook, sheet_name=excel_worksheet_detalle)

# Genera una columna con el resumen de los meses cancelados en cada pago
meses_cancelados = df_pagos['Concepto'].apply(lambda x: separa_meses(x, as_string=True))
df_pagos.insert(column='Meses', value=meses_cancelados, loc=df_pagos.shape[1])

# Elimina los registros que no no corresponden a pago de vigilancia
df_pagos = df_pagos.loc[esVigilancia(df_pagos['Categoría'])]

# Conserva las líneas con las cuotas y los totales mensuales
cuotas_mensuales  = (df_resumen[df_resumen['Beneficiario'] == 'CUOTAS MENSUALES'].values.tolist())[0]
totales_mensuales = (df_resumen[df_resumen['Dirección']    == 'TOTAL'           ].values.tolist())[0]
num_100_pct       = (df_resumen[df_resumen['Dirección']    == 'PAGOS 100% CUOTA'].values.tolist())[0]
pagos_100_pct     = [cuenta_pagos_completos(columns[idx]) for idx in range(last_col-nMeses_resumen+1, last_col+1)]

# Determina la cantidad de pagos recibidos por mes
pagos_mensuales = genera_pagos_mensuales()
columnas = df_resumen.columns.values.tolist()

# Genera una propuesta de cambio de categoría de pago de cada vecino
df_categoría = genera_propuesta_categoría()

# Toma los saldos a favor de algunos vecinos para forzar su selección en el paso siguiente
df_saldos  = read_excel(excel_workbook, sheet_name=excel_worksheet_saldos, skiprows=[0, 1])
df_saldos  = df_saldos[['Beneficiario', 'Dirección', 'Saldo']]
df_saldos.dropna(subset=['Beneficiario'], inplace=True)

# Elimina los registros que no tienen una categoría definida, aquellos donde
# el mes a evaluar ya está cancelado, y aquellos en los cuales el beneficiario
# no participa en el pago de vigilancia
df_resumen.dropna(subset=['Categoría'], inplace=True)
if solo_deudores:
    df_resumen = df_resumen.loc[seleccionaRegistro(df_resumen['Beneficiario'],
                                                   df_resumen['Categoría'],
                                                   df_resumen[datetime(año, mes, 1)])]

# Ordena los beneficiarios en orden alfabético
if ordenado:
    df_resumen.sort_values(by='Beneficiario', inplace=True)

# Inserta las columnas 'Promedio' y 'Variación' con el día promedio de pago y su
# desviación estándar
prom  = df_resumen.apply(dia_promedio, axis=1)
desv  = df_resumen.apply(desviación, axis=1)
df_resumen.insert(column='Promedio',  value=prom, loc=df_resumen.shape[1])
df_resumen.insert(column='Variación', value=desv, loc=df_resumen.shape[1])

df_vecinos = read_excel(excel_workbook, sheet_name=excel_worksheet_Vecinos)
df_vecinos = df_vecinos[['Beneficiario', 'SMS', 'WhatsApp']]

columns = df_resumen.columns.values.tolist()
last_col = columns.index(datetime(año, mes, 1))

# Imprime el archivo con los resultados del análisis

# Crea el archivo con el análisis
print(f"Creando análisis '{nombre_análisis.format(datetime(año, mes, 1))}'...")
print('')

análisis = ""

mensaje = '* {}, {}{}\n' + \
          '  {}{} usualmente hacia el día {}\n' + \
          '  Su último pago fue el {} y {}{}{}{}{}\n\n'
mensaje_2 = '* {}, {}{}\n' + \
            '  {}{}{}{}\n\n'
rango_fechas_txt = ' entre el {} y el {}'

# Encabezado
am_pm = 'pm' if datetime.now().hour > 12 else 'm' if datetime.now().hour == 12 else 'am'
análisis += f'GyG ANALISIS DE PAGOS, {fecha_análisis}\n({datetime.now():%d/%m/%Y %I:%M} {am_pm})\n\n'

# Resumen
genera_resumen()

# Análisis de pago
análisis += 'Análisis de pago:\n\n'
dirección_anterior = None
for index, r in df_resumen.iterrows():
    if round(r['Promedio'] + r['Variación'], 0) <= 100:   # 100: anteriormente 'hoy'
        min_day = min_d = int(r['Promedio']) - int(round(r['Variación']))
        max_day = max_d = int(r['Promedio']) + int(round(r['Variación']))
        if min_day <= -30:      # dos meses atrás
            fecha_usual = f"{int(r['Promedio'])} ± {int(round(r['Variación']))} días"
            rango_fechas = ','
        else:
            fecha_usual = int(r['Promedio'])
            if min_d <= 0:
                min_day = f'{30 + min_d} del mes anterior'
                max_day = f'{max_d} del mes actual'
            if max_d > 30:
                max_day = f'{max_d - 30} del mes siguiente'
                if min_d > 0:
                    min_day = f'{min_d} del mes actual'
            if min_day == max_day:
                rango_fechas = ""
            else:
                rango_fechas = rango_fechas_txt.format(min_day, max_day)
                if min_d > 0 and max_d <= 30:
                    rango_fechas += ' del mes actual'
                rango_fechas += ','
        v = df_vecinos[df_vecinos['Beneficiario'] == r['Beneficiario']]
        tlf_sms = tlf_wapp = ''
        if not v.empty:
            if notnull(v['SMS']).bool():
                tlf_sms = ', SMS: ' + v['SMS'].to_string(index=False)
            if notnull(v['WhatsApp']).bool():
                tlf_wapp = ', WhatsApp: ' + v['WhatsApp'].to_string(index=False)
        telefonos = tlf_sms + tlf_wapp
        info_pago, info_deuda = no_participa_desde(r)
        if (r['Categoría'] == 'Cuota completa'):
            tipo_pago = f"Paga {r['Categoría'].lower()}"
        else:
            tipo_pago  = "'" + r['Categoría'] + "'. " if r['Categoría'] != 'Colaboración' else ""
            tipo_pago += "Colabora"
        max_fecha = max(df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']]['Fecha'])
        meses_fecha = meses_último_pago(r['Beneficiario'], max_fecha)
        último_pago = '{:%d de %B} ({})'.format(max_fecha, meses_fecha)
        if not(solo_deudores and len(info_deuda) == 0):
            if dirección_anterior == None:
                dirección_anterior = get_street(r['Dirección'])
            if (get_street(r['Dirección']) != dirección_anterior) and not ordenado:
                análisis += '\n'
                dirección_anterior = get_street(r['Dirección'])
            análisis += mensaje.format(
                        r['Beneficiario'], r['Dirección'], telefonos,
                        tipo_pago, rango_fechas, fecha_usual,
                        último_pago,
                        info_pago, info_deuda,
                        mas_de_un_mes(r['Beneficiario']),
                        programa_de_comida(r['Comida']),
                        propuesta_de_cambio(r['Beneficiario']))
    else:
        v = df_vecinos[df_vecinos['Beneficiario'] == r['Beneficiario']]
        tlf_sms = tlf_wapp = ''
        if not v.empty:
            if notnull(v['SMS']).bool():
                tlf_sms = ', SMS: ' + v['SMS'].to_string(index=False)
            if notnull(v['WhatsApp']).bool():
                tlf_wapp = ', WhatsApp: ' + v['WhatsApp'].to_string(index=False)
        telefonos = tlf_sms + tlf_wapp
        if r['Categoría'] == 'No participa':
            tipo_pago = r['Categoría']
        else:
            tipo_pago = 'Paga ' + r['Categoría'].lower()
        if notnull(r['Promedio']):
            detalles_pago = ' y su único pago fue el {:%d de %B de %Y}'.format(
                                max(df_pagos[df_pagos['Beneficiario'] == r['Beneficiario']]['Fecha']))
        else:
            detalles_pago = ' y no se tienen pagos registrados'

        info_pago, info_deuda = no_participa_desde(r)
        info_deuda = re.findall(r'([0-9]+\.[0-9]+)', info_deuda)[0]
        detalles_pago += f"\n  Tiene cuotas pendientes por Bs. {info_deuda}"
        if not(solo_deudores and len(info_deuda) == 0):
            if dirección_anterior == None:
                dirección_anterior = get_street(r['Dirección'])
            if (get_street(r['Dirección']) != dirección_anterior) and not ordenado:
                análisis += '\n'
                dirección_anterior = get_street(r['Dirección'])
            análisis += mensaje_2.format(
                        r['Beneficiario'], r['Dirección'], telefonos,
                        tipo_pago, detalles_pago,
                        programa_de_comida(r['Comida']),
                        propuesta_de_cambio(r['Beneficiario']))
    df_categoría = df_categoría[df_categoría['Beneficiario'] != r['Beneficiario']]

# Otras propuestas de cambio de categoría
if df_categoría.shape[0] != 0:
    análisis += '\n'

    # Imprime las propuestas de cambio de categoría
    análisis += f'Otras propuestas de cambio de categoría en base a los últimos {nMeses} meses:\n\n'

    for index, r in df_categoría.iterrows():
        categoría_actual = "'" + r['Categoría'].lower() + "'"
        propuesta = "'" + r['Propuesta'] + "'"
        análisis += f"* {r['Beneficiario']}. Actualmente {categoría_actual}; se propone cambiar a {propuesta}" + \
                    f"{programa_de_comida(r['Comida'])}\n"


# Graba los archivos de análisis (encoding para Windows y para macOS X)
filename = os.path.join(attach_path, 'Apple', nombre_análisis.format(datetime(año, mes, 1)))
with open(filename, 'w', encoding=GyG_constantes.Apple_encoding) as output:
    output.write(análisis)

filename = os.path.join(attach_path, 'Windows', nombre_análisis.format(datetime(año, mes, 1)))
with open(filename, 'w', encoding=GyG_constantes.Windows_encoding) as output:
    output.write(análisis)
