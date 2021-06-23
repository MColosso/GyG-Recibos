# GYG Actualiza Registro de Pagos Off-Line
#
# Actualiza la hoja de cálculo con el registro de nuevos pagos; especialmente útil en ocasiones
# en las cuales no se dispone de conexión a Internet
#

"""
    PENDIENTE POR HACER
      √ Botón «Cierra» cuando hay pagos sin grabar, agregar botón «Grabar»
         -> Generar ventanas de mensajes en cada ocasión:
              - En el caso del botón «Cierra», incluir botones para «Guarda», «Cancela» y
                «Cierra» -> ícono de «advertencia»
              - En el caso de los botones «Registra nuevo pago» y «Actualiza este pago»,
                incluir botón «Ok» -> ícono de «información»

    NOTAS
      - ¿Los botones "Pago anterior" y "Siguiente pago" deben hacer una búsqueda circular?
         -> No, podría generar un lazo infinito en la búsqueda en caso de ser el primer pago
            de dicho vecino.
      - Al grabar la hoja de cálculo se borran todos los valores asociados a fórmulas, los cuales
        son leidos como NaN en Pandas y como None en OpenPyXl. Estos valores sólo pueden ser
        recuperados grabando la hoja de cálculo nuevamente con Excel, por lo que NO ES
        FACTIBLE utilizar esta versión de la librería OpenPyXl, o esta estrategia de operación,
        hasta que que se subsane esta situación.
         -> Se cambió la estrategia para el registro de pagos: se cuenta con dos hojas de cálculo:
            la hoja original (la cual conserva los valores de las celdas) y una hoja 'off-line' en la
            cual se almacenan los cambios realizados (y que perderá los valores asociados a las
            fórmulas, al ser manejada con OpenPyXl, dado que se requiere tener acceso a las fórmulas)
            Los nombres de columnas, así como los beneficiarios y direcciones en la hoja de resumen,
            son tomados de la hoja de cálculo original; el resto del manejo es hecho sobre la hoja
            de cálculo off-line. √
             -> IMPORTANTE: Esto hace que, para reemplazar la hoja de cálculo original con la versión
                off-line, primeramente se deba recalcular la hoja off-line con Excel.

    HISTORICO
      - Versión inicial (08/07/2021)

"""

print('Cargando librerías...')
import GyG_constantes
from GyG_cuotas import *
from GyG_utilitarios import *

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from pandas import read_excel, isnull, notnull
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from PIL import ImageTk, Image

import sys
import re
import string
from unicodedata import normalize
from shutil import copyfile
import locale

dummy = locale.setlocale(locale.LC_ALL, 'es_es')


#
# CONSTANTES ============================================================================
# 

excel_workbook_std = GyG_constantes.pagos_wb_estandar             # '1.1. GyG Recibos.xlsm'
excel_workbook     = "1.1. GyG Recibos OFF-LINE.xlsm"
hoja_resumen_reord = GyG_constantes.pagos_ws_resumen_reordenado   # 'R.VIGILANCIA (reordenado)'
hoja_resumen       = GyG_constantes.pagos_ws_resumen              # 'RESUMEN VIGILANCIA'
hoja_vigilancia    = GyG_constantes.pagos_ws_vigilancia           # 'Vigilancia'

_índice_vecinos    = None
_índice_inverso    = None
_índice_dirección  = None

background_color   = '#EDEDED'

SOLO_VECINOS_ACTIVOS = False
DISTRIB_N_ROWS     = 2
DISTRIB_N_COLS     = 5
FORMATO_MES        = "%m-%Y"

lst_labels = list()
lst_entries = list()

pagos_por_guardar = False

DEBUG = False


#
# === RUTINAS UTILITARIAS ===============================================================
#

def convierte_en_float(valor: str) -> float:
    valor = valor.replace(".", "").replace(",", ".")
    try:
        valor = float(valor)
    except ValueError:
        # valor = None
        pass
    return valor


def género(de_la, beneficiario):
    if de_la == 'de':
        if any(b in beneficiario for b in ['Familia', 'Sra.', 'Dra.']):
            return f'de la {beneficiario}'
        elif any(b in beneficiario for b in ['Sr.', 'Dr.']):
            return f'del {beneficiario}'
        else:
            return f'de {beneficiario}'
    elif de_la == 'la':
        if any(b in beneficiario for b in ['Familia', 'Sra.', 'Dra.']):
            return f'la {beneficiario}'
        elif any(b in beneficiario for b in ['Sr.', 'Dr.']):
            return f'el {beneficiario}'
        else:
            return f'{beneficiario}'
    else:
        return f"{de_la} {beneficiario}"

#
# ------------------------------------- Rutinas para el manejo de la lista de vecinos ---
#

def analiza_texto(texto: str) -> list:
    STOPWORDS = ['familia', 'calle', 'av', 'avenida', 'numero', 'nro', 'numeros', 'nros', 'y', 'el']
    CVT_SEQUENCES = {'z': 's', 'h': '',
                     'ss': 's', 'cc': 'c', 'll': 'l', 'tt': 't', 'nn': 'n', 'bb': 'b'}

    def remueve_acentos(texto: str) -> list:
        # -> NFD y eliminar diacríticos
        s = re.sub(
                r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", 
                normalize('NFD', texto), 0, re.I
            )
        # -> NFC
        return normalize('NFC', s)

    def remueve_stopwords(lista: list) -> list:
        return [token for token in lista if token not in STOPWORDS]

    def elimina_secuencias(texto: str) -> str:
        for key, value in CVT_SEQUENCES.items():
            texto = texto.replace(key, value)
        return texto

    # Remueve acentos, convierte a minúsculas, elimina signos de puntuación, convierte 'z's en 's's,
    # elimina las 'h's y convierte las secuencias dobles (s, c, l, t, n) en sencillas
    texto = remueve_acentos(texto) \
        .lower() \
        .translate(str.maketrans('', '', string.punctuation))
    lista = elimina_secuencias(texto) \
        .split()
    return remueve_stopwords(lista)


def indexa_listado_de_vecinos() -> list:
    global _índice_vecinos, _índice_inverso, _índice_dirección

    último_día_del_mes = datetime.today() + relativedelta(months=1, day=1) - relativedelta(days=1)

    # Abre la hoja de cálculo de Recibos de Pago
    df_worksheet = read_excel(excel_workbook_std, sheet_name=hoja_resumen_reord)

    # Elimina los registros que no tienen una categoría definida
    df_worksheet.dropna(subset=['Categoría'], inplace=True)

    # Elimina aquellos vecinos que compraron (o se iniciaron) en fecha posterior
    # o que vendieron (o cambiaron su razón social) en fecha anterior a la fecha
    # de análisis
    if SOLO_VECINOS_ACTIVOS:
        # Ajusta las columnas F.Desde y F.Hasta en aquellos en los que estén vacíos:
        # 01/01/2016 y 01/mes_siguiente/año
        df_worksheet.loc[df_worksheet[isnull(df_worksheet['F.Desde'])].index, 'F.Desde'] = datetime(2016, 1, 1)
        df_worksheet.loc[df_worksheet[isnull(df_worksheet['F.Hasta'])].index, 'F.Hasta'] = \
                                                                     último_día_del_mes + relativedelta(days=1)
        df_worksheet = df_worksheet[
                            (df_worksheet['F.Desde'] < último_día_del_mes) & \
                            (df_worksheet['F.Hasta'] >= último_día_del_mes)
                       ]

    # Cambia el nombre de la columna 2016 a datetimme(2016, 1, 1)
    df_worksheet.rename(columns={2016:datetime(2016, 1, 1)}, inplace=True)

    # Selecciona sólo las columnas a utilizar
    df_worksheet = df_worksheet[['Beneficiario', 'Dirección']]

    # Prepara lista de vecinos y el índice inverso del mismo
    _índice_vecinos = list()
    _índice_inverso = dict()
    _índice_dirección = list()

    vecino_ID = 0
    for _, r in df_worksheet.iterrows():
        texto_a_analizar = ''.join([
            r['Beneficiario'],
            '. ',
            r['Dirección']
        ])
        for token in analiza_texto(texto_a_analizar):
            if token not in _índice_inverso:
                _índice_inverso[token] = set()
            _índice_inverso[token].add(vecino_ID)
        vecino = r['Beneficiario']
        if vecino not in _índice_vecinos:
            _índice_vecinos.append(vecino)
            _índice_dirección.append(r['Dirección'])
            vecino_ID += 1


def busca_en_índices(texto: str, modo: str="AND", dirección=False) -> list:
    lista_generada = analiza_texto(texto)
    resultado = None
    for token in lista_generada:
        if token in _índice_inverso:
            if resultado is None:
                resultado = _índice_inverso[token]
            elif modo == "AND":
                resultado = resultado.intersection(_índice_inverso[token])
            elif modo == "OR":
                resultado.union(_índice_inverso[token])
            elif modo == "Smart":
                if resultado.intersection(_índice_inverso[token]) == set():
                    resultado = resultado.union(_índice_inverso[token])
                else:
                    resultado = resultado.intersection(_índice_inverso[token])
            else:
                raise ValueError(f"busca_en_indices(): error en modo: {modo}")

    if resultado:
        if dirección:
            resultado = [_índice_vecinos[idx] + ', ' + _índice_dirección[idx] for idx in resultado]
        else:
            resultado = [_índice_vecinos[idx] for idx in resultado]

    elif resultado == set():
        resultado = None

    return resultado


def maneja_búsqueda(a_buscar: str):
    lista_vecinos = list()
    resultado = busca_en_índices(a_buscar, modo="Smart", dirección=False)
    if resultado:
        for vecino in resultado:
            lista_vecinos.append(vecino)
    return lista_vecinos


#
# ------------------------ Rutinas para el manejo de la presentación de la aplicación ---
#

def actualiza_lista_de_vecinos():
    vecino = cmbVecinos.get()
    cmbVecinos["values"] = maneja_búsqueda(vecino)


def habilita_busqueda_de_pagos(event):
    global último_renglón

    btnBusca_anterior.config(state=tk.NORMAL, text="Busca último pago")
    btnBusca_siguiente.config(state=tk.DISABLED)
    btnDistribución.config(state=tk.DISABLED)
    # "+ 1" para compensar el "- 1" en btnBuscaAnterior_presionado()
    último_renglón = ws_vigilancia.max_row + 1
    reinicializa_renglon()


def btnBuscaAnterior_presionado():
    global último_renglón, iterador, sentido_ascendente, cambio_sentido

    btnBusca_anterior["text"] = "<< Pago anterior"
    btnDistribución.config(state=tk.NORMAL)
    btnBusca_siguiente.config(state=tk.NORMAL)
    sentido_ascendente = False
    # cambio_sentido = True
    iterador = reversed(benef_pagos[:último_renglón - 1])
    if DEBUG: print(f"DEBUG: btnBuscaAnterior_presionado(), último renglón = {último_renglón}")
    if DEBUG: print(f"        -> iterador = reversed(benef_pagos[:{último_renglón} - 1]), len = {len(benef_pagos[:último_renglón - 1])}")
    busca_vecino()


def btnBuscaSiguiente_presionado():
    global último_renglón, iterador, sentido_ascendente, cambio_sentido

    sentido_ascendente = True
    # if cambio_sentido:
    #     # último_renglón += 1
    #     cambio_sentido = False
    # iterador = benef_pagos[último_renglón + 1:]
    iterador = benef_pagos[último_renglón:]
    if DEBUG: print(f"DEBUG: btnBuscaSiguiente_presionado(), último renglón = {último_renglón}, len = {len(iterador)}")
    if DEBUG: print(f"        -> iterador = rbenef_pagos[{último_renglón}+ 1:]")
    busca_vecino()


def reinicializa_renglon():
    entry_vars = [var_Beneficiario, var_Direccion, var_Email, var_Fecha, var_Monto, var_Monto_USD,
                 var_Concepto, var_Categoria]
    check_vars = [var_Generar, var_Enviado, var_chk_USD]
    for var in entry_vars:
        var.delete(0, tk.END)
    for var in check_vars:
        var.set(False)
    for cell in range(DISTRIB_N_ROWS * DISTRIB_N_COLS):
        lst_labels[cell]["text"] = ""
        lst_entries[cell].delete(0, tk.END)
        lst_entries[cell].configure(relief="flat", bg="gainsboro", state=tk.DISABLED)

    # Deshabilita los botones para grabar el pago
    btnActualizaPagoActual.configure(state=tk.DISABLED)
    btnGrabaNuevoPago.configure(state=tk.DISABLED)


def busca_vecino():
    global último_renglón

    # Busca el último pago realizado por el vecino indicado, actualizando las fórmulas en caso
    # de ser necesario
    renglón_buscado = None
    nuevo_renglón = len(benef_pagos) + 1
    reinicializa_renglon()

    for x in iterador:
        if x.value == cmbVecinos.get():
            renglón_buscado = x.row
            for idx, col_name in enumerate(columns_pagos):
                from_cell = f"{get_column_letter(idx+1)}{renglón_buscado}"
                to_cell = f"{get_column_letter(idx+1)}{nuevo_renglón}"
                # if cell content is a formula, asume it as relative addressing, else, copy it directly
                if isinstance(ws_vigilancia[from_cell].value, str) and ws_vigilancia[from_cell].value[0] == '=':
                    valor_celda = Translator(ws_vigilancia[from_cell].value, origin=from_cell) \
                                                .translate_formula(to_cell)
                else:
                    valor_celda = ws_vigilancia[from_cell].value
                if valor_celda is None:
                    valor_celda = ''
                # print(f"           {col_name+':':20} {valor_celda}")
                if col_name == 'Beneficiario':
                    var_Beneficiario.insert(0, valor_celda)
                elif col_name == 'Dirección':
                    var_Direccion.insert(0, valor_celda)
                elif col_name == 'E-mail o celular':
                    var_Email.insert(0, valor_celda)
                elif col_name == 'Fecha':
                    var_Fecha.insert(0, valor_celda.strftime("%d-%m-%Y"))
                elif col_name == 'Monto':
                    var_Monto.insert(0, valor_celda if es_fórmula(valor_celda) else edita_número(valor_celda))
                elif col_name == 'Monto $':
                    var_Monto_USD.insert(0, valor_celda if es_fórmula(valor_celda) else edita_número(valor_celda))
                elif col_name == 'Concepto':
                    var_Concepto.insert(0, valor_celda)
                elif col_name == 'Categoría':
                    var_Categoria.insert(0, valor_celda)
                elif col_name == 'Generar':
                    var_Generar.set(bool(valor_celda))
                elif col_name == 'Enviado':
                    var_Enviado.set(bool(valor_celda))
                elif col_name == '$':
                    var_chk_USD.set(bool(valor_celda))
            último_renglón = renglón_buscado
            break
    if DEBUG: print(f"        -> último renglón = {último_renglón}")

def btnDistribución_presionado():
    global lista_meses

    # Reinicia todas las celdas de la distribución
    for cell in range(DISTRIB_N_ROWS * DISTRIB_N_COLS):
        lst_labels[cell]["text"] = ""
        lst_entries[cell].delete(0, tk.END)
        lst_entries[cell].configure(relief="flat", bg="gainsboro", state=tk.DISABLED)

    # Muestra desde el mes anterior hasta el mes siguiente a los indicados en el Concepto
    # del pago
    lista_meses = separa_meses(var_Concepto.get())
    if lista_meses and var_Categoria.get() == GyG_constantes.CATEGORIA_VIGILANCIA:
        d = datetime.strptime(lista_meses[0], FORMATO_MES)
        lista_meses.insert(0, (d - relativedelta(months=1)).strftime(FORMATO_MES))
        d = datetime.strptime(lista_meses[-1], FORMATO_MES)
        lista_meses.append((d + relativedelta(months=1)).strftime(FORMATO_MES))

        # Agrega sólo las celdas relacionadas
        row = benef_resumen.index(var_Beneficiario.get()) + 1
        for cell, value in enumerate(lista_meses):
            col = columns_resumen.index(value) + 1
            to_cell = f"{get_column_letter(col)}{row}"
            lst_labels[cell]["text"] = datetime.strptime(value, FORMATO_MES).strftime("%b %Y")
            lst_entries[cell].configure(state=tk.NORMAL, bg="white", relief="groove")
            if ws_resumen[to_cell].value is not None:
                lst_entries[cell].insert(0, edita_número(ws_resumen[to_cell].value))

    # Habilita los botones para grabar el pago
    btnActualizaPagoActual.configure(state=tk.NORMAL)
    btnGrabaNuevoPago.configure(state=tk.NORMAL)


def cambio_en_categoria(var, indx, mode):
    # texto = "Distribución del pago" if str_Categoria.get() == "Vigilancia" else "Activa botón para guardar"
    # btnDistribución.configure(text=texto)

    if str_Categoria.get() == GyG_constantes.CATEGORIA_VIGILANCIA:
        btnDistribución.configure(state=tk.NORMAL)
        btnActualizaPagoActual.configure(state=tk.DISABLED)
        btnGrabaNuevoPago.configure(state=tk.DISABLED)
    else:
        btnDistribución.configure(state=tk.DISABLED)
        btnActualizaPagoActual.configure(state=tk.NORMAL)
        btnGrabaNuevoPago.configure(state=tk.NORMAL)

        for cell in range(DISTRIB_N_ROWS * DISTRIB_N_COLS):
            lst_labels[cell]["text"] = ""
            lst_entries[cell].delete(0, tk.END)
            lst_entries[cell].configure(relief="flat", bg="gainsboro", state=tk.DISABLED)


def btnActualiza_presionado():
    # Actualiza el pago actual
    actualiza_hoja_de_cálculo(renglón_a_copiar=último_renglón, nuevo_renglón=último_renglón, 
                              este_renglón=True)


def btnGraba_presionado():
    global benef_pagos, último_renglón

    # Genera un nuevo registro de pago
    actualiza_hoja_de_cálculo(renglón_a_copiar=último_renglón, nuevo_renglón=ws_vigilancia.max_row + 1,
                              este_renglón=False)

    # Anexa el vecino a la lista de pagos
    benef_pagos.append(ws_vigilancia[f"{column('Beneficiario')}{ws_vigilancia.max_row}"])
    último_renglón = ws_vigilancia.max_row


def valida_campos_Ok() -> bool:
    resultado = True

    # Valida Fecha
    try:
        fecha = datetime.strptime(var_Fecha.get(), "%d-%m-%Y")
        # Restaura formato del campo Fecha
        var_Fecha['foreground'] = 'black'
    except ValueError:
        resultado = False
        # Cambia formato del campo Fecha
        var_Fecha['foreground'] = 'red'

    # Valida Monto
    monto = var_Monto.get()
    if not(isinstance(convierte_en_float(monto), float) or \
            isinstance(evalúa_fórmula(monto), float)):
        resultado = False
        # Cambia formato del campo Monto
        var_Monto['foreground'] = 'red'
    else:
        # Restaura formato del campo Monto
        var_Monto['foreground'] = 'black'

    # Valida Monto en US$
    monto = var_Monto_USD.get()
    if len(monto) > 0:
        if not(isinstance(convierte_en_float(monto), float) or \
                isinstance(evalúa_fórmula(monto), float)):
            resultado = False
            # Cambia formato del campo Monto en US$
            var_Monto_USD['foreground'] = 'red'
        else:
            # Restaura formato del campo Monto
            var_Monto_USD['foreground'] = 'black'

    # Valida distribución del pago
    for celda in lst_entries:
        monto = celda.get()
        if monto is not None and len(monto) > 0:
            if not(isinstance(convierte_en_float(monto), float) or \
                    isinstance(evalúa_fórmula(monto), float)):
                resultado = False
                # Cambia formato de la celda actual en la distribución del pago
                celda['foreground'] = 'red'
            else:
                # Restaura formato de la celda
                celda['foreground'] = 'black'

    return resultado


def actualiza_hoja_de_cálculo(renglón_a_copiar: int, nuevo_renglón: int, este_renglón: bool):
    global pagos_por_guardar, var_Categoria, lista_categorías

    if not valida_campos_Ok():
        return

    # Hoja VIGILANCIA
    if renglón_a_copiar != nuevo_renglón:
        for col in range(ws_vigilancia.max_column):
            from_cell = f"{get_column_letter(col+1)}{renglón_a_copiar}"
            to_cell = f"{get_column_letter(col+1)}{nuevo_renglón}"
            # if cell content is a formula, asume it as relative addressing, else, copy it directly
            if es_fórmula(ws_vigilancia[from_cell].value):
                ws_vigilancia[to_cell] = Translator(ws_vigilancia[from_cell].value, origin=from_cell) \
                                            .translate_formula(to_cell)
            else:
                ws_vigilancia[to_cell] = ws_vigilancia[from_cell].value
            if ws_vigilancia[from_cell].has_style:
                ws_vigilancia[to_cell]._style = ws_vigilancia[from_cell]._style

    ws_vigilancia[f"{column('Generar')}{nuevo_renglón}"] = 'ü' if var_Generar.get() else None
    ws_vigilancia[f"{column('Enviado')}{nuevo_renglón}"] = 'ü' if var_Enviado.get() else None
    ws_vigilancia[f"{column('Fecha')}{nuevo_renglón}"] = datetime.strptime(var_Fecha.get(), "%d-%m-%Y")
    ws_vigilancia[f"{column('Monto')}{nuevo_renglón}"] = convierte_en_float(var_Monto.get())
    ws_vigilancia[f"{column('Beneficiario')}{nuevo_renglón}"] = var_Beneficiario.get()
    ws_vigilancia[f"{column('E-mail o celular')}{nuevo_renglón}"] = var_Email.get()
    ws_vigilancia[f"{column('Dirección')}{nuevo_renglón}"] = var_Direccion.get()
    ws_vigilancia[f"{column('Concepto')}{nuevo_renglón}"] = var_Concepto.get()
    ws_vigilancia[f"{column('Categoría')}{nuevo_renglón}"] = var_Categoria.get()
    ws_vigilancia[f"{column('Monto $')}{nuevo_renglón}"] = convierte_en_float(var_Monto_USD.get())
    ws_vigilancia[f"{column('$')}{nuevo_renglón}"] = 'ü' if var_chk_USD.get() else None

    # Hoja RESUMEN VIGILANCIA
    if var_Categoria.get() == GyG_constantes.CATEGORIA_VIGILANCIA:
        row_resumen = benef_resumen.index(var_Beneficiario.get()) + 1
        lista_celdas = [f"{get_column_letter(columns_resumen.index(mes) + 1)}{row_resumen}" for mes in lista_meses]

        for celda, contenido in zip(lista_celdas, lst_entries):
            valor = contenido.get() if es_fórmula(contenido.get()) else convierte_en_float(contenido.get())
            # print(f"{celda}: {valor}, ")
            if valor:
                ws_resumen[celda] = valor

    cell_Enviado = ws_vigilancia[f"{column('Enviado')}{renglón_a_copiar}"]
    if cell_Enviado.has_style:
        ws_vigilancia[f"{column('Generar')}{nuevo_renglón}"]._style = cell_Enviado._style
        ws_vigilancia[f"{column('$')}{nuevo_renglón}"]._style = cell_Enviado._style

    # Actualiza tabla de categorías
    categoria = var_Categoria.get()
    if categoria not in lista_categorías:
        lista_categorías.insert(0, categoria)
        var_Categoria["values"] = lista_categorías

    # Guarda
    pagos_por_guardar = True
    btnGuarda.configure(text= 'GUARDA', state=tk.NORMAL)
    último_renglón = nuevo_renglón

    # messagebox.showinfo(message=f"El pago de {var_Beneficiario.get()} fue actualizado" if este_renglón \
    #                             else f"Se agregó un nuevo pago para {var_Beneficiario.get()}")
    ventana_información(f"El pago {género('de', var_Beneficiario.get())} fue actualizado" if este_renglón \
                        else f"Se agregó un nuevo pago para {género('la', var_Beneficiario.get())}")


def btnGuarda_presionado():
    global pagos_por_guardar

    wb_pagos.save(excel_workbook)
    btnGuarda.configure(text='Guarda', state=tk.DISABLED)
    pagos_por_guardar = False
    print(f'"{excel_workbook}" guardado...')
    # messagebox.showinfo(message=''.join([
    #     f'Grabada hoja de cálculo "{excel_workbook}"...',
    #     "\n\n",
    #     "Recuerde que para reemplazar la hoja de cálculo original con la versión off-line, ",
    #     "primeramente se debe recalcular la hoja off-line con Excel."
    # ]))
    ventana_información(''.join([
        f'Grabada hoja de cálculo "{excel_workbook}"...',
        "\n\n",
        "Recuerde que para reemplazar la hoja de cálculo original con la versión off-line, ",
        "primeramente se debe recalcular la hoja off-line con Excel."
    ]))


def btnCierra_presionado():
    # if pagos_por_guardar and \
    #         not messagebox.askokcancel(title='Salir sin guardar',
    #                                    message='Hay pagos pendientes por guardar.\n¿Desea salir sin guardar?',
    #                                    icon=None):
    if pagos_por_guardar and \
            not ventana_advertencia(mensaje='Hay pagos pendientes por guardar.\n¿Desea salir sin guardarlos?',
                                    comando=btnGuarda_presionado):
        pass
    else:
        root.quit()


def ventana_información(mensaje):
    global img      # This is to avoid garbage collector delete it

    child_info = tk.Toplevel(root, background=background_color)
    child_info.title("GyG Información")
    # child_info.geometry("320x100+150+150")    # width x height + x position + y position
    child_info.geometry("+200+200")           # x position + y position

    topFrame = ttk.Frame(child_info)
    topFrame.grid(row=0, column=0, sticky=tk.NSEW)      # padx=5, pady=5,

    info = os.path.join('./recursos/imagenes', 'info.png')
    img = ImageTk.PhotoImage(Image.open(info).resize((30, 30), Image.ANTIALIAS))
    imagen = ttk.Label(topFrame, image=img, width=10)
    imagen.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.E)

    tk.Message(topFrame, text=mensaje, font=("Helvetica", 14), width=250, background=background_color) \
        .grid(row=0, column=1, pady=10, sticky=tk.N+tk.E)

    bottomFrame = ttk.Frame(child_info)
    bottomFrame.grid(row=1, column=0, padx=10, pady=10, sticky=tk.NSEW)

    ttk.Label(bottomFrame, text='', width=20).grid(row=0, column=0, sticky=tk.S)
    btnOK = ttk.Button(bottomFrame, text='Ok', width=10, command=child_info.destroy)
    btnOK.grid(row=0, column=1, sticky=tk.E+tk.S)


def ventana_advertencia(mensaje, comando):
    global img      # This is to avoid garbage collector delete it

    def command_wrap():
        comando()
        root.destroy()

    child_warn = tk.Toplevel(root, background=background_color)
    child_warn.title("GyG Advertencia")
    # child_warn.geometry("320x100+150+150")    # width x height + x position + y position
    child_warn.geometry("+200+200")           # x position + y position

    topFrame = ttk.Frame(child_warn)
    topFrame.grid(row=0, column=0, sticky=tk.NSEW)

    advertencia = os.path.join('./recursos/imagenes', 'warn.png')
    img = ImageTk.PhotoImage(Image.open(advertencia).resize((30, 30), Image.ANTIALIAS))
    imagen = ttk.Label(topFrame, image=img, width=10)
    imagen.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.E)

    tk.Message(topFrame, text=mensaje, font=("Helvetica", 14), width=250, background=background_color) \
        .grid(row=0, column=1, pady=10, sticky=tk.N+tk.E+tk.W)

    bottomFrame = ttk.Frame(child_warn)
    bottomFrame.grid(row=1, column=0, padx=10, pady=10, sticky=tk.NSEW)

    ttk.Label(bottomFrame, text='', width=5).grid(row=0, column=0)
    btnGuarda = ttk.Button(bottomFrame, text='Guarda', command=command_wrap)
    btnGuarda.grid(row=0, column=1, padx=2, sticky=tk.W)
    btnCancela = ttk.Button(bottomFrame, text='Cancela', command=child_warn.destroy)
    btnCancela.grid(row=0, column=2, padx=2, sticky=tk.W)
    btnOK = ttk.Button(bottomFrame, text='Ok', command=root.destroy)
    btnOK.grid(row=0, column=3, padx=2, sticky=tk.W)


def btnListaPagos_presionado():
    NRO_RENGLONES_SELECCIONADOS = 20
    NRO_RENGLONES_A_MOSTRAR = 8

    child_w = tk.Toplevel(root)
    child_w.title("GyG Últimos pagos registrados")
    child_w.geometry("795x200+150+150")    # width x height + x position + y position

    columnas = ('Recibo', 'Fecha', 'Monto', 'Beneficiario', 'Dirección', 'Concepto', 'Categoría')
    anchuras = (50, 80, 100, 140, 120, 180, 120)
    alineaciones = (tk.CENTER, tk.CENTER, tk.E, tk.W, tk.W, tk.W, tk.W)
    últimos_pagos = ttk.Treeview(child_w, columns=columnas, height=NRO_RENGLONES_A_MOSTRAR, show='headings')

    # títulos de columnas
    for columna, ancho, alineación in zip(columnas, anchuras, alineaciones):
        últimos_pagos.heading(columna, text=columna)
        últimos_pagos.column(columna, minwidth=0, width=ancho, anchor=alineación, stretch=tk.NO)
    últimos_pagos.grid(row=0, column=0, columnspan=2)
    últimos_pagos_sb = ttk.Scrollbar(child_w, orient="vertical", command=últimos_pagos.yview)
    últimos_pagos_sb.grid(row=0, column=1, sticky=tk.S + tk.E + tk.N)
    últimos_pagos.configure(yscrollcommand=últimos_pagos_sb.set)

    # # -----------------------------------
    # #Creates a Tkinter-compatible photo image, which can be used everywhere Tkinter expects an image object.
    # advertencia = os.path.join('./recursos/imagenes', 'warn.png')
    # img = ImageTk.PhotoImage(Image.open(advertencia).resize((50, 50), Image.ANTIALIAS))
    # # img = Image.open(advertencia).resize((50, 50), Image.ANTIALIAS)

    # #The Label widget is a standard Tkinter widget used to display a text or image on the screen.
    # imagen = ttk.Label(child_w, image=img, width=60)
    # imagen.grid(row=1, column=0, sticky=tk.N+tk.W)
    # print(imagen.__doc__)
    # # panel.configure(image=img)
    # # -----------------------------------

    btnCierraListado = ttk.Button(child_w, text='Cierra', command=child_w.destroy)
    btnCierraListado.grid(row=1, column=1, sticky=tk.N+tk.E)

    num_renglones = ws_vigilancia.max_row
    for recibo in range(num_renglones - NRO_RENGLONES_SELECCIONADOS + 1, num_renglones + 1):
        monto = evalúa_fórmula(ws_vigilancia[f"{column('Monto')}{recibo}"].value)
        renglón = [
            f"{(recibo - 1):0{GyG_constantes.long_num_recibo}}",
            ws_vigilancia[f"{column('Fecha')}{recibo}"].value.strftime("%d/%m/%Y"),
            # alinea_texto(edita_número(monto, num_decimals=0), anchura=15),
            edita_número(monto, num_decimals=0),
            edita_beneficiario(ws_vigilancia[f"{column('Beneficiario')}{recibo}"].value),
            edita_dirección(ws_vigilancia[f"{column('Dirección')}{recibo}"].value),
            reagrupa_meses(ws_vigilancia[f"{column('Concepto')}{recibo}"].value),
            ws_vigilancia[f"{column('Categoría')}{recibo}"].value
        ]
        últimos_pagos.insert("", "end", iid=recibo, values=renglón)

    últimos_pagos.see(num_renglones)


# ---------------------------------------------------------------------------------
# Código tomado de: Scrollable Frames in Tkinter
#                   https://blog.teclado.com/tkinter-scrollable-frames/
# ---------------------------------------------------------------------------------
# class ScrollableFrame(ttk.Frame):
#     def __init__(self, container, *args, **kwargs):
#         super().__init__(container, *args, **kwargs)
#         canvas = tk.Canvas(self)
#         scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
#         self.scrollable_frame = ttk.Frame(canvas)
#
#         self.scrollable_frame.bind(
#             "<Configure>",
#             lambda e: canvas.configure(
#                 scrollregion=canvas.bbox("all")
#             )
#         )
#
#         canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
#
#         canvas.configure(yscrollcommand=scrollbar.set)
#
#         canvas.pack(side="left", fill="both", expand=True)
#         scrollbar.pack(side="right", fill="y")
# ---------------------------------------------------------------------------------

# ---------------------------------------------------------------------------------
# Código tomado de: How to know all style options of a ttk widget
#                   https://stackoverflow.com/questions/45389166/how-to-know-all-style-options-of-a-ttk-widget
# ---------------------------------------------------------------------------------
# def stylename_elements_options(stylename):
#     '''Function to expose the options of every element associated to a widget
#        stylename.'''
#     try:
#         # Get widget elements
#         style = ttk.Style()
#         layout = str(style.layout(stylename))
#         print('Stylename = {}'.format(stylename))
#         print('Layout    = {}'.format(layout))
#         elements=[]
#         for n, x in enumerate(layout):
#             if x=='(':
#                 element=""
#                 for y in layout[n+2:]:
#                     if y != ',':
#                         element=element+str(y)
#                     else:
#                         elements.append(element[:-1])
#                         break
#         print('\nElement(s) = {}\n'.format(elements))
#
#         # Get options of widget elements
#         for element in elements:
#             print('{0:30} options: {1}'.format(
#                 element, style.element_options(element)))
#
#     except tk.TclError:
#         print('_tkinter.TclError: "{0}" in function'
#               'widget_elements_options({0}) is not a regonised stylename.'
#               .format(stylename))
# ---------------------------------------------------------------------------------

def initUI():
    global var_NroRecibo, var_Generar, var_Enviado, var_Fecha, var_Monto
    global var_Beneficiario, var_Email, var_Direccion, var_Concepto
    global var_Categoria, var_Monto_USD, var_chk_USD
    global cmbVecinos, btnBusca_anterior, btnBusca_siguiente, btnDistribución
    global btnActualizaPagoActual, btnGrabaNuevoPago, btnGuarda, str_Categoria

    root = tk.Tk()
    root.title("GyG Registra pagos")
    # root.geometry("790x490+100+100")    # width x height + x position + y position
    root.geometry("+100+100")           # x position + y position

    mainFrame = ttk.Frame(root)
    mainFrame.grid(row=0, column=0, sticky=tk.N+tk.S+tk.E+tk.W)

    style = ttk.Style()
    style.map("TButton",
        foreground=[('pressed', 'white'), ('active', 'blue')],
        background=[('!disabled', 'black'), ('active', 'white')]
    )
    # style.configure("TFrame",
    #     background=background_color)

    # style.map("TCombobox",
    #     # border=2,
    #     textarea=[('relief', "groove")]
    # )

    ttk.Style().configure('TCombobox', relief='groove')

    # -----------------------------------------------------------------------------------
    headFrame = ttk.Frame(mainFrame)
    headFrame.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W+tk.E)

    ttk.Label(headFrame, text="Registra Pagos Off-Line",
                         font=("Helvetica", 16)).grid(row=0, column=0, sticky=tk.W)

    ttk.Label(headFrame, text="", width=53).grid(row=0, column=1)

    btnListaPagos = ttk.Button(headFrame, text='Últimos pagos', command=btnListaPagos_presionado)
    btnListaPagos.grid(row=0, column=2, sticky=tk.E)

    # -----------------------------------------------------------------------------------

    topFrame = ttk.Frame(mainFrame)
    topFrame.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W+tk.E)

    ttk.Label(topFrame, text="Vecino con pago a registrar").grid(row=0, column=0, sticky=tk.W)

    cmbVecinos = ttk.Combobox(topFrame, width=25, postcommand=actualiza_lista_de_vecinos)
    cmbVecinos.grid(row=1, column=0, ipadx=5)
    cmbVecinos.bind("<<ComboboxSelected>>", habilita_busqueda_de_pagos)

    btnBusca_anterior = ttk.Button(topFrame, text="Busca último pago", state=tk.DISABLED,
                                             command=btnBuscaAnterior_presionado)
    btnBusca_anterior.grid(row=1, column=1, padx=5, ipadx=10, sticky=tk.W)

    btnBusca_siguiente = ttk.Button(topFrame, text="Siguiente pago >>", state=tk.DISABLED,
                                              command=btnBuscaSiguiente_presionado)
    btnBusca_siguiente.grid(row=1, column=2, padx=5, ipadx=10, sticky=tk.W)

    ttk.Label(topFrame, text="").grid(row=2, column=0, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    middleFrame = ttk.Frame(mainFrame)
    middleFrame.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W+tk.E)

    ttk.Label(middleFrame, text="Vecino").grid(row=3, column=0, sticky=tk.W)
    ttk.Label(middleFrame, text="Dirección").grid(row=3, column=1, sticky=tk.W)
    ttk.Label(middleFrame, text="E-mail o teléfono").grid(row=3, column=2, sticky=tk.W)

    var_Beneficiario = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Beneficiario.grid(row=4, column=0, padx=2, sticky=tk.W)
    var_Direccion = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Direccion.grid(row=4, column=1, padx=2, sticky=tk.W)
    var_Email = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Email.grid(row=4, column=2, padx=2, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    var_Generar = tk.IntVar()
    var_Enviado = tk.IntVar()
    var_chk_USD = tk.IntVar()
    ttk.Label(middleFrame, text="Opciones").grid(row=3, column=3)
    ttk.Checkbutton(middleFrame, text="A generar", variable=var_Generar).grid(row=4, column=3, padx=5, sticky=tk.W)
    ttk.Checkbutton(middleFrame, text="Enviado", variable=var_Enviado).grid(row=5, column=3, padx=5, sticky=tk.W)
    ttk.Checkbutton(middleFrame, text="Recibo en US$", variable=var_chk_USD).grid(row=6, column=3, padx=5, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    ttk.Label(middleFrame, text="Fecha").grid(row=5, column=0, sticky=tk.W)
    ttk.Label(middleFrame, text="Monto en Bs.").grid(row=5, column=1, sticky=tk.W)
    ttk.Label(middleFrame, text="Monto en US$").grid(row=5, column=2, sticky=tk.W)

    var_Fecha = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Fecha.grid(row=6, column=0, padx=2, sticky=tk.W)
    var_Monto = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Monto.grid(row=6, column=1, padx=2, sticky=tk.W)
    var_Monto_USD = tk.Entry(middleFrame, width=23, bd=2, relief="groove")
    var_Monto_USD.grid(row=6, column=2, padx=2, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    ttk.Label(middleFrame, text="Concepto").grid(row=8, column=0, columnspan=2, sticky=tk.W)
    ttk.Label(middleFrame, text="Categoría").grid(row=8, column=2, sticky=tk.W)

    str_Categoria = tk.StringVar()
    var_Concepto = tk.Entry(middleFrame, width=48, bd=2, relief="groove")
    var_Concepto.grid(row=9, column=0, columnspan=2, sticky=tk.W)
    # var_Categoria = tk.Entry(middleFrame, textvariable=str_Categoria, width=23, bd=2, relief="groove")
    var_Categoria = ttk.Combobox(middleFrame, width=23, textvariable=str_Categoria)
    var_Categoria.grid(row=9, column=2, ipadx=5, sticky=tk.W)
    var_Categoria["values"] = lista_categorías
    str_Categoria.trace_add('write', cambio_en_categoria)
    # var_Categoria.trace("w", lambda name, index, mode, sv=var_Categoria: cambio_en_categoria(var_Categoria))

    ttk.Label(middleFrame, text="").grid(row=10, column=0, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    btnDistribución = ttk.Button(middleFrame, text="Distribución del pago", state=tk.DISABLED,
                                              command=btnDistribución_presionado)
    btnDistribución.grid(row=12, column=0, pady=5, ipadx=10, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    # scrollableFrame = ScrollableFrame(mainFrame, height=5)
    scrollableFrame = ttk.Frame(mainFrame, height=5)
    scrollableFrame.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W+tk.E)

    for cell in range(DISTRIB_N_ROWS * DISTRIB_N_COLS):
        row = (cell // DISTRIB_N_COLS) * 2
        col = cell % DISTRIB_N_COLS
        lst_labels.append(ttk.Label(scrollableFrame, text=""))
        lst_labels[-1].grid(row=row, column=col, sticky=tk.W)
        lst_entries.append(tk.Entry(scrollableFrame, width=15, bd=2, bg="lightgray", relief="flat", state=tk.DISABLED))
        lst_entries[-1].grid(row=row+1, column=col, padx=2, sticky=tk.W)

    # -----------------------------------------------------------------------------------

    bottomFrame = ttk.Frame(mainFrame)
    bottomFrame.grid(row=4, column=0, padx=5, pady=5, sticky=tk.W+tk.E+tk.S)

    ttk.Label(bottomFrame, text="").grid(row=0, column=0, sticky=tk.W)

    ttk.Label(bottomFrame, text="", width=26).grid(row=1, column=0)     # width=37
    btnGrabaNuevoPago = ttk.Button(bottomFrame, text="Registra nuevo pago", state=tk.DISABLED,
                                        command=btnGraba_presionado)
    btnGrabaNuevoPago.grid(row=1, column=1, padx=5, sticky=tk.E)
    btnActualizaPagoActual = ttk.Button(bottomFrame, text="Actualiza este pago", state=tk.DISABLED,
                                        command=btnActualiza_presionado)
    btnActualizaPagoActual.grid(row=1, column=2, padx=5, sticky=tk.E)
    btnGuarda = ttk.Button(bottomFrame, text="Guarda", state=tk.DISABLED,
                                        command=btnGuarda_presionado)
    btnGuarda.grid(row=1, column=3, padx=5, sticky=tk.W)
    btnCierra = ttk.Button(bottomFrame, text="Cierra",
                                        command=btnCierra_presionado)
    btnCierra.grid(row=1, column=4, padx=5, sticky=tk.W)

    ttk.Label(bottomFrame, text="").grid(row=2, column=0, sticky=tk.W)

    return root

#
# -------------------------------------- Rutinas para el manejo de la hoja de cálculo ---
#

def column(column_name: str, as_string: bool=True):
    if column_name in columns_pagos:
        index = columns_pagos.index(column_name) + 1
        return get_column_letter(index) if as_string else index
    return ''

def es_fórmula(value):
    return isinstance(value, str) and len(value) > 0 and value[0] == '='

def evalúa_fórmula(fórmula: str):
    if es_fórmula(fórmula):
        try:
            fórmula = float(eval(fórmula[1:].replace(',', '.')))
        except:
            # Si se genera un error al evaluar la fórmula, mantener la fórmula original
            # print(f"*** Error evaluando la fórmula: {str(sys.exc_info()[1])}\n    {fórmula}")
            pass
    return fórmula


#
# === PROCESO ===========================================================================
#

# Verifica si existe la hoja de cálculo off-line
if not os.path.isfile(excel_workbook):
    print(f'Copiando la hoja de cálculo "{excel_workbook_std}"')
    print(f'                         en "{excel_workbook}"...')
    copyfile(excel_workbook_std, excel_workbook)

# Genera el listado de vecinos y su índice inverso
print('Indexando lista de vecinos...')
indexa_listado_de_vecinos()

# Cargando las hojas de cálculo de Excel
print(f'Cargando hoja de cálculo "{excel_workbook}"...')

# Carga los nombres de columnas y la columna Beneficiario en la hoja de resumen
ws_resumen = read_excel(excel_workbook_std, sheet_name=hoja_resumen)

columns_resumen = list()
for cell in ws_resumen.columns:
    if isinstance(cell, datetime):
        columns_resumen.append(cell.strftime('%m-%Y'))
    else:
        columns_resumen.append(cell)

benef_resumen = ws_resumen['Beneficiario'].tolist()
benef_resumen.insert(0, '__dummy__')

# Carga la hoja de pagos a vigilancia para obtener la lista de categorías, ordenadas
# en base al último uso
ws_vigilancia = read_excel(excel_workbook, sheet_name=hoja_vigilancia)
categorias = ws_vigilancia['Categoría'].tolist()
lista_categorías = list()
for categoria in reversed(categorias):
    if categoria not in lista_categorías:
        lista_categorías.append(categoria)

# Carga la hoja de cálculo de pagos
wb_pagos = load_workbook(excel_workbook, keep_vba=True)
ws_vigilancia = wb_pagos[hoja_vigilancia]
ws_resumen = wb_pagos[hoja_resumen]

# Nombres de columnas en hoja de pagos
first_row = ws_vigilancia[1]
columns_pagos = [col_name.value for col_name in first_row]

# Selecciona la columna «Beneficiario»
benef_pagos = [cell for cell in ws_vigilancia[column('Beneficiario')]]

# print(f"columns_resumen\n{columns_resumen}")
# print(f"\nBENEFICIARIOS_RESUMEN\n{beneficiarios_resumen}")
# sys.exit()

#
# ---------------------------------------------------------- PRESENTACIÓN EN PANTALLA ---
#

print("Creando interfase de usuario...")
root = initUI()

print('Ejecutando aplicación...')
root.mainloop()

print("Cerrando...")
print()
