# GyG GRAFICAS
#
# Genera las gráficas de:
#   - Gestión de Cobranzas,
#   - Pagos 100% Equivalentes,
#   - Pagos recibidos en el mes equivalentes a cuotas completas,
#   - Cuotas recibidas en el Mes, y
#   - Cuotas por oportunidad de pago
# en base a la última información cargada en la hoja de cálculo.

"""
    POR HACER
    -   

    HISTORICO
    -   Si se selecciona una fecha correspondiente a fin de mes, en las gráficas "Gestión de Cobranzas"
        (gráfica_1) y "Cuotas recibidas en el mes" (gráfica_4) debe mostrarse el promedio de cuotas al
        fin de mes en lugar del día específico. (17/03/2021)
         -> Si la fecha de referencia NO ES FIN DE MES, el promedio de cuotas debe mostrarse en el día
            específico. (Corregido: 20/03/2021)
    -   En la gráfica de "Cuotas por oportunidad de pago" (gráfica_5()), mostrar los promedios de las
        diferentes oportunidades de pago (en el mes, anticipadas y atrasados) para el período indicado
        (16/03/2021)
    -   Generar una gráfica similar a "Cuotas por oportunidad de pago" (gráfica_5) donde se muestren
        los resultados a fin de mes, y no a un día en particular
         -> Cambiar la opción "Gráficas del mes anterior: <mes> <año>" a "Gráficas del mes de «combo
            box: <mes> <año>»" (default: mes inmediato anterior), con subtítulo para gráfica_5, de
            "corte al día 28 de cada mes: feb/2020 a feb/2021" a "corte al fin de cada mes: "feb/2020
            a feb/2021". Seleccionando esta opción se logrará el efecto deseado. (15/03/2021)
         => ERROR: Al seleccionar "Gráficas del mes de nov 2020", se muestra como referencia "28 nov 2020"
            en lugar de "30 nov 2020". (Corregido: 16/03/2021)
            Al seleccionar una fecha correspondiente a fin de mes (ej. 28/02/2021), en gráfica_1 a
            gráfica_4, se muestra como subtítulo: "al 28 feb 2021" en lugar de "feb 2021". (Corregido:
            17/03/2021)
            Adicionalmente, en "Gestión de Cobranzas" (gráfica_1) y "Cuotas recibidas en el mes" (gráfica_4)
            el promedio de los meses anteriores (jun a oct/2020) indican 74 cuotas en lugar de 79.
             -> Se está tomando como referencia el día 28 (último día del mes de febrero, mes anterior al
                actual: 16/03/2021) (Corregido: 16/03/2021)
    -   En la Gráfica 5 (Cuotas por oportunidad de pago) se añadieron barras para resaltar el mes actual
        y el mismo mes en los años anteriores para faciliar su comparación (23/01/2021)
    -   Revisar:
            Generando gráficas...
              - Cuotas por oportunidad de pago...
            Traceback (most recent call last):
              File "./graficas_GUI.py", line 252, in separa_meses
                mensaje_final = [f"{meses.index(last_month)+1:02}-{last_year} {x}"] + mensaje_final
            ValueError: None is not in list
        El concepto "Cancelación Vigilancia, meses de Junio, Julio, Agosto, Septiembre y Octubre 2020
        y US$ 10 como anticipo para meses subsiguientes" generó el error anterior. Se corrigió cambiando
        el concepto a "Cancelación Vigilancia, meses de Junio, Julio, Agosto, Septiembre y Octubre 2020
        (saldo a favor: US$ 10,00)" (10/10/2020)
    -   Error al generar gráfica de Cuotas equivalentes (gráfica_3) al 15/07/2020: "IndexError:
        list index out of range"
         -> La rutina "distribución_de_pagps()" calculaba la distribución en base a un número
            fijo de meses (g1_nro_meses) en lugar de utilizar la cantidad de meses indicados en
            los parámetros (17/07/2020)
    -   Generar la gráfica 'Pagos recibidos en el mes equivalentes a cuotas completas' como barras
        horizontales cuando se realiza vía batch.
         -> Corregido. (02/07/2020)
    -   Correcciones pendientes:
        -   Verificar por qué 'Gestión de Cobranzas' y 'Cuotas Recibidas en el Mes' muestran
            47 cuotas al 15/06/2020 y las tres gráficas restantes muestran 46 cuotas
            completas.
             -> Las diferencias eran debidas a errores de redondeo: en unas gráficas se utilizaba
                '.as(int)' (que trunca los decimaes) y en otros 'int(round())'.
                Corregido (16/06/2020)
        -   En 'Pagos 100% Equivalentes' y 'Pagos recibidos en el mes equivalentes a cuotas
            completas', limitar los pagos recibidos a la fecha de corte. Actualmente muestra
            TODOS los pagos recibidos en el mes de referencia.
             -> Corregido (16/06/2020)
        -   En 'Cuotas Recibidas en el Mes', si la cantidad de cuotas recibidas en el mes y en
            el promedio de los meses anteriores es similar, se confunden las anotaciones de número
            de cuotas.
             -> Añadir textos '<mes> <año>:' y 'meses anteriores:', alineados por la derecha, a las
                anotaciones.
                Corregido (16/06/2020)
    -   Agregar Punto de Equilibrio a gráfica de 'Cuotas Recibidas en el Mes' y agregar '% del
        estimado' en las anotaciones del número de cuotas a la fecha.
         -> Ajustados (14/06/2020)
    -   Revisar: Con el cambio realizado para unificar el criterio para las gráficas a cuotas
        completas, 'distribución_de_pagos()' dejó de considerar el 'corte_al_día_de_referencia', por
        lo que 'gráfica_5()' muestra resultados incorrectos.
         -> Corregido (12/06/2020)
    -   Agregar intervalos de confianza para los promedios de meses anteriores en las gráficas de
        línea. (11/06/2020)
    -   En algunas gráficas se muestran pagos recibidos, y en otras, cuotas completas, lo que genera
        una diferencia en la información mostrada (hay vecinos que colaboran con un monto inferior a
        la cuota, y otros, un monto superior). Ajustar para manejar un único criterio: cuotas comple-
        tas.
         -> Unificado (05/06/2020)
    -   El Total de Cuotas y las Cuotas Mes Actual pueden solaparse -> separar...
         -> Se separaron las leyendas de las diferentes curvas de la gráfica 'Cuotas por oportunidad
            de pago' en caso de que estén solapadas (03/06/2020)
    -   Se ajustaron las gráficas 'Pagos 100% Equivalentes' y 'Pagos recibidos en el mes equivalentes
        a cuotas completas' utilizando "fig.update_yaxes(autorange='reversed')" en lugar de invertir
        manualmente el eje Y (31/05/2020)
    -   Gráfica lineal con pagos recibidos distribuidos según el período al cual apliquen: meses
        anteriores, pagos del mes actual y anticipo de meses subsiguientes... (30/05/2020)
    -   Mostrar el número de pagos recibidos directamente sobre la gráfica y no como un comentario
        aparte.
         -> Se actualizaron las gráficas 'Gestión de Cobranzas' y 'Cuotas Recibidas en el Mes' con las
            cantidades de cuotas en el mes y en el promedio de meses anteriores en la fecha de refe-
            rencia (27/05/2020)
    -   Agregar una gráfica donde se comparen la cantidad de cuotas completas recibidas día a día con
        el promedio de los últimos 'n' meses, con el objeto de informar a los vecinos sobre la gestión
        de cobranzas. (23/05/2020)
    -   La gráfica 'Pagos recibidos en el mes equivalentes a cuotas completas' no está tomando en cuenta
        el día del análisis en meses anteriores al actual.
         -> En distribución_de_pagos() y gráfica_3() se filtraron aquellos pagos posteriores a la fecha
            de referencia (23/05/2020)
    -   Ajustado el nombre de la curva 'Promedio 5 últimos meses' a 'Promedio de dic/2019 a abr/2020'
        en la gráfica 1 (Gestión de Cobranzas)
        (15/05/2020)
    -   Corregido el error en la gráfica de Cuotas Equivalentes en la opción horizontal donde se
        mostraban distribuciones y textos erroneos en las barras (12/05/2020)
    -   Ajustadas las gráficas Pagos 100% Equivalentes y Cuotas Equivalentes para que puedan ser
        desplegadas como barras horizontales (01/05/2020)
    -   Dado que los pagos de meses atrasados se hece en base a la última cuota, la gráfica de Cuotas
        100% Equivalentes podrá mostrar cifras "infladas" al utilizar como cuota de referencia la vigente
        para dicho mes, especialmente cuando ha habido cambios en el monto.
         -> Hacer un análisis similar al realizado en "Análisis de Pagos", desmembrando el pago en cada
            uno de los meses afectados y dividiéndolo por la cuota vigente para la fecho de pago.
        Se incorporó la grafica_3() para tal efecto (29/04/2020)
    -   Incluir opción para generar las gráficas del mes anterior, sin tener que seleccionar la fecha
        correspondiente:
         -> fecha_de_referencia = <último día del mes inmediato anterior> √
         -> Los gráficos son guardados en archivos del tipo <nombre de gráfica> <año>-<mes> (<mes>).png √
         -> toma_opciones_por_defecto genera archivos <nombre de gráfica>.png o <nombre de grafica>-
            <mes> (<mes>).png dependiendo si se ejecuta en último de mes o no √
         -> mantenimiento.py debe ajustarse para incluir la posibilidad de borrar las gráficas generadas
            correspondientes a los tres últimos meses √
        Cambios realizados (20/04/2020)
    -   Se agregar opciones para seleccionar la cantidad de meses a mostrar o promediar
        en las gráficas en modo GUI (11-04-2020)
    -   Se independizaron las gráficas de la información contenida en la pestaña
        'Cobranzas', tomándolas directamente de 'Vigilancia' y 'Resumen Vigilancia',
        permitiendo generar gráficas con una fecha de referencia diferente
        (10/04/2020)
    -   Versión inicial en base a la información cargada en la pestaña 'Cobranzas'
        (01/04/2020)

    REVISIONES
    -   Revisar gráfica_3(): Se muestran 62 pagos para el mes de abril, mientras que gráfica_2()
        muestra 69, que coincide con el total de abril / cuota de abril.
         -> Gráfica_3() muestra que 26 PAGOS PARA ABRIL SE HICIERON EN ABRIL. Eso quiere decir que
            el resto (69 cuotas) debieron hacerse en marzo. En efecto, 6 pagos fueron para meses
            posteriores a marzo
        Revisado 03/05/2020

"""
print('Cargando librerías...')
import GyG_constantes
from GyG_utilitarios import *
from GyG_cuotas import Cuota
import PySimpleGUI as sg

# import plotly
import plotly.graph_objs as go
from plotly.offline import plot
from pandas import read_excel, pivot_table, DataFrame
from scipy import stats
from numpy import mean, std
from datetime import datetime
from dateutil.relativedelta import relativedelta
# import swifter        # .swifter.progress_bar(False).apply()

import sys
import os
import locale
dummy = locale.setlocale(locale.LC_ALL, 'es_es')

import warnings
warnings.simplefilter("ignore", category=RuntimeWarning)


# Define algunas constantes
excel_workbook      = GyG_constantes.pagos_wb_estandar     # '1.1. GyG Recibos.xlsm'
# excel_ws_cobranzas  = GyG_constantes.pagos_ws_cobranzas
excel_ws_vigilancia = GyG_constantes.pagos_ws_vigilancia
excel_ws_resumen    = GyG_constantes.pagos_ws_resumen

num_meses_gestion_cobranzas   =  5      # Nro de meses a promediar en gráfica 'Gestión de Cobranzas'
num_meses_cuotas_equivalentes = 12      # Nro de meses a mostrar en gráfica 'Cuotas Equivalentes'

PUNTO_DE_EQUILIBRIO           = 55      # Cantidad de familias usadas en el cálculo de la cuota

FORMATO_MES         = '%b %Y'           # <mes abreviado> '.' <año>

CALENDAR_ICON       = os.path.join(GyG_constantes.rec_imágenes, '62925-spiral-calendar-icon.png')
CALENDAR_SIZE       = (16, 16)
CALENDAR_SUBSAMPLE  = 8

TITULO_GRAFICA      = '<span style="font-size: 18px"><b>{}</b></span><br>' + \
                      '<span style="font-size: 14px"><i>{}</i></span>'

VERBOSE             = False             # Muestra mensajes adicionales

GRAFICA_5_PROM_MESES_ANTERIORES = False
GRAFICA_5_NRO_ULTIMOS_MESES = 12


toma_opciones_por_defecto = False
if len(sys.argv) > 1:
    toma_opciones_por_defecto = sys.argv[1] == '--toma_opciones_por_defecto'
    if not toma_opciones_por_defecto:
        print(f"Uso: {sys.argv[0]} [--toma_opciones_por_defecto]")
        sys.exit(1)


#
# RUTINAS DE UTILIDAD
#

def es_fin_de_mes(fecha):
    fecha_aux = fecha + relativedelta(months=1)
    fin_de_mes = datetime(fecha_aux.year, fecha_aux.month, 1) - relativedelta(days=1)
    return (fecha == fin_de_mes) or seleccionado_fin_de_mes


def promedio(lista):
    if len(lista) == 0:
        return 0
    else:
        return sum(lista[:-1]) / len(lista[:-1])


def distribución_de_pagos(nro_meses, corte_al_dia_de_referencia=False):
    from openpyxl import load_workbook
    from pyparsing import Word, Regex, Literal, OneOrMore, ParseException
    import warnings
    warnings.simplefilter("ignore", category=UserWarning)

    # meses            = ['enero',      'febrero', 'marzo',     'abril',
    #                     'mayo',       'junio',   'julio',     'agosto',
    #                     'septiembre', 'octubre', 'noviembre', 'diciembre']
    # meses_abrev      = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
    #                     'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    # conectores       = ['a', '-']
    # textos_anticipos = ['adelanto', 'anticipo'   ]
    # textos_saldos    = ['ajuste',   'complemento', 'diferencia', 'saldo']
    # modificadores    = ['anticipo', 'saldo']

    # tokens_validos = meses + meses_abrev + conectores


    def ajusta_fecha(ref):
        """ ajusta_fecha toma una fecha 'ref' de la forma '%m-%Y' (ej. 'may.2019') y la convierte en
            '%Y-%m' ('2019-05') para facilitar su comparación """
        return f"{ref[3:7]}-{ref[0:2]}"

    # def separa_meses(mensaje, as_string=False, muestra_modificador=False):
    #     import re
        
    #     tokens_validos = meses + meses_abrev + conectores + modificadores

    #     mensaje = re.sub("\([^()]*\)", "", mensaje)
    #     mensaje = mensaje.lower().replace('-', ' a ').replace('/', ' ')
    #     for token in textos_anticipos:
    #         mensaje.replace(token, modificadores[0])
    #     for token in textos_saldos:
    #         mensaje.replace(token, modificadores[1])
    #     mensaje = re.sub(r"\W ", " ", mensaje).split()
    #     mensaje_ed = [x for x in mensaje if (x in tokens_validos) or x.isdigit()]
    #     last_year = None
    #     last_month = None
    #     acción = ''
    #     mensaje_anterior = None
    #     mensaje_final = list()
    #     maneja_conector = False
    #     for x in reversed(mensaje_ed):
    #         token = meses[meses_abrev.index(x)] if x in meses_abrev else x
    #         if token.isdigit():
    #             if mensaje_anterior != None:
    #                 mensaje_final = mensaje_anterior + mensaje_final
    #             last_year = token
    #             last_month = None
    #             mensaje_anterior = None
    #         elif token in meses:
    #             if mensaje_anterior != None:
    #                 mensaje_final = mensaje_anterior + mensaje_final
    #             if maneja_conector:
    #                 try:
    #                     n_last_month = meses.index(last_month)
    #                 except:
    #                     continue    # ignora los mensajes que contienen textos del tipo:
    #                                 # "(saldo a favor: Bs. 69.862,95)"
    #                 n_token = meses.index(token)
    #                 for t in reversed(range(n_token + 1, n_last_month)):
    # #                        mensaje_final = [f"{meses_abrev[t]}.{last_year}"] + mensaje_final
    #                     mensaje_final = [f"{t+1:02}-{last_year}"] + mensaje_final
    #                 maneja_conector = False
    #             last_month = token
    # #                mensaje_anterior = [f"{meses_abrev[meses.index(last_month)]}.{last_year}"]
    #             mensaje_anterior = [f"{meses.index(last_month)+1:02}-{last_year}"]
    #         elif x in conectores:
    #             maneja_conector = True
    #         elif x in modificadores and muestra_modificador:
    # #                mensaje_final = [f"{meses_abrev[meses.index(last_month)]}.{last_year} {x}"] + mensaje_final
    #             mensaje_final = [f"{meses.index(last_month)+1:02}-{last_year} {x}"] + mensaje_final
    #             mensaje_anterior = None

    #     if mensaje_anterior != None:
    #         mensaje_final = mensaje_anterior + mensaje_final

    #     if as_string:
    #         mensaje_final = '|'.join(mensaje_final)

    #     return mensaje_final


    def análisis_de_pagos(fecha_referencia):

        # Gramática de expresiones aritméticas
        operator  = Regex(r'(?<!--[\+\-\^\*/%])[\+\-]|[\^\*/%!]')
        function  = Regex(r'[a-zA-Z_][a-zA-Z0-9_]*(?=([ \t]+)?\()')
        variable  = Regex(r'[+-]?[a-zA-Z_][a-zA-Z0-9_]*(?!([ \t]+)?\()')
        number    = Regex(r'[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?')       # Ej. -125,54e-5
        lbrace    = Word('(')
        rbrace    = Word(')')
        linebreak = Word('\n')
        skip      = Word(' \t')                                             # espacio o tabulador

        lexOnly = operator | function | variable | number | lbrace \
            | rbrace | linebreak | skip
        lexAllOnly = OneOrMore(lexOnly)

        def evalua_celda(formula, modificador):
            try:
                parsed_expression = lexAllOnly.parseString(formula[1:], parseAll=True)
            except ParseException as err:
                print(f"{err.line}\n{' '*(err.column-1)}^\n{err}\n")
                return None
            result = ''
            count_braces = 0
            start_idx = 0
            curr_idx = 0
            for token in parsed_expression:
                if token == '(':
                    count_braces += 1
                elif token == ')':
                    count_braces -= 1
                elif token in ['+', '-'] and count_braces == 0:
                    result += ''.join(parsed_expression[start_idx: curr_idx])
                    if len(result) > 0 and modificador in [GyG_constantes.modificadores[0], '']:   # Cancela un Anticipo
                        return eval(result)
                    result = ''
                    start_idx = curr_idx
                curr_idx += 1
            result = ''.join(parsed_expression[start_idx:])
            return eval(result)

        pagos = df_pagos.drop(df_pagos.index[df_pagos['Fecha'] != fecha_referencia])
        total_meses_anteriores = total_mes_actual = total_meses_siguientes = 0
        num_pagos = 0
        for index, pago in pagos.iterrows():
            beneficiario = pago['Beneficiario']
            meses_pagados = separa_meses(pago['Concepto'], muestra_modificador=True)
            num_pagos += 1
            r = df_resumen[df_resumen['Beneficiario'] == beneficiario]
            for mes in separa_meses(pago['Concepto'], muestra_modificador=False):
                modificador = ''
                for mes_pago in meses_pagados:
                    if mes_pago.startswith(mes):
                        modificador = mes_pago[len(mes) + 1:]
                try:
                    pago_mes = r[mes].values[0]
                except:
                    print(f"\n*** ERROR: {str(sys.exc_info()[1])} (Beneficiario: {pago['Beneficiario']}, mes: {mes})")
                    sys.exit()
                if pago_mes == None:
                    monto = 0
                elif isinstance(pago_mes, str):
                    try:
                        _ = eval(pago_mes[1:])   # Comprueba si es una fórmula válida
                    except:
                        print(f'*** Error evaluando "{pago_mes}", ({beneficiario}, {mes})')
                        monto = 0
                    else:
                        monto = evalua_celda(pago_mes, modificador)
                else:
                    monto = pago_mes
                if ajusta_fecha(mes) < ajusta_fecha(fecha_referencia):
                    total_meses_anteriores += monto
                elif ajusta_fecha(mes) > ajusta_fecha(fecha_referencia):
                    total_meses_siguientes += monto
                else:   # mes == fecha_referencia
                    total_mes_actual += monto

        return [num_pagos, total_meses_anteriores, total_mes_actual, total_meses_siguientes]


    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Calcula cómo se distribuyen los pagos por mes')

    fecha_inicial = fecha_de_referencia - relativedelta(months=nro_meses)
    mes_inicial = datetime(fecha_inicial.year, fecha_inicial.month, 1)

    cuotas_obj = Cuota(excel_workbook)

    # Lee la hoja con el resumen de pagos. Usa la libreria OpenPyXL para preservar
    # las fórmulas en las columnas y genera un dataframe de Pandas como resultado

    # Maneja la primera columna ('Beneficiario') como su valor asociado
    df_resumen = read_excel(excel_workbook, sheet_name=excel_ws_resumen, usecols=['Beneficiario'])

    # Y el resto de las columnas como fórmulas
    wb = load_workbook(filename=excel_workbook)
    ws = wb[excel_ws_resumen]
    col_value = None
    for idx, col in enumerate(ws.columns):
        if idx == 0:
            continue
        curr_col = [expr.value for expr in col]
        col_name = curr_col.pop(0)
        try:
            if idx == 2:
                col_name = 'Check'
            elif col_name[0] == '=':
                col_name = col_value + relativedelta(months=1)
        except:
            pass
        col_value = col_name
        df_resumen[col_name] = curr_col

    # Ajusta los nombres de columnas
    new_cols = list()
    for col in df_resumen.columns.values.tolist():
        if isinstance(col, datetime):
            new_cols.append(col.strftime('%m-%Y'))
        else:
            new_cols.append(col)
    df_resumen.columns = new_cols

    df_pagos = df_vigilancia[(df_vigilancia['Fecha'] <= fecha_de_referencia) & \
                             (df_vigilancia['Fecha'] >= mes_inicial)].copy()
    df_num_cuotas = pivot_table(df_pagos[df_pagos['Fecha'] <= fecha_de_referencia],
                                values=['_num. cuotas_'], index=['Mes'], aggfunc=sum) #.astype(int)  #.reset_index()
    if corte_al_dia_de_referencia:
        df_pagos = df_pagos[df_pagos['Día'] <= fecha_de_referencia.day]
        ws_cant_pagos = pivot_table(df_pagos, values=['_num. cuotas_'], index=['Mes'], aggfunc=sum)   #.astype(int)
    else:
        ws_cant_pagos = df_num_cuotas[(df_num_cuotas.index <= fecha_de_referencia) & \
                                      (df_num_cuotas.index >= mes_inicial)]
    total_cuotas = ws_cant_pagos['_num. cuotas_'].to_list()
    df_pagos['Fecha'] = df_pagos['Fecha'].apply(lambda x: f'{x:%m-%Y}')


    def edit(valor, width=9, decimals=0):
        valor_ed = edita_número(valor, num_decimals=decimals)
        return f"{valor_ed:>{width}}"

    def edit_pct(valor, total, width=5, decimals=1):
        return edit(valor/total*100, width-1, decimals) + '%'


    if VERBOSE:
        print()
        print('   Fecha    Total Bs. Pag. Cuota   Meses anteriores           Mes actual           Anticipos')
        print('  -------  ---------- ---- -----  ------------------ -------------------- -------------------')

    lista_resultados = list()
    idx_total_cuotas = 0
    for offset in reversed(range(nro_meses+1)):
        f_ref = f"{mes_actual - relativedelta(months=offset) + relativedelta(day=1):%m-%Y}"
        resultado = análisis_de_pagos(f_ref)
        total = sum(resultado[1:])
        resultado.append(total_cuotas[idx_total_cuotas])
        lista_resultados.append(resultado)
        idx_total_cuotas += 1

        if VERBOSE:
            print(f"  {f_ref}: {edit(total, width=10)} ({resultado[0]:>3}, {int(round(resultado[4])):>3})  " + \
                  f"{edit(resultado[1])} ({edit_pct(resultado[1], total, width=6)})  " + \
                  f"{edit(resultado[2], width=10)} ({edit_pct(resultado[2], total, width=6)})  " + \
                  f"{edit(resultado[3])} ({edit_pct(resultado[3], total, width=6)})")

    if VERBOSE: print()

    return lista_resultados


#
# GRAFICAS
#

def genera_gráficas():
    global g3_horizontal

    print('Generando gráficas...')

    if g1:
        gráfica_1()     # Gestión de Cobranzas
    if g2:
        gráfica_2()     # Pagos 100% Equivalentes (montos distribuidos a lo largo de los meses)
    if g3:
        gráfica_3()     # Cuotas Equivalentes (montos recibidos en el mes)
    if g4:
        gráfica_4()     # Cuotas recibidas
    if g5:
        gráfica_5()     # Cuotas recibidas por tipo


#
# GRAFICA 1 - Gestión de Cobranzas en base al estimado de pagos --------------------------------------------
#

def gráfica_1():
    global g1_nro_meses

    grafica_nombre = 'Gestión de Cobranzas'
    if es_fin_de_mes(fecha_de_referencia):
        grafica_png = f"1. Gestion_de_cobranzas {fecha_de_referencia.strftime('%Y-%m (%b)')}.png"
    else:
        grafica_png =  '1. Gestion_de_cobranzas.png'


    if in_GUI_mode:
        # num_meses_gestion_cobranzas = int(values['_g1_nro_meses_'])
        print(f'  - {grafica_nombre}...')

    # Define las fechas de referencia
    dia_de_referencia = fecha_de_referencia.day
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    fecha_inicial = fecha_de_referencia - relativedelta(months=g1_nro_meses)
    mes_inicial = datetime(fecha_inicial.year, fecha_inicial.month, 1)
    # --------------------
    mes_final = mes_actual - relativedelta(months=1)
    str_mes_inicial = mes_inicial.strftime('%b'+('/%Y' if mes_inicial.year != mes_final.year else ''))
    str_mes_final = mes_final.strftime('%b/%Y')
    # --------------------

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Lee los estimados del servicio de vigilancia')
    ws_resumen = read_excel(excel_workbook, sheet_name=excel_ws_resumen)
    # ws_cuotas_mensuales = ws_resumen[(ws_resumen['Beneficiario'] == 'CUOTAS MENSUALES')]
    ws_pago_estimado = ws_resumen[(ws_resumen['Beneficiario'] == \
                                                    'PAGO ESTIMADO DE VIGILANCIA (Vigilantes, Pasivos y Consumibles)')]


    # Selecciona los pagos de vigilancia entre la fecha de referencia y g1_nro_meses atrás
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Filtra los pagos realizados por Categoría y Fecha')
    ws_vigilancia = df_vigilancia[(df_vigilancia['Fecha'] <= fecha_de_referencia) & \
                                  (df_vigilancia['Fecha'] >= mes_inicial)]
    # ws_vigilancia = ws_vigilancia[['Mes', 'Día', 'Monto', 'Fecha']]

    # Valida que los datos estén completos y agrega registros en cero para los datos faltantes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Valida la completitud de valores en el período ' + \
                                                                     'y rellena con ceros')
    fecha = mes_inicial
    a_agregar = []
    while fecha <= fecha_de_referencia:
        if ws_vigilancia[ws_vigilancia['Fecha'] == fecha].empty:
            a_agregar.append({'Mes':datetime(fecha.year, fecha.month, 1), 'Día':fecha.day, 'Monto':0.0, 'Fecha':fecha})
        fecha = fecha + relativedelta(days=1)

    ws_vigilancia = ws_vigilancia.append(a_agregar, ignore_index=True)

    # Normalizar los montos dividiéndolo por el estimado del mes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Normaliza los pagos de vigilancia en base al estimado ' + \
                                                                     'de pago del mes')
    # 'Monto Normalizado' = porcentaje del monto del pago recibido respecto al estimado mensual
    ws_vigilancia['_% del estimado_'] = ws_vigilancia.apply(
                                            lambda ws: ws['Monto'] / ws_pago_estimado[ws['Mes']], axis=1)

    # Totaliza los pagos de vigilancia por día y mes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Totaliza los pagos de vigilancia por día y mes')
    ws_vigilancia = pivot_table(ws_vigilancia, values=['Monto', '_% del estimado_', '_num. cuotas_'], index=['Día', 'Mes'], aggfunc=sum)
    ws_vigilancia = ws_vigilancia.reset_index() 

    # Genera columnas a graficar
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los datos del mes actual')
    ws_mes_actual = ws_vigilancia[ws_vigilancia['Mes'] == mes_actual].copy().reset_index()
    ws_mes_actual['_% acumulado a la fecha_'] = ws_mes_actual['_% del estimado_'].cumsum()
    ws_mes_actual['_cuotas a la fecha_'] = ws_mes_actual['_num. cuotas_'].cumsum()
    total_recaudado = ws_mes_actual['Monto'].sum()
    pct_del_estimado = total_recaudado / ws_pago_estimado[mes_actual] * 100

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los datos del promedio de los meses anteriores')
    ws_meses_anteriores = ws_vigilancia[ws_vigilancia['Mes'] < mes_actual]
    ws_std_cuotas = pivot_table(ws_meses_anteriores, values=['_% del estimado_'], index=['Día'], aggfunc=std)  #.reset_index()
    ws_meses_anteriores = pivot_table(ws_meses_anteriores, values='_% del estimado_', index=['Día'], fill_value=0)
    ws_meses_anteriores = ws_meses_anteriores.reset_index()
    ws_meses_anteriores['_% acumulado a la fecha_'] = ws_meses_anteriores['_% del estimado_'].cumsum()
    ws_cuotas_anteriores = ws_vigilancia[ws_vigilancia['Mes'] < mes_actual]
    ws_cuotas_anteriores = pivot_table(ws_cuotas_anteriores, values='_num. cuotas_', index=['Día'], fill_value=0)
    ws_cuotas_anteriores = ws_cuotas_anteriores.reset_index()
    ws_meses_anteriores['_cuotas a la fecha_'] = ws_cuotas_anteriores['_num. cuotas_'].cumsum()

    # Mes actual
    trace_mes_actual = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = ws_mes_actual['_% acumulado a la fecha_'],
                mode = 'lines',
                name = fecha_de_referencia.strftime(FORMATO_MES),
                line = dict(
                            width = 3.5,
                            color = '#f44611',      # naranja puro
                        )
            )

    # Regresión lineal
    x = ws_meses_anteriores['Día']
    y = ws_mes_actual['_% acumulado a la fecha_']
    slope, intercept, r_value, p_value, std_err = stats.linregress(x[:dia_de_referencia], y[:dia_de_referencia])
    linea_lr = [slope * xi + intercept for xi in x]

    trace_regresion = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = linea_lr,
                mode = 'lines',
                name = 'Tendencia ' + fecha_de_referencia.strftime(FORMATO_MES),
                line = dict(
                            width = 2,
                            color = '#f44611',      # naranja puro
                            dash = 'dot',
                        )
            )

    # Promedio de los últimos 5 meses
    trace_prom_5_ultimos_meses = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = ws_meses_anteriores['_% acumulado a la fecha_'],
                mode = 'lines',
                # name = f"Promedio {g1_nro_meses} últimos meses",
                name = f"Promedio de {str_mes_inicial} a {str_mes_final}",
                line = dict(
                            width = 3,
                            color = 'blue',
                        ),
                fillcolor = 'rgba(68, 68, 68, 0.2)' if g1_conf_interval else None,
                fill = 'tonexty' if g1_conf_interval else None,
            )

    if g1_conf_interval:
        # Upper bound promedio de los últimos 5 meses
        trace_ub_prom_meses_anteriores = go.Scatter(
                    name = 'Upper bound',
                    x = ws_meses_anteriores['Día'],
                    y = ws_meses_anteriores['_% acumulado a la fecha_'] + ws_std_cuotas['_% del estimado_'],
                    mode = 'lines',
                    line = dict(
                                width = 0,
                            ),
                    fillcolor = 'rgba(68, 68, 68, 0.2)',
                    fill = 'tonexty',
                    showlegend=False,
                )

        trace_lb_prom_meses_anteriores = go.Scatter(
                    name = 'Lower bound',
                    x = ws_meses_anteriores['Día'],
                    y = ws_meses_anteriores['_% acumulado a la fecha_'] - ws_std_cuotas['_% del estimado_'],
                    mode = 'lines',
                    line = dict(
                                width = 0,
                            ),
                    showlegend=False,
                )

    # Pago Vigilantes (estimado)
    trace_pago_vigilantes = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = [1.0 for _ in range(31)],
                mode = 'lines',
                name = 'Pago Vigilantes (estimado)',
                line = dict(
                            color = 'gray',
                            dash = 'solid',
                            width = 2,
                        )
            )

    if g1_conf_interval:
        data = [trace_regresion,
                trace_pago_vigilantes,
                trace_lb_prom_meses_anteriores, trace_prom_5_ultimos_meses, trace_ub_prom_meses_anteriores,
                trace_mes_actual]
    else:
        data = [trace_mes_actual,
                trace_prom_5_ultimos_meses,
                trace_pago_vigilantes,
                trace_regresion]

    título = 'Gestión de Cobranzas'
    subtítulo = 'en base al estimado de pagos, '
    if es_fin_de_mes(fecha_de_referencia):
        subtítulo += fecha_de_referencia.strftime("%b %Y")
    else:
        subtítulo += 'al ' + fecha_de_referencia.strftime("%d %b %Y")

    comentario = f"Recaudado: Bs. {edita_número(total_recaudado, num_decimals=0)} " + \
                 f"({edita_número(pct_del_estimado, num_decimals=0)}% del estimado)"
    día = fecha_de_referencia.day

    layout = go.Layout(
                title = TITULO_GRAFICA.format(título, subtítulo),
                width = 1000,
                height = 600,
                xaxis = dict(
                            nticks = 31,
                            showgrid = True, gridwidth=1, gridcolor='lightgray',
                            showline=True, linewidth=1, linecolor='black',
                        ),
                yaxis = dict(
                            title = '% Cobertura de Pagos de Vigilantes, Pasivos y Consumibles',
                            tickformat = ',.0%',
                            showline=True, linewidth=1, linecolor='black',
                        ),
                legend = dict(
                            orientation = 'h',
                            tracegroupgap = 20,
                        ),
                annotations = [
                        dict(
                            text = comentario,
                            font = dict(
                                    size = 14,
                                ),
                            showarrow = False,
                            xref = 'paper',
                            yref = 'paper',
                            x = 0.98,
                            y = 0.05,
                            align = 'right',
                            valign = 'bottom',
                        ),
                    ],
#                paper_bgcolor = 'white',
                plot_bgcolor = 'white',
            )

    fig = go.Figure(data = data, layout = layout)

    ultimo_indice_act = ws_mes_actual['_% acumulado a la fecha_'].index[-1]
    ultimo_indice_ant = ws_meses_anteriores['_% acumulado a la fecha_'].index[-1]
    if not es_fin_de_mes(fecha_de_referencia):
        ultimo_indice_ant = ultimo_indice_act

    dx, dy = 20, 40
    actual_ge_anteriores = ws_mes_actual['_% acumulado a la fecha_'][ultimo_indice_act] >= \
                               ws_meses_anteriores['_% acumulado a la fecha_'][ultimo_indice_ant]

    delta_x = -dx if actual_ge_anteriores else dx
    if ultimo_indice_act == 0:
        delta_x = abs(delta_x)
    elif ultimo_indice_act == 30:
        delta_x = -abs(delta_x)
    fig.add_annotation(
        x = ws_meses_anteriores['Día'][ultimo_indice_act],
        y = ws_mes_actual['_% acumulado a la fecha_'][ultimo_indice_act],
        text = f'{int(round(ws_mes_actual["_cuotas a la fecha_"][día - 1]))} cuotas<span style="font-size: 10px"><br>' + \
               f'<i>({int(round(ws_mes_actual["_% acumulado a la fecha_"][día - 1]*100))}% del estim.)</i></span>',
        xref = "x",
        yref = "y",
        ax = delta_x,
        ay = -dy if actual_ge_anteriores else dy,
        showarrow = True,
        arrowhead = 3, arrowsize = 1.5,
        standoff = 3,
    )
    delta_x = -dx if not actual_ge_anteriores else dx
    if ultimo_indice_ant == 0:
        delta_x = abs(delta_x)
    elif ultimo_indice_ant == 30:
        delta_x = -abs(delta_x)
    fig.add_annotation(
        x = ws_meses_anteriores['Día'][ultimo_indice_ant],
        y = ws_meses_anteriores['_% acumulado a la fecha_'][ultimo_indice_ant],
        text = f'{int(round(ws_meses_anteriores["_cuotas a la fecha_"][día - 1]))} cuotas<span style="font-size: 10px"><br>' + \
               f'<i>({int(round(ws_meses_anteriores["_% acumulado a la fecha_"][día - 1]*100))}% del estim.)</i></span>',
        xref = "x",
        yref = "y",
        ax = delta_x,
        ay = -dy if not actual_ge_anteriores else dy,
        showarrow = True,
        arrowhead = 3, arrowsize = 1.5,
        standoff = 3,
    )

    if g1_show:
        print(f'    . Desplegando imagen...')
        fig.show()

    if g1_save:
        print(f'    . Grabando "{grafica_png}"...')
        img_bytes = fig.to_image(format="png")
        with open(os.path.join(GyG_constantes.ruta_graficas, grafica_png), 'wb') as f:
            f.write(img_bytes)

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Gráfica concluida')



#
# GRAFICA 2 - Pagos 100% Equivalentes ----------------------------------------------------------------------
#

def gráfica_2():
    global g2_nro_meses

    VERTICAL_BARS = not g2_horizontal

    grafica_nombre = 'Pagos 100% equivalentes'
    if es_fin_de_mes(fecha_de_referencia):
        grafica_png = f"2. Pagos_100pct_equivalentes {fecha_de_referencia.strftime('%Y-%m (%b)')}.png"
    else:
        grafica_png =  '2. Pagos_100pct_equivalentes.png'


    if in_GUI_mode:
        # num_meses_cuotas_equivalentes = int(values['_g2_nro_meses_'])
        print(f'  - {grafica_nombre}...')

    dia_de_referencia = fecha_de_referencia.day
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    df_num_cuotas = pivot_table(df_vigilancia[df_vigilancia['Fecha'] <= fecha_de_referencia],
                                values=['_num. cuotas_'], index=['Mes'], aggfunc=sum) #.astype(int)  #.reset_index()

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los rangos de la imagen')
    meses = [mes_actual - relativedelta(months=mes) for mes in reversed(range(g2_nro_meses +1))]
    meses_ed = [mes.strftime(FORMATO_MES) for mes in meses]
    num_cuotas_completas = [int(round(df_num_cuotas.loc[mes].values[0])) for mes in meses]
    promedio = int(round(mean(num_cuotas_completas[:-1]), ndigits=0))
    promedios = [promedio for _ in meses]
    pto_equilibrio = [PUNTO_DE_EQUILIBRIO for _ in meses]


    título =    'Pagos 100% Equivalentes'
    if es_fin_de_mes(fecha_de_referencia):
        subtítulo = fecha_de_referencia.strftime("%b %Y")
    else:
        subtítulo = 'al ' + fecha_de_referencia.strftime("%d %b %Y")

    # 100% Equivalentes
    trace_100pct_equiv = go.Bar(
                x = meses_ed if VERTICAL_BARS else num_cuotas_completas,
                y = num_cuotas_completas if VERTICAL_BARS else meses_ed,
                orientation = 'v' if VERTICAL_BARS else 'h',
                name = 'Cuotas completas',
                text = [f"{y:.0f}" for y in num_cuotas_completas],
                textposition = "outside",
                marker_color = 'cornflowerblue',        # '#5b9bd5',
                # marker = dict(
                #             color = 'mediumblue',
                #         ),
            )

    # Promedio últimos 12 meses
    trace_prom_12_ultimos_meses = go.Scatter(
                x = meses_ed if VERTICAL_BARS else promedios,
                y = promedios if VERTICAL_BARS else meses_ed,
                mode = 'lines',
                name = f"Promedio de los {g2_nro_meses} meses anteriores = {promedio}; " + \
                       f"actual = {num_cuotas_completas[-1]}",
                line = dict(
                            color = '#ed7d31',          # 'orange',
                            width = 2,
                        ),
            )

    # Punto de equilibrio
    trace_punto_equilibrio = go.Scatter(
                x = meses_ed if VERTICAL_BARS else pto_equilibrio,
                y = pto_equilibrio if VERTICAL_BARS else meses_ed,
                mode = 'lines',
                name = f"Punto de equilibrio = {pto_equilibrio[0]} cuotas completas",
                line = dict(
                            color = 'red',           # '#ffc000',
                            dash = 'dash',
                            width = 2,
                        ),
            )

    data = [trace_100pct_equiv, trace_prom_12_ultimos_meses, trace_punto_equilibrio]

    layout = dict(
                title = TITULO_GRAFICA.format(título, subtítulo),
                width = 1000,
                height = 600,
                legend = dict(
                            orientation = 'h',
                            tracegroupgap = 20,
                        ),
                xaxis_tickangle = -30 if VERTICAL_BARS else 0,
#                paper_bgcolor = 'white',
                plot_bgcolor = 'white',
            )

    fig = go.Figure(data = data, layout = layout)

    if not VERTICAL_BARS:
        fig.update_yaxes(autorange='reversed')

    if g2_show:
        print(f"    . Desplegando imagen {'vertical' if VERTICAL_BARS else 'horizontal'}...")
        fig.show()

    if g2_save:
        print(f'    . Grabando "{grafica_png}"...')
        img_bytes = fig.to_image(format="png")
        with open(os.path.join(GyG_constantes.ruta_graficas, grafica_png), 'wb') as f:
            f.write(img_bytes)



#
# GRAFICA 3 - Pagos recibidos en el mes equivalentes a cuotas completas ------------------------------------
#

def gráfica_3():
    global g3_nro_meses
    from GyG_cuotas import Cuota

    VERTICAL_BARS = not g3_horizontal

    grafica_nombre = 'Cuotas equivalentes'
    if es_fin_de_mes(fecha_de_referencia):
        grafica_png = f"3. Cuotas_equivalentes {fecha_de_referencia.strftime('%Y-%m (%b)')}.png"
    else:
        grafica_png =  '3. Cuotas_equivalentes.png'


    if in_GUI_mode:
        # num_meses_cuotas_equivalentes = int(values['_g3_nro_meses_'])
        print(f'  - {grafica_nombre}...')

    dia_de_referencia = fecha_de_referencia.day
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    fecha_inicial = fecha_de_referencia - relativedelta(months=g3_nro_meses)
    mes_inicial = datetime(fecha_inicial.year, fecha_inicial.month, 1)

    distribución = distribución_de_pagos(nro_meses=g3_nro_meses)

    pctMesAnt = list()
    pctMesAct = list()
    pctMesSig = list()
    str_pctMesAnt = list()
    str_pctMesAct = list()
    str_pctMesSig = list()
    for nPagos, mesAnt, mesAct, mesSig, nCuotas in distribución:
        total_mes = mesAnt + mesAct + mesSig
        if total_mes == 0.00:
            pctMesAnt.append(pMesAnt := 0.00)
            pctMesAct.append(pMesAct := 0.00)
            pctMesSig.append(pMesSig := 0.00)
        else:
            pctMesAnt.append(pMesAnt := mesAnt / total_mes)
            pctMesAct.append(pMesAct := mesAct / total_mes)
            pctMesSig.append(pMesSig := mesSig / total_mes)
        str_pctMesAnt.append(edita_número(pMesAnt * 100, num_decimals=1) + '%')
        str_pctMesAct.append(edita_número(pMesAct * 100, num_decimals=1) + '%')
        str_pctMesSig.append(edita_número(pMesSig * 100, num_decimals=1) + '%')

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los rangos de la imagen')
    df_num_cuotas = pivot_table(df_vigilancia[df_vigilancia['Fecha'] <= fecha_de_referencia],
                                values=['_num. cuotas_'], index=['Mes'], aggfunc=sum) #.astype(int)  #.reset_index()
    ws_cant_pagos = df_num_cuotas[(mes_inicial <= df_num_cuotas.index) & \
                                  (df_num_cuotas.index <= fecha_de_referencia)]
    meses = ws_cant_pagos.index.to_list()
    num_cuotas_completas = ws_cant_pagos['_num. cuotas_'].to_list()
    meses_ed = [mes.strftime(FORMATO_MES) for mes in meses]
    promedio = int(round(mean(num_cuotas_completas[:-1]), ndigits=0))

    bar_mesAnt = [num_pagos * pctMesAnt[idx] for idx, num_pagos in enumerate(num_cuotas_completas)]
    bar_mesAct = [num_pagos * pctMesAct[idx] for idx, num_pagos in enumerate(num_cuotas_completas)]
    bar_mesSig = [num_pagos * pctMesSig[idx] for idx, num_pagos in enumerate(num_cuotas_completas)]
    sep_1, sep_2 = ('<br />', '') if VERTICAL_BARS else (' cuotas (', ')')
    str_pctMesAct = [f"{int(round(numPagos * pctMesAct[idx], 0))}{sep_1}" + \
                     f"{edita_número(pctMesAct[idx] * 100, num_decimals=1)}%{sep_2}" \
                            for idx, numPagos in enumerate(num_cuotas_completas)]
    promedios = [promedio for _ in meses]
    pto_equilibrio = [PUNTO_DE_EQUILIBRIO for _ in meses]

    título =    'Pagos recibidos en el mes equivalentes a cuotas completas'
    if es_fin_de_mes(fecha_de_referencia):
        subtítulo = fecha_de_referencia.strftime("%b %Y")
    else:
        subtítulo = 'al ' + fecha_de_referencia.strftime("%d %b %Y")


    fig = go.Figure()

    # Distribución de cuotas
    fig.add_trace(
                    go.Bar(
                        name = 'Meses anteriores',
                        x = meses_ed if VERTICAL_BARS else bar_mesAnt,
                        y = bar_mesAnt if VERTICAL_BARS else meses_ed,
                        orientation = 'v' if VERTICAL_BARS else 'h',
                        text = str_pctMesAnt,
                        textposition = "inside",
                        marker_color = '#4475cd',            # '#5485dd',
                        legendgroup = 'bar',
                    ))
    fig.add_trace(
                    go.Bar(
                        name = 'Mes actual',
                        x = meses_ed if VERTICAL_BARS else bar_mesAct,
                        y = bar_mesAct if VERTICAL_BARS else meses_ed,
                        orientation = 'v' if VERTICAL_BARS else 'h',
                        text = str_pctMesAct,
                        textposition = "inside",
                        marker_color = 'cornflowerblue',     # '#6495ed',
                        legendgroup = 'bar',
                    ))
    fig.add_trace(
                    go.Bar(
                        name = 'Meses subsiguientes',
                        x = meses_ed if VERTICAL_BARS else bar_mesSig,
                        y = bar_mesSig if VERTICAL_BARS else meses_ed,
                        orientation = 'v' if VERTICAL_BARS else 'h',
                        text = str_pctMesSig,
                        textposition = "inside",
                        marker_color = '#84b5ff',           # '#74a5fd',
                        legendgroup = 'bar',
                    ))
    fig.add_trace(
                    go.Bar(
                        name = '',
                        x = meses_ed if VERTICAL_BARS else [0 for _ in meses],
                        y = [0 for _ in meses] if VERTICAL_BARS else meses_ed,
                        orientation = 'v' if VERTICAL_BARS else 'h',
                        text = [f"{y:.0f}" for y in num_cuotas_completas],
                        textposition = "outside",
                        marker_color = 'white',             # '#5b9bd5',
                        showlegend=False,
                    ))
    fig.update_layout(
                    title = TITULO_GRAFICA.format(título, subtítulo),
                    width = 1000,
                    height = 600,
                    legend = dict(
                                orientation = 'h',
                                tracegroupgap = 20,
                            ),
                    xaxis_tickangle = -30 if VERTICAL_BARS else 0,
    #                xaxis_title = 'Cantidad de cuotas completas',
    #                paper_bgcolor = 'white',
                    plot_bgcolor = 'white',
                    barmode='stack',
                )

    # # Punto de equilibrio
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed if VERTICAL_BARS else pto_equilibrio,
                        y = pto_equilibrio if VERTICAL_BARS else meses_ed,
                        mode = 'lines',
                        name = f"Punto de equilibrio = {pto_equilibrio[0]} cuotas completas",
                        line = dict(
                                    color = 'red',          # 'yellow',           # '#ffc000',
                                    dash = 'dash',
                                    width = 2,
                                ),
                       legendgroup = 'lines',
                    ))

    # Promedio últimos 12 meses
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed if VERTICAL_BARS else promedios,
                        y = promedios if VERTICAL_BARS else meses_ed,
                        mode = 'lines',
                        name = f"Promedio de los {g3_nro_meses} meses anteriores = {promedio}; " + \
                               f"actual = {int(round(num_cuotas_completas[-1], 0))}",
                        line = dict(
                                    color = 'blue',         # '#ed7d31',          # 'orange',
                                    width = 2,
                                ),
                       legendgroup = 'lines',
                    ))

    if not VERTICAL_BARS:
        fig.update_yaxes(autorange='reversed')

    if g3_show:
        print(f"    . Desplegando imagen {'vertical' if VERTICAL_BARS else 'horizontal'}...")
        fig.show()

    if g3_save:
        print(f'    . Grabando "{grafica_png}"...')
        img_bytes = fig.to_image(format="png")
        with open(os.path.join(GyG_constantes.ruta_graficas, grafica_png), 'wb') as f:
            f.write(img_bytes)



#
# GRAFICA 4 - Cuotas recibidas en el mes -------------------------------------------------------------------
#

def gráfica_4():
    global g4_nro_meses

    grafica_nombre = 'Cuotas Recibidas en el Mes'
    if es_fin_de_mes(fecha_de_referencia):
        grafica_png = f"4. Cuotas_recibidas {fecha_de_referencia.strftime('%Y-%m (%b)')}.png"
    else:
        grafica_png =  '4. Cuotas_recibidas.png'


    if in_GUI_mode:
        # num_meses_gestion_cobranzas = int(values['_g4_nro_meses_'])
        print(f'  - {grafica_nombre}...')

    # Define las fechas de referencia
    dia_de_referencia = fecha_de_referencia.day
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    fecha_inicial = fecha_de_referencia - relativedelta(months=g4_nro_meses)
    mes_inicial = datetime(fecha_inicial.year, fecha_inicial.month, 1)
    # --------------------
    mes_final = mes_actual - relativedelta(months=1)
    str_mes_inicial = mes_inicial.strftime('%b'+('/%Y' if mes_inicial.year != mes_final.year else ''))
    str_mes_final = mes_final.strftime('%b/%Y')
    # --------------------

    # Selecciona los pagos de vigilancia entre la fecha de referencia y g4_nro_meses atrás
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Filtra los pagos realizados por Categoría y Fecha')
    ws_vigilancia = df_vigilancia[(df_vigilancia['Fecha'] <= fecha_de_referencia) & \
                                  (df_vigilancia['Fecha'] >= mes_inicial)]

    # Valida que los datos estén completos y agrega registros en cero para los datos faltantes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Valida la completitud de valores en el período ' + \
                                                                     'y rellena con ceros')
    fecha = mes_inicial
    a_agregar = []
    while fecha <= fecha_de_referencia:
        if ws_vigilancia[ws_vigilancia['Fecha'] == fecha].empty:
            a_agregar.append({'Mes':datetime(fecha.year, fecha.month, 1), 'Día':fecha.day, 'Monto':0.0, 'Fecha':fecha})
        fecha = fecha + relativedelta(days=1)

    ws_vigilancia = ws_vigilancia.append(a_agregar, ignore_index=True)

    # Normalizar los montos dividiéndolo por el estimado del mes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Normaliza los pagos de vigilancia en base a la cuota ' + \
                                                                     'vigente a la fecha')

    ws_vigilancia['_% de la cuota_'] = ws_vigilancia.apply(
                                            lambda ws: ws['Monto'] / ws['_cuota_'], axis=1)

    # Totaliza los pagos de vigilancia por día y mes
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Totaliza los pagos de vigilancia por día y mes')
    ws_vigilancia = pivot_table(ws_vigilancia, values=['Monto', '_% de la cuota_'], index=['Día', 'Mes'], aggfunc=sum)
    ws_vigilancia = ws_vigilancia.reset_index() 

    # Genera columnas a graficar
    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los datos del mes actual')
    ws_mes_actual = ws_vigilancia[ws_vigilancia['Mes'] == mes_actual].copy().reset_index()
    ws_mes_actual['_% acumulado a la fecha_'] = ws_mes_actual['_% de la cuota_'].cumsum()

    total_recaudado = ws_mes_actual['Monto'].sum()

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los datos del promedio de los meses anteriores')
    ws_meses_anteriores = ws_vigilancia[ws_vigilancia['Mes'] < mes_actual]
    ws_std_cuotas = pivot_table(ws_meses_anteriores, values=['_% de la cuota_'], index=['Día'], aggfunc=std)  #.reset_index()
                        # Usar df_std_cuotas.loc[<variable con el mes a buscar de la forma datetime(year, month, 1)>]
    ws_meses_anteriores = pivot_table(ws_meses_anteriores, values='_% de la cuota_', index=['Día'], fill_value=0)
    ws_meses_anteriores = ws_meses_anteriores.reset_index()
    ws_meses_anteriores['_% acumulado a la fecha_'] = ws_meses_anteriores['_% de la cuota_'].cumsum()
    pto_equilibrio = [PUNTO_DE_EQUILIBRIO for _ in ws_meses_anteriores['Día']]

    # Mes actual
    trace_mes_actual = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = ws_mes_actual['_% acumulado a la fecha_'],
                mode = 'lines',
                name = fecha_de_referencia.strftime(FORMATO_MES),
                line = dict(
                            width = 3.5,
                            color = '#f44611',      # naranja puro
                        )
            )

    # Regresión lineal
    x = ws_meses_anteriores['Día']
    y = ws_mes_actual['_% acumulado a la fecha_']
    slope, intercept, r_value, p_value, std_err = stats.linregress(x[:dia_de_referencia], y[:dia_de_referencia])
    linea_lr = [slope * xi + intercept for xi in x]

    trace_regresion = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = linea_lr,
                mode = 'lines',
                name = 'Tendencia ' + fecha_de_referencia.strftime(FORMATO_MES),
                line = dict(
                            width = 2,
                            color = '#f44611',      # naranja puro
                            dash = 'dot',
                        )
            )

    # Promedio de los últimos 5 meses
    trace_prom_5_ultimos_meses = go.Scatter(
                x = ws_meses_anteriores['Día'],
                y = ws_meses_anteriores['_% acumulado a la fecha_'],
                mode = 'lines',
                # name = f"Promedio {g4_nro_meses} últimos meses",
                name = f"Promedio de los {g4_nro_meses} meses anteriores: {str_mes_inicial} a {str_mes_final}",
                line = dict(
                            width = 3,
                            color = 'blue',
                        ),
                fillcolor = 'rgba(68, 68, 68, 0.2)' if g4_conf_interval else None,
                fill = 'tonexty' if g4_conf_interval else None,
            )

    if g4_conf_interval:
        # Upper bound promedio de los últimos 5 meses
        trace_ub_prom_meses_anteriores = go.Scatter(
                    name = 'Upper bound',
                    x = ws_meses_anteriores['Día'],
                    y = ws_meses_anteriores['_% acumulado a la fecha_'] + ws_std_cuotas['_% de la cuota_'],
                    mode = 'lines',
                    line = dict(
                                width = 0,
                            ),
                    fillcolor = 'rgba(68, 68, 68, 0.2)',
                    fill = 'tonexty',
                    showlegend=False,
                )

        trace_lb_prom_meses_anteriores = go.Scatter(
                    name = 'Lower bound',
                    x = ws_meses_anteriores['Día'],
                    y = ws_meses_anteriores['_% acumulado a la fecha_'] - ws_std_cuotas['_% de la cuota_'],
                    mode = 'lines',
                    line = dict(
                                width = 0,
                            ),
                    showlegend=False,
                )

    # # Punto de equilibrio
    trace_punto_equilibrio = go.Scatter(
                    x = ws_meses_anteriores['Día'],
                    y = pto_equilibrio,
                    mode = 'lines',
                    name = f"Pto. equilibrio = {pto_equilibrio[0]} cuotas",     # ... completas
                    line = dict(
                                color = 'red',          # 'yellow',           # '#ffc000',
                                dash = 'dash',
                                width = 1.5,
                            ),
                )

    if g4_conf_interval:
        data = [trace_punto_equilibrio,
                trace_regresion,
                trace_lb_prom_meses_anteriores, trace_prom_5_ultimos_meses, trace_ub_prom_meses_anteriores,
                trace_mes_actual]
    else:
        data = [trace_mes_actual,
                trace_prom_5_ultimos_meses,
                trace_regresion,
                trace_punto_equilibrio]

    título =    'Cuotas Recibidas en el Mes'
    if es_fin_de_mes(fecha_de_referencia):
        subtítulo = fecha_de_referencia.strftime("%b %Y")
    else:
        subtítulo = 'al ' + fecha_de_referencia.strftime("%d %b %Y")
    día = fecha_de_referencia.day

    layout = go.Layout(
                title = dict(
                            text = TITULO_GRAFICA.format(título, subtítulo),
                        ),
                width = 1000,
                height = 600,
                xaxis = dict(
                            nticks = 31,
                            showgrid = True, gridwidth=1, gridcolor='lightgray',
                            showline=True, linewidth=1, linecolor='black',
                        ),
                yaxis = dict(
                            title = 'Cantidad de cuotas recibidas en el mes',
                            tickformat = ',.0',
                            showline=True, linewidth=1, linecolor='black',
                        ),
                legend = dict(
                            orientation = 'h',
                            tracegroupgap = 20,
                        ),
#                paper_bgcolor = 'white',
                plot_bgcolor = 'white',
            )

    fig = go.Figure(data = data, layout = layout)

    ultimo_indice_act = ws_mes_actual['_% acumulado a la fecha_'].index[-1]
    ultimo_indice_ant = ws_meses_anteriores['_% acumulado a la fecha_'].index[-1]
    if not es_fin_de_mes(fecha_de_referencia):
        ultimo_indice_ant = ultimo_indice_act

    dx, dy = 20, 40
    actual_ge_anteriores = ws_mes_actual['_% acumulado a la fecha_'][ultimo_indice_act] >= \
                               ws_meses_anteriores['_% acumulado a la fecha_'][ultimo_indice_ant]

    # Verifica si los puntos correspondientes a ambas anotaciones están muy cerca entre sí
    muestra_identificador = abs(ws_mes_actual['_% acumulado a la fecha_'][ultimo_indice_act] - \
                                    ws_meses_anteriores['_% acumulado a la fecha_'][ultimo_indice_ant]) <= 2

    delta_x = -dx if actual_ge_anteriores else dx
    if ultimo_indice_act == 0:
        delta_x = abs(delta_x)
    elif ultimo_indice_act == 30:
        delta_x = -abs(delta_x)
    fig.add_annotation(
        x = ws_meses_anteriores['Día'][ultimo_indice_act],
        y = ws_mes_actual['_% acumulado a la fecha_'][ultimo_indice_act],
        text = (f"{fecha_de_referencia.strftime(FORMATO_MES)}<br>" if muestra_identificador else "") + \
                f"{int(round(ws_mes_actual['_% acumulado a la fecha_'][día - 1]))} cuotas",
        align = 'right',
        # text = f"{int(round(ws_mes_actual['_% acumulado a la fecha_'][día - 1]))} cuotas",
        ax = delta_x,
        ay = -dy if actual_ge_anteriores else dy
    )
    delta_x = -dx if not actual_ge_anteriores else dx
    if ultimo_indice_ant == 0:
        delta_x = abs(delta_x)
    elif ultimo_indice_ant == 30:
        delta_x = -abs(delta_x)
    fig.add_annotation(
        x = ws_meses_anteriores['Día'][ultimo_indice_ant],
        y = ws_meses_anteriores['_% acumulado a la fecha_'][ultimo_indice_ant],
        text = (f"meses anteriores<br>" if muestra_identificador else "") + \
                f"{int(round(ws_meses_anteriores['_% acumulado a la fecha_'][día - 1]))} cuotas",
        align = 'right',
        # text = f"{int(round(ws_meses_anteriores['_% acumulado a la fecha_'][día - 1]))} cuotas",
        ax = delta_x,
        ay = -dy if not actual_ge_anteriores else dy
    )
    fig.update_annotations(dict(
        xref = "x",
        yref = "y",
        showarrow = True,
        arrowhead = 3, arrowsize = 1.5,
        standoff = 3,
    ))

    if g4_show:
        print(f'    . Desplegando imagen...')
        fig.show()

    if g4_save:
        print(f'    . Grabando "{grafica_png}"...')
        img_bytes = fig.to_image(format="png")
        with open(os.path.join(GyG_constantes.ruta_graficas, grafica_png), 'wb') as f:
            f.write(img_bytes)

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Gráfica concluida')



#
# GRAFICA 5 - Cuotas recibidas por oportunidad de pago -----------------------------------------------------
#

def gráfica_5():
    global g5_nro_meses
    from GyG_cuotas import Cuota

    grafica_nombre = 'Cuotas por oportunidad de pago'
    if es_fin_de_mes(fecha_de_referencia):
        grafica_png = f"5. Cuotas_por_oportunidad_de_pago {fecha_de_referencia.strftime('%Y-%m (%b)')}.png"
    else:
        grafica_png =  '5. Cuotas_por_oportunidad_de_pago.png'

    if in_GUI_mode:
        # num_meses_cuotas_equivalentes = int(values['_g5_nro_meses_'])
        print(f'  - {grafica_nombre}...')

    dia_de_referencia = fecha_de_referencia.day
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    mes_anterior = mes_actual - relativedelta(months=1)
    fecha_inicial = fecha_de_referencia - relativedelta(months=g5_nro_meses)
    mes_inicial = datetime(fecha_inicial.year, fecha_inicial.month, 1)
    str_mes_inicial = mes_inicial.strftime('%b'+('/%Y' if mes_inicial.year != fecha_de_referencia.year else ''))
    str_mes_final = fecha_de_referencia.strftime('%b/%Y')
    str_mes_inicial_promedios =mes_inicial.strftime('%b'+('/%Y' if mes_inicial.year != mes_anterior.year else ''))
    str_mes_final_promedios = mes_anterior.strftime('%b/%Y')

    distribución = distribución_de_pagos(nro_meses=g5_nro_meses,
                                         corte_al_dia_de_referencia=not seleccionado_fin_de_mes)

    if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Genera los rangos de la imagen')
    GRAFICA_5_BARRA_COMPARACION = g5_nro_meses > 12
    nPagosMesAnt = list()
    nPagosMesAct = list()
    nPagosMesSig = list()
    nPagosTotal = list()
    for nPagos, mesAnt, mesAct, mesSig, nCuotas in distribución:
        total_mes = mesAnt + mesAct + mesSig
        nPagosMesAnt.append(0 if total_mes == 0.00 else int(round(nCuotas * mesAnt / total_mes)))
        nPagosMesAct.append(0 if total_mes == 0.00 else int(round(nCuotas * mesAct / total_mes)))
        nPagosMesSig.append(0 if total_mes == 0.00 else int(round(nCuotas * mesSig / total_mes)))
        nPagosTotal.append(0 if total_mes == 0.00 else int(round(nCuotas)))

    promPagosMesAnt = int(round(promedio(nPagosMesAnt)))
    promPagosMesAct = int(round(promedio(nPagosMesAct)))
    promPagosMesSig = int(round(promedio(nPagosMesSig)))
    promPagosTotal = int(round(promedio(nPagosTotal)))

    sAnt = '' if promPagosMesAnt == 1 else 's'
    sAct = '' if promPagosMesAct == 1 else 's'
    sSig = '' if promPagosMesSig == 1 else 's'
    sTotal = '' if promPagosTotal == 1 else 's'

    txt_promedios = f"Promedios de meses anteriores - {str_mes_inicial_promedios} a {str_mes_final_promedios}: " + \
                    f"{promPagosMesAnt} cuota{sAnt} retrasada{sAnt}, " + \
                    f"{promPagosMesAct} cuota{sAct} para el mes, " + \
                    f"{promPagosMesSig} anticipo{sSig}; " + \
                    f"total: {promPagosTotal} cuota{sTotal} completa{sTotal}"

    if GRAFICA_5_BARRA_COMPARACION:
        max_nPagosTotal = max(nPagosTotal)
        nPagosBarraComp = [0.00 for _ in range(len(nPagosTotal))]
        for i in range(len(nPagosBarraComp) - 1, -1, -12):
            nPagosBarraComp[i] = max_nPagosTotal

    meses = [fecha_inicial + relativedelta(months=mes) for mes in range(g5_nro_meses + 1)]
    # meses_ed =  [mes.strftime(FORMATO_MES) + '<span style="font-size: 10px">' + \
    #            f'<br>({nPagosTotal[idx]} cuotas)</span>' for idx, mes in enumerate(meses)]
    meses_ed =  [mes.strftime(FORMATO_MES) for mes in meses]

    if GRAFICA_5_PROM_MESES_ANTERIORES:
        avg = int(sum(nPagosTotal[:-1]) / len(nPagosTotal[:-1]))
        nPromMesesAnt = [avg for nP in nPagosTotal[:-1]]

    título =     'Cuotas recibidas por oportunidad de pago'
    if seleccionado_fin_de_mes:
        subtítulo = f'corte al fin de cada mes: {str_mes_inicial} a {str_mes_final}'
    else:
        subtítulo = f'corte al día {fecha_de_referencia.strftime("%d")} de cada mes: {str_mes_inicial} a {str_mes_final}'

    fig = go.Figure()

    # Distribución de cuotas
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed,
                        y = nPagosMesAnt,
                        mode = 'lines+text',
                        name = 'Meses anteriores',
                        line = dict(
                                    width = 3,
                                    color = '#4475cd',      # cranberry
                                ),
                        text = [str(nP) for nP in nPagosMesAnt],
                        textposition = 'top center',
                    ))
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed,
                        y = nPagosMesAct,
                        mode = 'lines+text',
                        name = 'Mes actual',
                        line = dict(
                                    width = 3,
                                    color = 'cornflowerblue',
                                ),
                        text = [str(nP) for nP in nPagosMesAct],
                        textposition = 'top center',
                    ))
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed,
                        y = nPagosMesSig,
                        mode = 'lines+text',
                        name = 'Meses subsiguientes',
                        line = dict(
                                    width = 3,
                                    color = '#84b5ff',
                                ),
                        text = [str(nP) for nP in nPagosMesSig],
                        textposition = 'top center',
                    ))
    fig.add_trace(
                    go.Scatter(
                        x = meses_ed,
                        y = nPagosTotal,
                        mode = 'lines+text',
                        name = 'Total pagos',
                        line = dict(
                                    width = 3,
                                    color = 'darkblue',
                                ),
                        text = [str(nP) for nP in nPagosTotal],
                        textposition = 'top center',
                    ))

    if GRAFICA_5_BARRA_COMPARACION:
        fig.add_trace(
                    go.Bar(
                        x = meses_ed,
                        y = nPagosBarraComp,
                        marker_color = 'rgb(183,226,240)',
                    )
            )
    if GRAFICA_5_PROM_MESES_ANTERIORES:
        fig.add_trace(
                        go.Scatter(
                            x = meses_ed[:-1],
                            y = nPromMesesAnt,
                            mode = 'lines+text',
                            name = 'Promedio',
                            line = dict(
                                        width = 1.5,
                                        color = 'red',
                                        dash = 'dash',
                                    ),
                            text = ['' for nP in nPromMesesAnt[:-1]] + [str(nPromMesesAnt[-1])],
                            textposition = 'top center',
                        ))

    # Define el desplazamiento vertical del nombre de la curva para evitar solapamientos
    valores = [{'key': 'Total cuotas', 'cuotas': nPagosTotal[-1]},
               {'key': 'Mes actual',   'cuotas': nPagosMesAct[-1]},
               {'key': 'Anticipos',    'cuotas': nPagosMesSig[-1]},
               {'key': 'Retrasados',   'cuotas': nPagosMesAnt[-1]}]
    df_valores = DataFrame(valores).sort_values(by='cuotas')
    df_valores = df_valores.set_index('key')
    v_anterior = df_valores.iloc[0]['cuotas']
    dy = 10
    diff = []
    sh_cuota = []   # True en ambas curvas si hay solapamiento
    offset = -dy
    for idx, r in df_valores.iterrows():
        v = r['cuotas']
        d = v - v_anterior
        offset = offset + dy if d <= 2 else 0
        diff.append(offset)
        sh_cuota.append(offset > 0)
        if offset > 0:
            sh_cuota[-2] = True
        v_anterior = v
    df_valores['offset'] = diff
    df_valores['muestra cuota'] = sh_cuota

    delta_x = 20
    delta_y = df_valores.loc['Retrasados']['offset']
    muestra_cuota = df_valores.loc['Retrasados']['muestra cuota']
    fig.add_annotation(
        x = 0.95,
        y = nPagosMesAnt[-1],
        name = 'Retrasados',
        text = 'Retrasados' + (f': {nPagosMesAnt[-1]}' if muestra_cuota else ''),
        yshift = delta_y,
    )
    delta_y = df_valores.loc['Mes actual']['offset']
    muestra_cuota = df_valores.loc['Mes actual']['muestra cuota']
    fig.add_annotation(
        x = 0.95,               # meses_ed[-1]
        y = nPagosMesAct[-1],
        name = 'Mes actual',
        text = 'Mes actual' + (f': {nPagosMesAct[-1]}' if delta_y > 0 else ''),
        yshift = delta_y,
    )
    delta_y = df_valores.loc['Anticipos']['offset']
    muestra_cuota = df_valores.loc['Anticipos']['muestra cuota']
    fig.add_annotation(
        x = 0.95,
        y = nPagosMesSig[-1],
        name = 'Anticipos',
        text = 'Anticipos' + (f': {nPagosMesSig[-1]}' if muestra_cuota else ''),
        yshift = delta_y,
    )
    delta_y = df_valores.loc['Total cuotas']['offset']
    muestra_cuota = df_valores.loc['Total cuotas']['muestra cuota']
    fig.add_annotation(
        x = 0.95,               # meses_ed[-1]
        y = nPagosTotal[-1],
        name = 'Total cuotas',
        text = 'Total cuotas' + (f': {nPagosTotal[-1]}' if muestra_cuota else ''),
        yshift = delta_y,
    )
    if GRAFICA_5_PROM_MESES_ANTERIORES:
        fig.add_annotation(
            x = 0.95,               # meses_ed[-1]
            y = nPromMesesAnt[-1],
            name = 'Promedio',
            text = 'Promedio',
            yshift = delta_y,
        )
    fig.update_annotations(dict(
        xref = "paper", yref = "y",
        xanchor = "left", yanchor = "bottom",
        xshift = delta_x,
        showarrow = False, arrowhead = 3, arrowsize = 1.5, standoff = 3,
    ))
    fig.update_layout(
        title = TITULO_GRAFICA.format(título, subtítulo),
        # xaxis_title = f'<span style="font-size: 12px"><i>{txt_promedios}</i></span>',
        width = 1000,
        height = 600,
        showlegend = False,
        # xaxis_tickangle = -30,
        plot_bgcolor = 'white',
    )
    fig.add_annotation(
        x = 0, xref = "paper", #xanchor = "left",
        y = -0.20 if g5_nro_meses > 25 else -0.17 if g5_nro_meses > 12 else -0.13, yref = "paper", #yanchor = "bottom",
        text = f'<span style="font-size: 12px"><i>{txt_promedios}</i></span>',
        showarrow = False
    )

    fig.update_xaxes(showline=True)
    fig.update_yaxes(showline=True)

    if g5_show:
        print(f"    . Desplegando imagen...")
        fig.show()

    if g5_save:
        print(f'    . Grabando "{grafica_png}"...')
        img_bytes = fig.to_image(format="png")
        with open(os.path.join(GyG_constantes.ruta_graficas, grafica_png), 'wb') as f:
            f.write(img_bytes)


#
# PROCESO
#

# print(f'Cargando hoja de cálculo "{excel_workbook}"...')
# ws = read_excel(excel_workbook, sheet_name=excel_ws_cobranzas, header=None)

HOY = fecha_de_referencia = datetime.today()
MES_ACTUAL = mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
MES_ANTERIOR = mes_anterior = mes_actual - relativedelta(days=1)
AYER = ayer = fecha_de_referencia - relativedelta(days=1)

cuotas_obj = Cuota(excel_workbook)

if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Lee los pagos recibidos')
df_vigilancia = read_excel(excel_workbook, sheet_name=excel_ws_vigilancia)

# Selecciona los pagos de vigilancia entre la fecha de referencia y el 1ro. de enero de 2017
df_vigilancia = df_vigilancia[(df_vigilancia['Categoría'] == 'Vigilancia')  & \
                              (df_vigilancia['Fecha'] <= fecha_de_referencia) & \
                              (df_vigilancia['Fecha'] >= datetime(2017, 1, 1))]
df_vigilancia = df_vigilancia[['Beneficiario', 'Fecha', 'Monto', 'Concepto', 'Día', 'Mes']]

if VERBOSE: print(f'    . [{datetime.now().strftime("%H:%M:%S")}] Agrega la cuota vigente para la fecha de pago')
df_vigilancia['_cuota_'] = df_vigilancia.apply(
                                lambda r: cuotas_obj.cuota_vigente(r['Beneficiario'], r['Fecha']), axis=1)
df_vigilancia['_num. cuotas_'] = df_vigilancia.apply(
                                lambda r: float(r['Monto']) / float(r['_cuota_']), axis=1)

lista_ultimos_meses = [(mes_anterior - relativedelta(months=nMeses)).strftime(FORMATO_MES)
                            for nMeses in range(GRAFICA_5_NRO_ULTIMOS_MESES)]

if toma_opciones_por_defecto:

    in_GUI_mode = False

    fecha_de_referencia = mes_anterior
    mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
    seleccionado_fin_de_mes = True

    print()

    # Selecciona si se muestran las gráficas generadas
    se_muestran_las_gráficas = input_si_no('Se muestran las gráficas generadas', 'no', toma_opciones_por_defecto)

    # Selecciona si se graban las imágenes generadas
    se_graban_las_imágenes = input_si_no('Se graban las imágenes generadas', 'sí', toma_opciones_por_defecto)

    # Selecciona la forma de desplegar las gráficas de barras
    graficas_de_barra_en_horizontal = input_si_no('Se muestran las barras en horizontal', 'no', toma_opciones_por_defecto)

    # Selecciona si se muestran los intervalos de confianza de los meses anteriores
    intervalos_de_confianza = input_si_no('Se muestran los intervalos de confianza', 'no', toma_opciones_por_defecto)

    print()

    g1 = True
    g1_show = se_muestran_las_gráficas
    g1_save = se_graban_las_imágenes
    g1_nro_meses = num_meses_gestion_cobranzas
    g1_conf_interval = intervalos_de_confianza

    g2 = False
    g2_show = se_muestran_las_gráficas
    g2_save = se_graban_las_imágenes
    g2_nro_meses = num_meses_cuotas_equivalentes
    g2_horizontal = graficas_de_barra_en_horizontal

    g3 = True
    g3_show = se_muestran_las_gráficas
    g3_save = se_graban_las_imágenes
    g3_nro_meses = num_meses_cuotas_equivalentes
    g3_horizontal = True    # graficas_de_barra_en_horizontal

    g4 = False
    g4_show = se_muestran_las_gráficas
    g4_save = se_graban_las_imágenes
    g4_nro_meses = num_meses_gestion_cobranzas
    g4_conf_interval = intervalos_de_confianza

    g5 = True
    g5_show = se_muestran_las_gráficas
    g5_save = se_graban_las_imágenes
    g5_nro_meses = num_meses_gestion_cobranzas

    fecha_de_referencia = mes_anterior

    genera_gráficas()

else:

    in_GUI_mode = True

    print("Generando layout ...")
#    sg.theme('DarkAmber')   # Add a little color to your windows

    sg.SetOptions(
            icon=os.path.join(GyG_constantes.rec_imágenes, 'GyG_logo.png'),
        )

    fechas_layout = [   
                        [sg.Radio("", size=(2, 1), default=True,
                                  group_id='fechas', key='_fecha_manual_'),
                         sg.Text("Gráficas al:"),
                         sg.InputText(key='_fecha_referencia_', size=(10,1),
                                      default_text=fecha_de_referencia.strftime("%d-%m-%Y")),
                         sg.CalendarButton(button_text='Seleccione otra fecha',
                                           image_filename=CALENDAR_ICON, image_size=(16, 16), image_subsample=8,
                                           default_date_m_d_y=(fecha_de_referencia.month, fecha_de_referencia.day, fecha_de_referencia.year),
                                           target='_fecha_referencia_', format='%d-%m-%Y',
                                           close_when_date_chosen=True)],
                        [sg.Radio("", size=(2, 1), 
                                  group_id='fechas', key='_fecha_mes_anterior_'),
# ------------------------------------------
                         # sg.Text(f"Gráficas del mes anterior:  {mes_anterior.strftime(FORMATO_MES)}", size=(30, 1))],
                         sg.Text("Gráficas del mes de:"),
                         sg.Combo(lista_ultimos_meses, key="_mes_a_graficar_", size=(10, 1),
                                  default_value=lista_ultimos_meses[0])],
# ------------------------------------------
                        [sg.Radio("", size=(2, 1),
                                  group_id='fechas', key='_fecha_ayer_'),
                         sg.Text(f"Gráficas al día de ayer:  {ayer.strftime('%A %d de %B %Y')}", size=(40, 1 )),
                         sg.Text("", size=(34, 1))],    # espaciador
                    ]
    ANCHURA_NOMBRE_GRAFICO = 28
    graficas_layout = [
                        [sg.Text("", size=(ANCHURA_NOMBRE_GRAFICO - 3, 1)), sg.Text("Visualiza"),
                         sg.Text("Graba"), sg.Text("Horiz."), sg.Text("x ± σ")],
                        [sg.Checkbox("Gestión de Cobranzas", key="_g1_", size=(ANCHURA_NOMBRE_GRAFICO, 1), default=False),
                         sg.Checkbox("", key="_g1_show_", size=(7, 1), default=True),
                         sg.Checkbox("", key="_g1_save_", size=(5, 1)),
                         sg.Text("", size=(5, 1)),      # espaciador (gráfica siempre vertical)
                         sg.Checkbox("", key="_g1_conf_interval_", size=(5, 1)),
                         sg.Text("Nº meses:"), sg.Spin(key="_g1_nro_meses_", values=[str(i) for i in range(2, 25)],
                                                       initial_value=str(num_meses_gestion_cobranzas), size=(3, 1)),
                         sg.Text("", size=(1, 1))],
                        [sg.Checkbox("Pagos 100% equivalentes", key="_g2_", size=(ANCHURA_NOMBRE_GRAFICO, 1), default=False),
                         sg.Checkbox("", key="_g2_show_", size=(7, 1), default=True),
                         sg.Checkbox("", key="_g2_save_", size=(5, 1)),
                         sg.Checkbox("", key="_g2_horizontal_", size=(5, 1)),
                         sg.Text("", size=(5, 1)),      # espaciador (gráficas de barras sin intervalos de confianza)
                         sg.Text("Nº meses:"), sg.Spin(key="_g2_nro_meses_", values=[str(i) for i in range(3, 25)],
                                                       initial_value=str(num_meses_cuotas_equivalentes), size=(3, 1))],
                        [sg.Checkbox("Cuotas equivalentes", key="_g3_", size=(ANCHURA_NOMBRE_GRAFICO, 1), default=False),
                         sg.Checkbox("", key="_g3_show_", size=(7, 1), default=True),
                         sg.Checkbox("", key="_g3_save_", size=(5, 1)),
                         sg.Checkbox("", key="_g3_horizontal_", size=(5, 1), default=True),
                         sg.Text("", size=(5, 1)),      # espaciador (gráficas de barras sin intervalos de confianza)
                         sg.Text("Nº meses:"), sg.Spin(key="_g3_nro_meses_", values=[str(i) for i in range(3, 25)],
                                                       initial_value=str(num_meses_cuotas_equivalentes), size=(3, 1))],
                        [sg.Checkbox("Pagos Recibidos", key="_g4_", size=(ANCHURA_NOMBRE_GRAFICO, 1), default=False),
                         sg.Checkbox("", key="_g4_show_", size=(7, 1), default=True),
                         sg.Checkbox("", key="_g4_save_", size=(5, 1)),
                         sg.Text("", size=(5, 1)),      # espaciador (gráfica siempre vertical)
                         sg.Checkbox("", key='_g4_conf_interval_', size=(5, 1)),
                         sg.Text("Nº meses:"), sg.Spin(key="_g4_nro_meses_", values=[str(i) for i in range(2, 25)],
                                                       initial_value=str(num_meses_gestion_cobranzas), size=(3, 1)),
                         sg.Text("", size=(1, 1))],
                        [sg.Checkbox("Cuotas por oportunidad de pago", key="_g5_", size=(ANCHURA_NOMBRE_GRAFICO, 1),
                                                                       default=False),
                         sg.Checkbox("", key="_g5_show_", size=(7, 1), default=True),
                         sg.Checkbox("", key="_g5_save_", size=(5, 1)),
                         sg.Text("", size=(5, 1)),      # espaciador (gráfica siempre vertical)
                         sg.Text("", size=(5, 1)),      # espaciador (valores únicos por mes)
                         sg.Text("Nº meses:"), sg.Spin(key="_g5_nro_meses_", values=[str(i) for i in range(2, 25)],
                                                       initial_value=str(num_meses_gestion_cobranzas), size=(3, 1)),
                         sg.Text("", size=(1, 1))],
                    ]
    layout = [  [sg.Frame(" Fecha ",    fechas_layout,   size=(60, None))],
                [sg.Frame(" Gráficas ", graficas_layout, size=(60, None))],
                [sg.Button('Genera gráficas'), sg.Button('Finaliza')],
             ]

    # Create the Window
    window = sg.Window('GyG Gráficas', layout)

    # Event Loop to process "events"
    while True:             
        event, values = window.read()

        seleccionado_fin_de_mes = False
        
        if event in (None, 'Finaliza'):
            break
        elif event == 'Genera gráficas':
            if values['_fecha_manual_']:
                fecha_de_referencia = datetime.strptime(values['_fecha_referencia_'], '%d-%m-%Y')
            elif values['_fecha_mes_anterior_']:
# ----------------------------------------------
                # fecha_de_referencia = mes_anterior
                fecha_de_referencia = MES_ANTERIOR - \
                                      relativedelta(months=lista_ultimos_meses.index(values['_mes_a_graficar_']))
                fecha_de_referencia += relativedelta(day=31)
                seleccionado_fin_de_mes = True
# ----------------------------------------------
            elif values['_fecha_ayer_']:
                fecha_de_referencia = ayer
            if fecha_de_referencia > datetime.today():
                fecha_de_referencia = datetime.today()

            mes_actual = datetime(fecha_de_referencia.year, fecha_de_referencia.month, 1)
            mes_anterior = mes_actual - relativedelta(days=1)
            seleccionado_fin_de_mes = es_fin_de_mes(fecha_de_referencia)

            g1 = values['_g1_']
            g1_show = values['_g1_show_']
            g1_save = values['_g1_save_']
            g1_nro_meses = int(values['_g1_nro_meses_'])
            g1_conf_interval = values['_g1_conf_interval_']

            g2 = values['_g2_']
            g2_show = values['_g2_show_']
            g2_save = values['_g2_save_']
            g2_nro_meses = int(values['_g2_nro_meses_'])
            g2_horizontal = values['_g2_horizontal_']

            g3 = values['_g3_']
            g3_show = values['_g3_show_']
            g3_save = values['_g3_save_']
            g3_nro_meses = int(values['_g3_nro_meses_'])
            g3_horizontal = values['_g3_horizontal_']

            g4 = values['_g4_']
            g4_show = values['_g4_show_']
            g4_save = values['_g4_save_']
            g4_nro_meses = int(values['_g4_nro_meses_'])
            g4_conf_interval = values['_g4_conf_interval_']

            g5 = values['_g5_']
            g5_show = values['_g5_show_']
            g5_save = values['_g5_save_']
            g5_nro_meses = int(values['_g5_nro_meses_'])

            genera_gráficas()

    window.close()

# print()
# print('Proceso terminado . . .')
